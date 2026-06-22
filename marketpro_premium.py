#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
МаркетПРО Premium — коммерческий Telegram AI-помощник селлера.

Главные возможности:
- Карточка товара 360° из текста или фото
- Сохранённые товары и Brand Kit
- Аудит карточки по тексту, ссылке и скриншотам
- Анализ конкурента и AI-стратегия ниши
- Юнит-экономика PRO, финансовый доктор, акции и поставки
- Ответы и пакетный анализ отзывов
- Апелляции PRO
- Подключение WB/Ozon Seller API (без имитации данных)
- AI-директор и ежедневные отчёты
- Тарифы, платежи, лимиты, рефералы, промокоды
- История, экспорт, админ-статистика и рассылки

Один файл специально сохранён для простого развёртывания. Для масштабирования
его можно без изменения бизнес-логики разнести на модули.
"""

from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import html
import io
import json
import logging
import math
import os
import re
import sqlite3
import uuid
from contextlib import asynccontextmanager
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urlparse

import aiosqlite
import httpx
from aiogram import Bot, Dispatcher, F, Router
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openai import AsyncOpenAI

try:
    from aiogram.fsm.storage.redis import RedisStorage
except Exception:  # optional dependency
    RedisStorage = None

try:
    from cryptography.fernet import Fernet
except Exception:  # optional dependency
    Fernet = None

try:
    from yookassa import Configuration, Payment
except Exception:  # optional dependency
    Configuration = Payment = None

try:
    import anthropic
except Exception:  # optional dependency
    anthropic = None

try:
    from openpyxl import load_workbook
except Exception:  # optional dependency
    load_workbook = None


# ============================================================================
# CONFIG
# ============================================================================

def load_env_file(path: str) -> dict[str, str]:
    data: dict[str, str] = {}
    p = Path(path)
    if not p.exists():
        return data
    for raw in p.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        data[key.strip()] = value.strip().strip('"').strip("'")
    return data


FILE_ENV = load_env_file(os.getenv("MARKETPRO_ENV", "/root/.env_card"))


def env(name: str, default: str = "") -> str:
    return os.getenv(name, FILE_ENV.get(name, default))


def env_int(name: str, default: int) -> int:
    try:
        return int(env(name, str(default)))
    except ValueError:
        return default


def env_float(name: str, default: float) -> float:
    try:
        return float(env(name, str(default)))
    except ValueError:
        return default


@dataclass(frozen=True)
class Config:
    bot_token: str = env("BOT_TOKEN")
    bot_username: str = env("BOT_USERNAME", "PostGeniusHelperBot").lstrip("@")
    owner_id: int = env_int("OWNER_ID", 549639607)
    support_url: str = env("SUPPORT_URL", "https://t.me/Boss023rus")
    db_path: str = env("DB_PATH", "/root/marketpro_premium.db")
    legacy_db_path: str = env("LEGACY_DB_PATH", "/root/marketpro.db")
    redis_url: str = env("REDIS_URL")

    openai_key: str = env("OPENAI_KEY")
    openai_text_model: str = env("OPENAI_TEXT_MODEL", "gpt-4o-mini")
    openai_vision_model: str = env("OPENAI_VISION_MODEL", "gpt-4o")
    openai_image_model: str = env("OPENAI_IMAGE_MODEL", "gpt-image-1")
    anthropic_key: str = env("ANTHROPIC_KEY")
    anthropic_model: str = env("ANTHROPIC_MODEL", "claude-sonnet-4-5")

    yookassa_shop_id: str = env("YOOKASSA_SHOP_ID")
    yookassa_secret: str = env("YOOKASSA_SECRET")
    payment_return_url: str = env("PAYMENT_RETURN_URL", "https://t.me/PostGeniusHelperBot")

    encryption_key: str = env("ENCRYPTION_KEY")
    timezone_offset: int = env_int("TZ_OFFSET", 3)
    request_timeout: int = env_int("HTTP_TIMEOUT", 35)
    max_input_chars: int = env_int("MAX_INPUT_CHARS", 18000)
    max_photo_mb: int = env_int("MAX_PHOTO_MB", 15)
    fair_use_daily: int = env_int("FAIR_USE_DAILY", 250)
    image_daily: int = env_int("IMAGE_DAILY", 10)
    free_credits: int = env_int("FREE_CREDITS", 12)
    referral_inviter: int = env_int("REF_BONUS_INVITER", 7)
    referral_invited: int = env_int("REF_BONUS_INVITED", 5)


CFG = Config()
LOCAL_TZ = timezone(timedelta(hours=CFG.timezone_offset))

PLANS: dict[str, dict[str, Any]] = {
    "free": {"name": "Знакомство", "price": 0, "days": 0, "credits": CFG.free_credits, "products": 2, "images": 1, "shops": 0},
    "start": {"name": "Старт", "price": 490, "days": 30, "credits": 80, "products": 7, "images": 5, "shops": 0},
    "pro": {"name": "ПРО", "price": 990, "days": 30, "credits": 350, "products": 40, "images": 20, "shops": 1},
    "business": {"name": "Бизнес", "price": 1990, "days": 30, "credits": 1200, "products": 250, "images": 80, "shops": 4},
}

FEATURE_COST = {
    "card360_text": 3,
    "card360_photo": 5,
    "audit": 2,
    "audit_link": 3,
    "competitor": 3,
    "niche": 3,
    "review": 1,
    "review_batch": 3,
    "appeal": 2,
    "finance": 1,
    "doctor": 2,
    "supply": 1,
    "image": 8,
    "director": 3,
}

logging.basicConfig(
    level=getattr(logging, env("LOG_LEVEL", "INFO").upper(), logging.INFO),
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
log = logging.getLogger("marketpro")

if Payment and CFG.yookassa_shop_id and CFG.yookassa_secret:
    Configuration.account_id = CFG.yookassa_shop_id
    Configuration.secret_key = CFG.yookassa_secret


# ============================================================================
# HELPERS
# ============================================================================

def now() -> datetime:
    return datetime.now(LOCAL_TZ)


def iso_now() -> str:
    return now().isoformat()


def parse_dt(value: str | None) -> Optional[datetime]:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        return dt if dt.tzinfo else dt.replace(tzinfo=LOCAL_TZ)
    except ValueError:
        return None


def safe_float(value: str, minimum: float = 0, maximum: float = 10**9) -> float:
    cleaned = re.sub(r"[^0-9,.-]", "", value).replace(",", ".")
    number = float(cleaned)
    if not minimum <= number <= maximum:
        raise ValueError(f"Значение должно быть от {minimum:g} до {maximum:g}")
    return number


def safe_int(value: str, minimum: int = 0, maximum: int = 10**9) -> int:
    return int(safe_float(value, minimum, maximum))


def money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ") + " ₽"


def pct(value: float) -> str:
    return f"{value:.1f}%"


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def chunks(text: str, size: int = 3900) -> list[str]:
    if len(text) <= size:
        return [text]
    parts: list[str] = []
    current = ""
    for paragraph in text.split("\n"):
        candidate = current + ("\n" if current else "") + paragraph
        if len(candidate) <= size:
            current = candidate
            continue
        if current:
            parts.append(current)
        while len(paragraph) > size:
            parts.append(paragraph[:size])
            paragraph = paragraph[size:]
        current = paragraph
    if current:
        parts.append(current)
    return parts


def clean_ai_text(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^```(?:json|markdown|text)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


def extract_json(text: str) -> dict[str, Any]:
    clean = clean_ai_text(text)
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", clean, re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def redact_token(token: str) -> str:
    if len(token) <= 8:
        return "••••••••"
    return token[:4] + "••••" + token[-4:]


def normalize_url(url: str) -> str:
    url = url.strip()
    if not re.match(r"^https?://", url, re.I):
        url = "https://" + url
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("Некорректная ссылка")
    return url


def marketplace_from_url(url: str) -> str:
    host = urlparse(url).netloc.lower()
    if "wildberries" in host or "wb.ru" in host:
        return "wb"
    if "ozon" in host:
        return "ozon"
    if "avito" in host:
        return "avito"
    return "unknown"


def extract_article(url: str, marketplace: str) -> Optional[str]:
    patterns = {
        "wb": [r"/catalog/(\d+)", r"(?:nm|article|id)[=/](\d+)"],
        "ozon": [r"-(\d{6,})(?:/|\?|$)", r"/product/[^/?]*?(\d{6,})"],
    }
    for pattern in patterns.get(marketplace, []):
        m = re.search(pattern, url, re.I)
        if m:
            return m.group(1)
    return None


class TokenCipher:
    def __init__(self, key: str):
        self._fernet = None
        if key and Fernet:
            try:
                raw = key.encode()
                if len(raw) != 44:
                    raw = base64.urlsafe_b64encode(hashlib.sha256(raw).digest())
                self._fernet = Fernet(raw)
            except Exception as exc:
                log.error("Encryption init failed: %s", exc)

    @property
    def secure(self) -> bool:
        return self._fernet is not None

    def encrypt(self, value: str) -> str:
        if not value:
            return ""
        if not self._fernet:
            raise RuntimeError("ENCRYPTION_KEY не настроен: подключение кабинета нельзя сохранить безопасно")
        return self._fernet.encrypt(value.encode()).decode()

    def decrypt(self, value: str) -> str:
        if not value:
            return ""
        try:
            if self._fernet and not value.startswith("plain:"):
                return self._fernet.decrypt(value.encode()).decode()
            if value.startswith("plain:"):
                return base64.urlsafe_b64decode(value[6:].encode()).decode()
        except Exception:
            return ""
        return value


CIPHER = TokenCipher(CFG.encryption_key)


# ============================================================================
# DATABASE
# ============================================================================

SCHEMA = """
PRAGMA journal_mode=WAL;
PRAGMA foreign_keys=ON;

CREATE TABLE IF NOT EXISTS users (
    user_id INTEGER PRIMARY KEY,
    username TEXT NOT NULL DEFAULT '',
    first_name TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL,
    last_seen_at TEXT NOT NULL,
    onboarded INTEGER NOT NULL DEFAULT 0,
    experience TEXT NOT NULL DEFAULT 'newbie',
    primary_marketplace TEXT NOT NULL DEFAULT 'wb',
    is_blocked INTEGER NOT NULL DEFAULT 0,
    ref_by INTEGER NOT NULL DEFAULT 0,
    referral_rewarded INTEGER NOT NULL DEFAULT 0,
    daily_report_hour INTEGER NOT NULL DEFAULT 9,
    notifications_enabled INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS balances (
    user_id INTEGER PRIMARY KEY,
    credits INTEGER NOT NULL DEFAULT 0,
    image_credits INTEGER NOT NULL DEFAULT 0,
    lifetime_used INTEGER NOT NULL DEFAULT 0,
    daily_used INTEGER NOT NULL DEFAULT 0,
    daily_date TEXT NOT NULL DEFAULT '',
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS subscriptions (
    user_id INTEGER PRIMARY KEY,
    plan TEXT NOT NULL DEFAULT 'free',
    started_at TEXT NOT NULL DEFAULT '',
    expires_at TEXT NOT NULL DEFAULT '',
    auto_renew INTEGER NOT NULL DEFAULT 0,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    marketplace TEXT NOT NULL DEFAULT 'wb',
    article TEXT NOT NULL DEFAULT '',
    brand TEXT NOT NULL DEFAULT '',
    category TEXT NOT NULL DEFAULT '',
    audience TEXT NOT NULL DEFAULT '',
    benefits TEXT NOT NULL DEFAULT '',
    characteristics TEXT NOT NULL DEFAULT '',
    keywords TEXT NOT NULL DEFAULT '',
    source_text TEXT NOT NULL DEFAULT '',
    current_content TEXT NOT NULL DEFAULT '',
    cost REAL NOT NULL DEFAULT 0,
    price REAL NOT NULL DEFAULT 0,
    length_cm REAL NOT NULL DEFAULT 0,
    width_cm REAL NOT NULL DEFAULT 0,
    height_cm REAL NOT NULL DEFAULT 0,
    weight_kg REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'active',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS brand_kits (
    user_id INTEGER PRIMARY KEY,
    brand_name TEXT NOT NULL DEFAULT '',
    tone TEXT NOT NULL DEFAULT 'профессиональный и дружелюбный',
    colors TEXT NOT NULL DEFAULT '',
    fonts TEXT NOT NULL DEFAULT '',
    visual_style TEXT NOT NULL DEFAULT '',
    target_audience TEXT NOT NULL DEFAULT '',
    forbidden_phrases TEXT NOT NULL DEFAULT '',
    logo_file_id TEXT NOT NULL DEFAULT '',
    updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    product_id INTEGER,
    feature TEXT NOT NULL,
    input_text TEXT NOT NULL DEFAULT '',
    result TEXT NOT NULL,
    metadata TEXT NOT NULL DEFAULT '{}',
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE,
    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS marketplace_connections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    marketplace TEXT NOT NULL,
    name TEXT NOT NULL,
    client_id TEXT NOT NULL DEFAULT '',
    token_encrypted TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    last_sync_at TEXT NOT NULL DEFAULT '',
    last_error TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL,
    UNIQUE(user_id, marketplace, name),
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS imported_metrics (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    product_id INTEGER,
    source TEXT NOT NULL,
    period_start TEXT NOT NULL DEFAULT '',
    period_end TEXT NOT NULL DEFAULT '',
    metrics_json TEXT NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE,
    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS payments (
    payment_id TEXT PRIMARY KEY,
    user_id INTEGER NOT NULL,
    plan TEXT NOT NULL,
    amount REAL NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS referral_events (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invited_user_id INTEGER NOT NULL UNIQUE,
    inviter_user_id INTEGER NOT NULL,
    registered_rewarded INTEGER NOT NULL DEFAULT 0,
    paid_rewarded INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS promo_codes (
    code TEXT PRIMARY KEY,
    bonus_credits INTEGER NOT NULL DEFAULT 0,
    discount_percent INTEGER NOT NULL DEFAULT 0,
    max_uses INTEGER NOT NULL DEFAULT 0,
    uses INTEGER NOT NULL DEFAULT 0,
    expires_at TEXT NOT NULL DEFAULT '',
    active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS promo_uses (
    user_id INTEGER NOT NULL,
    code TEXT NOT NULL,
    used_at TEXT NOT NULL,
    PRIMARY KEY(user_id, code)
);

CREATE TABLE IF NOT EXISTS recommendations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    product_id INTEGER,
    severity TEXT NOT NULL,
    title TEXT NOT NULL,
    description TEXT NOT NULL,
    estimated_effect REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'new',
    created_at TEXT NOT NULL,
    resolved_at TEXT NOT NULL DEFAULT '',
    category TEXT NOT NULL DEFAULT 'general',
    confidence INTEGER NOT NULL DEFAULT 50,
    action_code TEXT NOT NULL DEFAULT '',
    evidence_json TEXT NOT NULL DEFAULT '{}',
    fingerprint TEXT NOT NULL DEFAULT '',
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS metric_snapshots (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    product_id INTEGER,
    marketplace TEXT NOT NULL DEFAULT '',
    article TEXT NOT NULL DEFAULT '',
    metric_date TEXT NOT NULL,
    orders REAL NOT NULL DEFAULT 0,
    revenue REAL NOT NULL DEFAULT 0,
    views REAL NOT NULL DEFAULT 0,
    clicks REAL NOT NULL DEFAULT 0,
    cart_adds REAL NOT NULL DEFAULT 0,
    buyouts REAL NOT NULL DEFAULT 0,
    returns REAL NOT NULL DEFAULT 0,
    stock REAL NOT NULL DEFAULT 0,
    ad_spend REAL NOT NULL DEFAULT 0,
    price REAL NOT NULL DEFAULT 0,
    rating REAL NOT NULL DEFAULT 0,
    reviews_count REAL NOT NULL DEFAULT 0,
    source TEXT NOT NULL DEFAULT '',
    raw_json TEXT NOT NULL DEFAULT '{}',
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE,
    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS decision_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    recommendation_id INTEGER,
    action TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'planned',
    note TEXT NOT NULL DEFAULT '',
    baseline_json TEXT NOT NULL DEFAULT '{}',
    result_json TEXT NOT NULL DEFAULT '{}',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(user_id) ON DELETE CASCADE,
    FOREIGN KEY(recommendation_id) REFERENCES recommendations(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS migrations (
    name TEXT PRIMARY KEY,
    applied_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS sync_runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    connection_id INTEGER,
    status TEXT NOT NULL,
    items_count INTEGER NOT NULL DEFAULT 0,
    details TEXT NOT NULL DEFAULT '',
    started_at TEXT NOT NULL,
    finished_at TEXT NOT NULL DEFAULT ''
);

CREATE INDEX IF NOT EXISTS idx_history_user ON history(user_id, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_products_user ON products(user_id, status);
CREATE INDEX IF NOT EXISTS idx_payments_status ON payments(status);
CREATE INDEX IF NOT EXISTS idx_metrics_user_date ON metric_snapshots(user_id, metric_date DESC);
CREATE INDEX IF NOT EXISTS idx_metrics_article ON metric_snapshots(user_id, marketplace, article, metric_date DESC);
CREATE INDEX IF NOT EXISTS idx_recommendations_user ON recommendations(user_id, status, severity);
"""


class Database:
    def __init__(self, path: str):
        self.path = path
        self._write_lock = asyncio.Lock()

    async def init(self) -> None:
        Path(self.path).parent.mkdir(parents=True, exist_ok=True)
        async with aiosqlite.connect(self.path) as db:
            await db.executescript(SCHEMA)
            await self._ensure_schema_upgrades(db)
            await db.commit()

    async def _ensure_schema_upgrades(self, db: aiosqlite.Connection) -> None:
        """Идемпотентно добавляет новые поля при обновлении существующей premium-базы."""
        upgrades = {
            "recommendations": {
                "category": "TEXT NOT NULL DEFAULT 'general'",
                "confidence": "INTEGER NOT NULL DEFAULT 50",
                "action_code": "TEXT NOT NULL DEFAULT ''",
                "evidence_json": "TEXT NOT NULL DEFAULT '{}'",
                "fingerprint": "TEXT NOT NULL DEFAULT ''",
            },
        }
        for table, columns in upgrades.items():
            cur = await db.execute(f"PRAGMA table_info({table})")
            existing = {row[1] for row in await cur.fetchall()}
            for name, definition in columns.items():
                if name not in existing:
                    await db.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")

    async def migrate_legacy(self, legacy_path: str) -> dict[str, int]:
        """Однократно переносит данные старой версии, не изменяя исходную БД."""
        result = {"users": 0, "history": 0, "payments": 0}
        legacy = Path(legacy_path)
        if not legacy.exists() or legacy.resolve() == Path(self.path).resolve():
            return result
        marker = await self.fetchone("SELECT name FROM migrations WHERE name='legacy_marketpro_v1'")
        if marker:
            return result

        def read_legacy() -> dict[str, list[tuple[Any, ...]]]:
            out: dict[str, list[tuple[Any, ...]]] = {}
            conn = sqlite3.connect(str(legacy))
            try:
                tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
                for table in ("users", "requests", "subscriptions", "history", "pending_payments", "referrals"):
                    if table in tables:
                        out[table] = conn.execute(f"SELECT * FROM {table}").fetchall()
            finally:
                conn.close()
            return out

        data = await asyncio.to_thread(read_legacy)
        old_users = data.get("users", [])
        req_map = {int(r[0]): int(r[1] or 0) for r in data.get("requests", []) if len(r) >= 2}
        ref_map = {int(r[0]): r for r in data.get("referrals", []) if r}
        for row in old_users:
            uid = int(row[0])
            username = str(row[1] or "") if len(row) > 1 else ""
            first_name = str(row[2] or "") if len(row) > 2 else ""
            created = str(row[3] or iso_now()) if len(row) > 3 else iso_now()
            onboarded = int(row[4] or 0) if len(row) > 4 else 0
            await ensure_user(uid, username, first_name)
            await self.execute("UPDATE users SET created_at=?,onboarded=? WHERE user_id=?", (created, onboarded, uid))
            used = req_map.get(uid, 0)
            bonus = 0
            ref = ref_map.get(uid)
            if ref and len(ref) > 3:
                bonus = int(ref[3] or 0)
            remaining = max(0, CFG.free_credits + bonus - used)
            await self.execute("UPDATE balances SET credits=MAX(credits,?),lifetime_used=MAX(lifetime_used,?) WHERE user_id=?", (remaining, used, uid))
            result["users"] += 1

        for row in data.get("subscriptions", []):
            if len(row) < 4:
                continue
            uid, plan, sub_end, paid_at = int(row[0]), str(row[1] or ""), str(row[2] or ""), str(row[3] or "")
            mapped = "pro" if plan and plan != "free" else "free"
            await ensure_user(uid)
            await self.execute("INSERT INTO subscriptions(user_id,plan,started_at,expires_at) VALUES(?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET plan=excluded.plan,started_at=excluded.started_at,expires_at=excluded.expires_at", (uid, mapped, paid_at, sub_end))
            legacy_expiry = parse_dt(sub_end)
            if mapped == "pro" and legacy_expiry and legacy_expiry > now():
                await self.execute(
                    "UPDATE balances SET credits=MAX(credits,?),image_credits=MAX(image_credits,?) WHERE user_id=?",
                    (PLANS["pro"]["credits"], PLANS["pro"]["images"], uid),
                )

        for row in data.get("history", []):
            if len(row) < 5:
                continue
            old_id, uid, feature, text_result, created = row[:5]
            exists = await self.fetchone("SELECT id FROM history WHERE user_id=? AND feature=? AND created_at=?", (int(uid), str(feature), str(created)))
            if not exists:
                await ensure_user(int(uid))
                await self.execute("INSERT INTO history(user_id,feature,result,created_at) VALUES(?,?,?,?)", (int(uid), str(feature), str(text_result or "")[:50000], str(created or iso_now())))
                result["history"] += 1

        for row in data.get("pending_payments", []):
            if len(row) < 5:
                continue
            pid, uid, plan, amount, created = row[:5]
            await ensure_user(int(uid))
            await self.execute("INSERT OR IGNORE INTO payments(payment_id,user_id,plan,amount,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?)", (str(pid), int(uid), "pro" if str(plan) else "pro", float(amount or 0), "pending", str(created or iso_now()), iso_now()))
            result["payments"] += 1

        await self.execute("INSERT INTO migrations(name,applied_at) VALUES('legacy_marketpro_v1',?)", (iso_now(),))
        return result

    @asynccontextmanager
    async def connect(self):
        db = await aiosqlite.connect(self.path)
        db.row_factory = aiosqlite.Row
        try:
            yield db
        finally:
            await db.close()

    async def execute(self, sql: str, params: Iterable[Any] = ()) -> int:
        async with self._write_lock, self.connect() as db:
            cur = await db.execute(sql, tuple(params))
            await db.commit()
            return cur.lastrowid

    async def execute_rowcount(self, sql: str, params: Iterable[Any] = ()) -> int:
        async with self._write_lock, self.connect() as db:
            cur = await db.execute(sql, tuple(params))
            await db.commit()
            return cur.rowcount

    async def fetchone(self, sql: str, params: Iterable[Any] = ()) -> Optional[dict[str, Any]]:
        async with self.connect() as db:
            cur = await db.execute(sql, tuple(params))
            row = await cur.fetchone()
            return dict(row) if row else None

    async def fetchall(self, sql: str, params: Iterable[Any] = ()) -> list[dict[str, Any]]:
        async with self.connect() as db:
            cur = await db.execute(sql, tuple(params))
            rows = await cur.fetchall()
            return [dict(row) for row in rows]

    async def transaction(self, operations: list[tuple[str, tuple[Any, ...]]]) -> None:
        async with self._write_lock, self.connect() as db:
            try:
                await db.execute("BEGIN IMMEDIATE")
                for sql, params in operations:
                    await db.execute(sql, params)
                await db.commit()
            except Exception:
                await db.rollback()
                raise


DB = Database(CFG.db_path)


async def ensure_user(user_id: int, username: str = "", first_name: str = "") -> None:
    stamp = iso_now()
    await DB.transaction([
        ("""INSERT INTO users(user_id,username,first_name,created_at,last_seen_at)
             VALUES(?,?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET
             username=excluded.username, first_name=excluded.first_name,
             last_seen_at=excluded.last_seen_at""", (user_id, username, first_name, stamp, stamp)),
        ("INSERT OR IGNORE INTO balances(user_id,credits,image_credits,daily_date) VALUES(?,?,?,?)",
         (user_id, CFG.free_credits, 1, now().date().isoformat())),
        ("INSERT OR IGNORE INTO subscriptions(user_id,plan) VALUES(?, 'free')", (user_id,)),
    ])


async def user_plan(user_id: int) -> tuple[str, Optional[datetime]]:
    row = await DB.fetchone("SELECT plan,expires_at FROM subscriptions WHERE user_id=?", (user_id,))
    if not row:
        return "free", None
    plan = row["plan"] or "free"
    expires = parse_dt(row["expires_at"])
    if plan != "free" and (not expires or expires <= now()):
        await DB.execute("UPDATE subscriptions SET plan='free',expires_at='' WHERE user_id=?", (user_id,))
        return "free", None
    return plan, expires


async def activate_plan(user_id: int, plan: str) -> None:
    if plan not in PLANS or plan == "free":
        raise ValueError("Неизвестный тариф")
    old_plan, old_end = await user_plan(user_id)
    start = now()
    base = old_end if old_end and old_end > start else start
    end = base + timedelta(days=PLANS[plan]["days"])
    await DB.transaction([
        ("""INSERT INTO subscriptions(user_id,plan,started_at,expires_at)
             VALUES(?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET
             plan=excluded.plan, started_at=excluded.started_at, expires_at=excluded.expires_at""",
         (user_id, plan, start.isoformat(), end.isoformat())),
        ("UPDATE balances SET credits=credits+?, image_credits=image_credits+? WHERE user_id=?",
         (PLANS[plan]["credits"], PLANS[plan]["images"], user_id)),
    ])


async def balance(user_id: int) -> dict[str, Any]:
    today = now().date().isoformat()
    row = await DB.fetchone("SELECT * FROM balances WHERE user_id=?", (user_id,))
    if not row:
        await ensure_user(user_id)
        row = await DB.fetchone("SELECT * FROM balances WHERE user_id=?", (user_id,))
    if row and row["daily_date"] != today:
        await DB.execute("UPDATE balances SET daily_used=0,daily_date=? WHERE user_id=?", (today, user_id))
        row["daily_used"] = 0
        row["daily_date"] = today
    return row or {}


async def charge(user_id: int, feature: str, image: bool = False) -> tuple[bool, str]:
    """Атомарно списывает кредиты и защищает баланс от параллельных запросов."""
    cost = FEATURE_COST.get(feature, 1)
    await balance(user_id)  # создаёт баланс и сбрасывает дневной счётчик при смене даты
    user = await DB.fetchone("SELECT is_blocked FROM users WHERE user_id=?", (user_id,))
    if user and user["is_blocked"]:
        return False, "Ваш доступ временно ограничен. Напишите в поддержку."
    if image:
        changed = await DB.execute_rowcount(
            """UPDATE balances SET image_credits=image_credits-1,daily_used=daily_used+1,lifetime_used=lifetime_used+1
               WHERE user_id=? AND image_credits>=1 AND daily_used<?""",
            (user_id, CFG.fair_use_daily),
        )
    else:
        changed = await DB.execute_rowcount(
            """UPDATE balances SET credits=credits-?,daily_used=daily_used+1,lifetime_used=lifetime_used+1
               WHERE user_id=? AND credits>=? AND daily_used<?""",
            (cost, user_id, cost, CFG.fair_use_daily),
        )
    if changed:
        return True, ""
    bal = await balance(user_id)
    if bal.get("daily_used", 0) >= CFG.fair_use_daily:
        return False, "Достигнут дневной fair-use лимит. Он защищает сервис от автоматической перегрузки."
    if image:
        return False, "Закончились кредиты изображений. Пополните тариф или пакет изображений."
    return False, f"Для операции нужно {cost} кредитов, доступно {bal.get('credits', 0)}."


async def refund(user_id: int, feature: str, image: bool = False) -> None:
    """Полностью откатывает списание при технической ошибке или отмене сценария."""
    if image:
        await DB.execute(
            """UPDATE balances SET image_credits=image_credits+1,
               daily_used=MAX(0,daily_used-1),lifetime_used=MAX(0,lifetime_used-1) WHERE user_id=?""",
            (user_id,),
        )
    else:
        await DB.execute(
            """UPDATE balances SET credits=credits+?,daily_used=MAX(0,daily_used-1),
               lifetime_used=MAX(0,lifetime_used-1) WHERE user_id=?""",
            (FEATURE_COST.get(feature, 1), user_id),
        )


async def save_history(user_id: int, feature: str, input_text: str, result: str,
                       product_id: int | None = None, metadata: dict[str, Any] | None = None) -> int:
    return await DB.execute(
        "INSERT INTO history(user_id,product_id,feature,input_text,result,metadata,created_at) VALUES(?,?,?,?,?,?,?)",
        (user_id, product_id, feature, input_text[:12000], result[:50000], json.dumps(metadata or {}, ensure_ascii=False), iso_now()),
    )


# ============================================================================
# METRICS, RECOMMENDATIONS AND HEALTH SCORE
# ============================================================================

METRIC_ALIASES: dict[str, tuple[str, ...]] = {
    "article": ("артикул", "nm id", "nmid", "sku", "offer id", "offer_id", "артикул wb", "артикул продавца"),
    "name": ("товар", "название", "наименование", "product", "product name"),
    "date": ("дата", "день", "date", "metric date", "период"),
    "marketplace": ("площадка", "marketplace"),
    "orders": ("заказы", "заказов", "orders", "количество заказов"),
    "revenue": ("выручка", "оборот", "revenue", "sales amount", "сумма заказов"),
    "views": ("показы", "просмотры", "views", "impressions", "открытия карточки"),
    "clicks": ("клики", "переходы", "clicks", "open card"),
    "cart_adds": ("корзины", "добавления в корзину", "cart", "cart adds", "add to cart"),
    "buyouts": ("выкупы", "выкуплено", "buyouts", "purchases"),
    "returns": ("возвраты", "returns", "отмены и возвраты"),
    "stock": ("остаток", "остатки", "stock", "stocks", "доступно"),
    "ad_spend": ("реклама", "расход рекламы", "затраты на рекламу", "ad spend", "spend"),
    "price": ("цена", "price", "средняя цена"),
    "rating": ("рейтинг", "rating"),
    "reviews_count": ("отзывы", "количество отзывов", "reviews", "reviews count"),
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", " ", text).strip()


def row_metric(row: dict[str, Any], metric: str, default: Any = "") -> Any:
    normalized = {normalize_header(k): v for k, v in row.items()}
    for alias in METRIC_ALIASES.get(metric, ()):
        key = normalize_header(alias)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return default


def metric_number(value: Any) -> float:
    if value in (None, "", "—", "-"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(" ", "").replace("%", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text) if text not in {"", "-", "."} else 0.0
    except ValueError:
        return 0.0


def metric_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return now().date().isoformat()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    parsed = parse_dt(text)
    return parsed.date().isoformat() if parsed else now().date().isoformat()


async def import_metric_rows(user_id: int, rows: list[dict[str, Any]], source: str) -> dict[str, int]:
    """Нормализует отчёт и сохраняет дневные снимки для правил AI-директора."""
    stats = {"rows": 0, "linked": 0, "products_created": 0}
    for raw in rows[:5000]:
        article = str(row_metric(raw, "article", "") or "").strip()
        name = str(row_metric(raw, "name", "") or "").strip()
        marketplace = str(row_metric(raw, "marketplace", "") or "").strip().lower()
        if marketplace not in {"wb", "ozon", "avito"}:
            marketplace = "wb" if "wb" in source.lower() or "wild" in source.lower() else "ozon" if "ozon" in source.lower() else ""
        product_id = None
        if article:
            product = await DB.fetchone("SELECT id FROM products WHERE user_id=? AND article=? AND status='active' ORDER BY id LIMIT 1", (user_id, article))
            if product:
                product_id = product["id"]
            elif name:
                product_id = await DB.execute(
                    "INSERT INTO products(user_id,name,marketplace,article,created_at,updated_at) VALUES(?,?,?,?,?,?)",
                    (user_id, name[:160], marketplace or "wb", article[:100], iso_now(), iso_now()),
                )
                stats["products_created"] += 1
        values = {key: metric_number(row_metric(raw, key, 0)) for key in (
            "orders", "revenue", "views", "clicks", "cart_adds", "buyouts", "returns",
            "stock", "ad_spend", "price", "rating", "reviews_count"
        )}
        await DB.execute(
            """INSERT INTO metric_snapshots(
                user_id,product_id,marketplace,article,metric_date,orders,revenue,views,clicks,cart_adds,
                buyouts,returns,stock,ad_spend,price,rating,reviews_count,source,raw_json,created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user_id, product_id, marketplace, article, metric_date(row_metric(raw, "date", "")),
             values["orders"], values["revenue"], values["views"], values["clicks"], values["cart_adds"],
             values["buyouts"], values["returns"], values["stock"], values["ad_spend"], values["price"],
             values["rating"], values["reviews_count"], source[:200], json.dumps(raw, ensure_ascii=False)[:12000], iso_now()),
        )
        stats["rows"] += 1
        if product_id:
            stats["linked"] += 1
    return stats


def recommendation_fingerprint(user_id: int, product_id: int | None, category: str, title: str) -> str:
    raw = f"{user_id}:{product_id or 0}:{category}:{title.lower().strip()}"
    return hashlib.sha256(raw.encode()).hexdigest()[:24]


async def upsert_recommendation(
    user_id: int, *, product_id: int | None, severity: str, category: str, title: str,
    description: str, estimated_effect: float = 0, confidence: int = 60,
    action_code: str = "", evidence: dict[str, Any] | None = None,
) -> int:
    fingerprint = recommendation_fingerprint(user_id, product_id, category, title)
    existing = await DB.fetchone(
        "SELECT id,status,resolved_at FROM recommendations WHERE user_id=? AND fingerprint=? ORDER BY id DESC LIMIT 1",
        (user_id, fingerprint),
    )
    payload = json.dumps(evidence or {}, ensure_ascii=False)[:12000]
    if existing:
        status = existing.get("status") or "new"
        hold_until = parse_dt(existing.get("resolved_at"))
        if status == "snoozed" and hold_until and hold_until > now():
            return existing["id"]
        if status == "resolved" and hold_until and hold_until > now() - timedelta(days=7):
            return existing["id"]
        next_status = "in_progress" if status == "in_progress" else "new"
        await DB.execute(
            """UPDATE recommendations SET severity=?,description=?,estimated_effect=?,confidence=?,action_code=?,
               evidence_json=?,status=?,created_at=?,resolved_at='' WHERE id=?""",
            (severity, description[:4000], estimated_effect, int(clamp(confidence, 0, 100)), action_code,
             payload, next_status, iso_now(), existing["id"]),
        )
        return existing["id"]
    return await DB.execute(
        """INSERT INTO recommendations(user_id,product_id,severity,title,description,estimated_effect,status,created_at,
           category,confidence,action_code,evidence_json,fingerprint) VALUES(?,?,?,?,?,?,'new',?,?,?,?,?,?)""",
        (user_id, product_id, severity, title[:240], description[:4000], estimated_effect, iso_now(),
         category, int(clamp(confidence, 0, 100)), action_code, payload, fingerprint),
    )


async def build_recommendations_for_user(user_id: int) -> list[dict[str, Any]]:
    """Детерминированный движок: создаёт задачи только из фактически сохранённых данных."""
    run_started = iso_now()
    snapshots = await DB.fetchall(
        "SELECT * FROM metric_snapshots WHERE user_id=? ORDER BY metric_date DESC,id DESC LIMIT 1200",
        (user_id,),
    )
    products = await DB.fetchall("SELECT * FROM products WHERE user_id=? AND status='active'", (user_id,))
    connections = await DB.fetchall("SELECT * FROM marketplace_connections WHERE user_id=?", (user_id,))

    if not snapshots:
        await upsert_recommendation(
            user_id, product_id=None, severity="high", category="data",
            title="Подключите данные магазина",
            description="Без продаж, остатков, рекламы и воронки AI-директор не сможет находить денежные потери. Подключите кабинет или импортируйте CSV/XLSX.",
            confidence=100, action_code="open_import",
        )

    for connection in connections:
        if connection.get("status") == "error":
            await upsert_recommendation(
                user_id, product_id=None, severity="high", category="integration",
                title=f"Восстановить подключение {connection['name']}",
                description=f"Последняя проверка кабинета завершилась ошибкой: {connection.get('last_error') or 'причина не записана'}.",
                confidence=100, action_code="open_connections", evidence={"connection_id": connection["id"]},
            )

    by_key: dict[str, list[dict[str, Any]]] = {}
    for item in snapshots:
        key = str(item.get("product_id") or item.get("article") or f"row:{item['id']}")
        by_key.setdefault(key, []).append(item)

    for series in by_key.values():
        latest = series[0]
        previous = next((x for x in series[1:] if x["metric_date"] != latest["metric_date"]), None)
        pid = latest.get("product_id")
        article = latest.get("article") or "без артикула"
        product = next((p for p in products if p["id"] == pid), None) if pid else None
        label = product["name"] if product else f"Артикул {article}"
        orders = float(latest.get("orders") or 0)
        revenue = float(latest.get("revenue") or 0)
        stock = float(latest.get("stock") or 0)
        ad_spend = float(latest.get("ad_spend") or 0)
        views = float(latest.get("views") or 0)
        clicks = float(latest.get("clicks") or 0)
        carts = float(latest.get("cart_adds") or 0)
        buyouts = float(latest.get("buyouts") or 0)
        returns = float(latest.get("returns") or 0)
        price = float(latest.get("price") or (product or {}).get("price") or 0)
        cost = float((product or {}).get("cost") or 0)

        if orders > 0 and stock >= 0:
            days_left = stock / orders if orders else math.inf
            if days_left < 5:
                lost = max(0, orders * max(price, 0) * 7)
                await upsert_recommendation(
                    user_id, product_id=pid, severity="critical", category="stock",
                    title=f"Срочно пополнить остаток: {label}",
                    description=f"При текущем темпе {orders:.1f} шт./день остатка хватит примерно на {days_left:.1f} дня. Возможная выручка под риском за неделю: {money(lost)}.",
                    estimated_effect=lost, confidence=90, action_code="open_supply", evidence={"stock": stock, "orders": orders, "days_left": days_left},
                )
            elif days_left < 14:
                await upsert_recommendation(
                    user_id, product_id=pid, severity="high", category="stock",
                    title=f"Подготовить поставку: {label}",
                    description=f"Остатка ориентировочно на {days_left:.1f} дня. Проверьте срок производства и логистики.",
                    confidence=85, action_code="open_supply", evidence={"stock": stock, "orders": orders, "days_left": days_left},
                )

        if previous and float(previous.get("orders") or 0) >= 3:
            prev_orders = float(previous.get("orders") or 0)
            delta = (orders - prev_orders) / prev_orders * 100
            if delta <= -30:
                effect = max(0, (prev_orders - orders) * max(price, 0) * 7)
                await upsert_recommendation(
                    user_id, product_id=pid, severity="high", category="sales",
                    title=f"Разобрать падение заказов: {label}",
                    description=f"Заказы снизились на {abs(delta):.1f}% относительно предыдущего периода ({prev_orders:.1f} → {orders:.1f}). Проверьте цену, наличие, рекламу, позицию и изменения карточки.",
                    estimated_effect=effect, confidence=80, action_code="open_audit", evidence={"orders_now": orders, "orders_previous": prev_orders},
                )

        if views >= 100 and clicks >= 0:
            ctr = clicks / views * 100
            if ctr < 1.5:
                await upsert_recommendation(
                    user_id, product_id=pid, severity="high", category="card",
                    title=f"Улучшить главную обложку: {label}",
                    description=f"CTR по переданным данным составляет {ctr:.2f}% при {views:.0f} показах. Сначала тестируйте обложку, цену и видимость, а не переписывайте всё описание.",
                    confidence=75, action_code="open_visual_audit", evidence={"views": views, "clicks": clicks, "ctr": ctr},
                )

        if clicks >= 30:
            cart_rate = carts / clicks * 100
            if cart_rate < 5:
                await upsert_recommendation(
                    user_id, product_id=pid, severity="medium", category="conversion",
                    title=f"Повысить конверсию карточки: {label}",
                    description=f"В корзину добавляют только {cart_rate:.1f}% посетителей карточки. Проверьте оффер, инфографику, цену, характеристики, сроки доставки и отзывы.",
                    confidence=70, action_code="open_audit", evidence={"clicks": clicks, "cart_adds": carts, "cart_rate": cart_rate},
                )

        denominator = max(buyouts, orders, 1)
        return_rate = returns / denominator * 100
        if returns >= 2 and return_rate >= 15:
            await upsert_recommendation(
                user_id, product_id=pid, severity="high", category="returns",
                title=f"Снизить возвраты: {label}",
                description=f"Доля возвратов/отмен по переданным данным около {return_rate:.1f}%. Разберите причины в отзывах и соответствие товара ожиданиям карточки.",
                confidence=80, action_code="open_reviews", evidence={"returns": returns, "base": denominator, "return_rate": return_rate},
            )

        if revenue > 0 and ad_spend > 0:
            drr = ad_spend / revenue * 100
            if drr >= 25:
                effect = max(0, ad_spend - revenue * 0.18)
                await upsert_recommendation(
                    user_id, product_id=pid, severity="high", category="ads",
                    title=f"Проверить рекламные расходы: {label}",
                    description=f"ДРР составляет примерно {drr:.1f}%. Сопоставьте его с маржой и отключите запросы без заказов до масштабирования.",
                    estimated_effect=effect, confidence=90, action_code="open_finance", evidence={"ad_spend": ad_spend, "revenue": revenue, "drr": drr},
                )

        if price > 0 and cost > 0 and price <= cost:
            await upsert_recommendation(
                user_id, product_id=pid, severity="critical", category="profit",
                title=f"Проверить убыточную цену: {label}",
                description=f"Цена {money(price)} не превышает себестоимость {money(cost)} ещё до комиссий и логистики.",
                confidence=100, action_code="open_finance", evidence={"price": price, "cost": cost},
            )

    for product in products:
        if not (product.get("current_content") or "").strip():
            await upsert_recommendation(
                user_id, product_id=product["id"], severity="low", category="card",
                title=f"Заполнить контент: {product['name']}",
                description="В паспорте товара нет сохранённого контента. Создайте карточку 360°, чтобы дальнейшие аудиты учитывали единый контекст.",
                confidence=100, action_code="open_card360",
            )

    # Automatically close fresh recommendations whose condition is no longer present.
    # Tasks explicitly taken in progress remain until the user closes them.
    await DB.execute(
        "UPDATE recommendations SET status='resolved',resolved_at=? WHERE user_id=? AND status='new' AND created_at<?",
        (iso_now(), user_id, run_started),
    )

    return await DB.fetchall(
        """SELECT * FROM recommendations WHERE user_id=? AND status IN ('new','in_progress')
           ORDER BY CASE status WHEN 'in_progress' THEN 0 ELSE 1 END,
                    CASE severity WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 ELSE 4 END,
                    estimated_effect DESC, id DESC""",
        (user_id,),
    )


async def health_report(user_id: int, recommendations: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    recommendations = recommendations if recommendations is not None else await build_recommendations_for_user(user_id)
    snapshots = await DB.fetchone("SELECT COUNT(*) c FROM metric_snapshots WHERE user_id=?", (user_id,))
    products = await DB.fetchone("SELECT COUNT(*) c FROM products WHERE user_id=? AND status='active'", (user_id,))
    connections = await DB.fetchone("SELECT COUNT(*) c FROM marketplace_connections WHERE user_id=? AND status='active'", (user_id,))
    penalties = {"critical": 16, "high": 9, "medium": 5, "low": 2}
    score = 100 - sum(penalties.get(r.get("severity", "low"), 2) for r in recommendations[:12])
    data_count = int((snapshots or {}).get("c", 0))
    product_count = int((products or {}).get("c", 0))
    connection_count = int((connections or {}).get("c", 0))
    completeness = min(100, (35 if product_count else 0) + (35 if data_count else 0) + (30 if connection_count else 0))
    if completeness < 70:
        score = min(score, 78)
    categories = {"stock": 100, "profit": 100, "card": 100, "customers": 100, "data": completeness}
    for rec in recommendations:
        category = rec.get("category")
        penalty = penalties.get(rec.get("severity", "low"), 2)
        target = "customers" if category in {"returns", "reviews"} else category
        if target in categories:
            categories[target] = max(0, categories[target] - penalty * 2)
    return {
        "score": int(clamp(score, 0, 100)),
        "data_completeness": completeness,
        "sections": categories,
        "recommendations": recommendations,
        "products": product_count,
        "snapshots": data_count,
        "connections": connection_count,
    }


# ============================================================================
# AI SERVICE
# ============================================================================

class AIService:
    def __init__(self):
        self.openai = AsyncOpenAI(api_key=CFG.openai_key) if CFG.openai_key else None
        self.anthropic = anthropic.AsyncAnthropic(api_key=CFG.anthropic_key) if anthropic and CFG.anthropic_key else None

    async def text(self, system: str, user: str, *, model: str | None = None,
                   max_tokens: int = 2200, temperature: float = 0.35) -> str:
        if not self.openai:
            raise RuntimeError("OPENAI_KEY не настроен")
        user = user[:CFG.max_input_chars]
        response = await self.openai.chat.completions.create(
            model=model or CFG.openai_text_model,
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
            max_tokens=max_tokens,
            temperature=temperature,
        )
        return clean_ai_text(response.choices[0].message.content or "")

    async def json(self, system: str, user: str, *, max_tokens: int = 2600) -> dict[str, Any]:
        strict = system + "\nВерни только валидный JSON без markdown и пояснений."
        raw = await self.text(strict, user, max_tokens=max_tokens, temperature=0.2)
        return extract_json(raw)

    async def vision(self, system: str, prompt: str, image_bytes: bytes, mime: str = "image/jpeg") -> str:
        if not self.openai:
            raise RuntimeError("OPENAI_KEY не настроен")
        b64 = base64.b64encode(image_bytes).decode()
        response = await self.openai.chat.completions.create(
            model=CFG.openai_vision_model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                ]},
            ],
            max_tokens=2800,
            temperature=0.25,
        )
        return clean_ai_text(response.choices[0].message.content or "")

    async def transcribe(self, audio_bytes: bytes, filename: str = "voice.ogg") -> str:
        if not self.openai:
            raise RuntimeError("OPENAI_KEY не настроен")
        bio = io.BytesIO(audio_bytes)
        bio.name = filename
        result = await self.openai.audio.transcriptions.create(model="whisper-1", file=bio, language="ru")
        return result.text.strip()

    async def image(self, prompt: str, size: str = "1024x1024") -> bytes:
        if not self.openai:
            raise RuntimeError("OPENAI_KEY не настроен")
        response = await self.openai.images.generate(
            model=CFG.openai_image_model,
            prompt=prompt[:4000],
            size=size,
            n=1,
        )
        item = response.data[0]
        if getattr(item, "b64_json", None):
            return base64.b64decode(item.b64_json)
        if getattr(item, "url", None):
            async with httpx.AsyncClient(timeout=60) as client:
                r = await client.get(item.url)
                r.raise_for_status()
                return r.content
        raise RuntimeError("Сервис изображений не вернул файл")

    async def image_edit(self, image_bytes: bytes, prompt: str, size: str = "1024x1024") -> bytes:
        """Редактирует исходное фото товара, чтобы лучше сохранять форму и упаковку."""
        if not self.openai:
            raise RuntimeError("OPENAI_KEY не настроен")
        source = io.BytesIO(image_bytes)
        source.name = "product.jpg"
        response = await self.openai.images.edit(
            model=CFG.openai_image_model,
            image=source,
            prompt=prompt[:4000],
            size=size,
            n=1,
        )
        item = response.data[0]
        if getattr(item, "b64_json", None):
            return base64.b64decode(item.b64_json)
        if getattr(item, "url", None):
            async with httpx.AsyncClient(timeout=90) as client:
                r = await client.get(item.url)
                r.raise_for_status()
                return r.content
        raise RuntimeError("Сервис изображений не вернул отредактированный файл")


AI = AIService()

CARD_SYSTEM = """Ты ведущий e-commerce стратег по Wildberries, Ozon и Авито.
Создавай коммерчески сильный результат без выдумывания характеристик. Если данных нет,
помечай их как «нужно уточнить». Не обещай гарантированный рост. Пиши по-русски,
конкретно, без вступлений и без фраз «конечно/вот результат». Учитывай правила площадки,
читабельность, естественное SEO и соответствие фактам."""

AUDIT_SYSTEM = """Ты строгий e-commerce аудитор. Отделяй фактические данные от гипотез.
Никогда не утверждай, что видел страницу или метрики, если они не переданы. Диагностируй
проблемы по этапам воронки: показы → CTR → карточка → корзина → заказ → выкуп → возврат.
Каждой рекомендации назначай приоритет, сложность, ожидаемый эффект и уровень уверенности."""

FINANCE_SYSTEM = """Ты финансовый аналитик маркетплейсов. Не подменяй расчёты общими советами.
Явно показывай допущения, риски и чувствительность результата. Не выдавай приблизительные
тарифы за официальные. Все выводы формулируй как управленческие рекомендации."""


async def brand_context(user_id: int) -> str:
    row = await DB.fetchone("SELECT * FROM brand_kits WHERE user_id=?", (user_id,))
    if not row:
        return "Brand Kit не заполнен."
    return (
        f"Бренд: {row['brand_name'] or 'не указан'}; тон: {row['tone']}; цвета: {row['colors']}; "
        f"стиль: {row['visual_style']}; аудитория: {row['target_audience']}; "
        f"запрещённые фразы: {row['forbidden_phrases']}"
    )


async def generate_card360(user_id: int, platform: str, product_data: str) -> str:
    brand = await brand_context(user_id)
    limits = {
        "wb": "название желательно до 60 символов; без внешних ссылок и неподтверждённых превосходных степеней",
        "ozon": "название по схеме тип + бренд + ключевые характеристики; подготовь Rich Content",
        "avito": "заголовок до 50 символов; описание с выгодами, доверием и призывом",
    }.get(platform, "соблюдай общие требования маркетплейса")
    prompt = f"""Площадка: {platform.upper()}. Правила: {limits}.
Brand Kit: {brand}
Данные товара:
{product_data}

Создай полный пакет «Карточка 360°»:
1. Проверка входных данных и список того, что нельзя выдумывать.
2. Целевая аудитория: 3 сегмента, боли, мотивы и возражения.
3. Позиционирование и УТП.
4. Пять SEO-названий с пометкой стратегии каждого; выбери лучший и объясни.
5. Финальное описание площадки.
6. Характеристики: заполненные и список «нужно уточнить».
7. SEO-кластеры: высокочастотные, средние, низкие и смысловые.
8. FAQ из 7 вопросов.
9. Структура 8 слайдов инфографики с точным текстом и задачей каждого слайда.
10. Концепция главного изображения и две идеи A/B-теста.
11. Три промта для AI-фотосессии с обязательным сохранением формы и маркировки товара.
12. Rich Content / расширенный контент.
13. Три рекламных объявления.
14. Сценарий видео 20–30 секунд.
15. План запуска на 14 дней.
16. Контроль качества: риски, переспам, неподтверждённые заявления, лимиты символов.
"""
    return await AI.text(CARD_SYSTEM, prompt, max_tokens=4200)


async def generate_card360_package(user_id: int, platform: str, product_data: str) -> dict[str, Any]:
    """Структурированный пакет: Telegram показывает краткий итог, детали открываются кнопками."""
    brand = await brand_context(user_id)
    rules = {
        "wb": "название желательно до 60 символов; без ссылок и неподтверждённых превосходных степеней",
        "ozon": "название по схеме тип + бренд + ключевые характеристики; подготовь Rich Content",
        "avito": "заголовок до 50 символов; описание с выгодами, доверием и призывом",
    }.get(platform, "соблюдай требования площадки и не выдумывай факты")
    schema = {
        "summary": "краткое резюме не более 500 символов",
        "missing_data": ["что нужно уточнить"],
        "audience": [{"segment": "", "pains": [], "motives": [], "objections": []}],
        "positioning": {"idea": "", "usp": "", "proof": []},
        "titles": [{"title": "", "strategy": "", "score": 0}],
        "recommended_title": "",
        "recommended_title_reason": "",
        "description": "",
        "characteristics": [{"name": "", "value": "", "status": "confirmed|clarify"}],
        "seo": {"high": [], "middle": [], "low": [], "semantic": []},
        "faq": [{"question": "", "answer": ""}],
        "slides": [{"number": 1, "goal": "", "headline": "", "text": "", "visual": ""}],
        "main_image": {"concept": "", "headline": "", "composition": "", "mobile_check": ""},
        "ab_tests": [{"hypothesis": "", "variant_a": "", "variant_b": "", "metric": ""}],
        "photo_prompts": [""],
        "rich_content": [{"block": "", "content": ""}],
        "ads": [{"channel": "", "headline": "", "text": ""}],
        "video": [{"seconds": "0-3", "scene": "", "voiceover": "", "caption": ""}],
        "launch_plan": [{"day": "1", "action": "", "metric": ""}],
        "quality": {"risks": [], "seo_spam": "", "unsupported_claims": [], "platform_checks": []},
    }
    prompt = f"""Площадка: {platform.upper()}. Правила: {rules}.
Brand Kit: {brand}
Данные товара:
{product_data}

Создай коммерческий пакет карточки 360°. Не выдумывай материал, состав, размер, комплект, сертификаты и свойства.
Неизвестное переноси в missing_data и characteristics со status=clarify.
Верни JSON точно по этой структуре, сохранив все ключи:
{json.dumps(schema, ensure_ascii=False)}
"""
    try:
        package = await AI.json(CARD_SYSTEM, prompt, max_tokens=5200)
        if not isinstance(package, dict) or not package.get("recommended_title"):
            raise ValueError("неполный JSON")
        quality = package.get("quality") if isinstance(package.get("quality"), dict) else {}
        risks = list(quality.get("risks") or [])
        title = str(package.get("recommended_title") or "")
        title_limit = 50 if platform == "avito" else 60 if platform == "wb" else 200
        if len(title) > title_limit:
            risks.append(f"Рекомендуемое название длиннее ориентира площадки: {len(title)} из {title_limit} символов")
        description = str(package.get("description") or "")
        if len(description) < 250:
            risks.append("Описание слишком короткое для полноценного раскрытия товара")
        if len(description) > 5000:
            risks.append("Описание требует сокращения перед публикацией")
        lowered = (title + " " + description).lower()
        for phrase in ("лучший", "№1", "номер один", "гарантированно", "100% результат"):
            if phrase in lowered:
                risks.append(f"Проверьте неподтверждённую формулировку: «{phrase}»")
        if package.get("missing_data"):
            risks.append("Перед публикацией заполните данные из раздела «Нужно уточнить»")
        quality["risks"] = list(dict.fromkeys(str(x) for x in risks))
        package["quality"] = quality
        return package
    except Exception as exc:
        log.warning("Structured card fallback: %s", exc)
        full = await generate_card360(user_id, platform, product_data)
        return {
            "summary": "Карточка создана в резервном текстовом режиме.",
            "recommended_title": "Смотрите полный результат",
            "recommended_title_reason": "Структурированный ответ модели не прошёл проверку.",
            "description": full,
            "missing_data": [], "audience": [], "positioning": {}, "titles": [],
            "characteristics": [], "seo": {}, "faq": [], "slides": [], "main_image": {},
            "ab_tests": [], "photo_prompts": [], "rich_content": [], "ads": [], "video": [],
            "launch_plan": [], "quality": {"risks": ["Проверьте результат вручную"]},
        }


def _format_items(value: Any, *, limit: int = 30) -> str:
    if value in (None, "", [], {}):
        return "—"
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return "\n".join(f"{k}: {_format_items(v, limit=limit)}" for k, v in value.items())
    if isinstance(value, list):
        lines = []
        for item in value[:limit]:
            if isinstance(item, dict):
                parts = [str(v) for v in item.values() if v not in (None, "", [], {})]
                lines.append("• " + " — ".join(parts))
            else:
                lines.append("• " + str(item))
        return "\n".join(lines) or "—"
    return str(value)


def render_card_summary(package: dict[str, Any]) -> str:
    missing = package.get("missing_data") or []
    missing_text = ", ".join(str(x) for x in missing[:5]) if missing else "критичных пробелов не найдено"
    positioning = package.get("positioning") or {}
    if isinstance(positioning, dict):
        pos = positioning.get("idea") or positioning.get("usp") or "—"
    else:
        pos = str(positioning)
    return (
        "✅ КАРТОЧКА 360° ГОТОВА\n\n"
        f"Лучшее название:\n{package.get('recommended_title') or '—'}\n\n"
        f"Позиционирование:\n{pos}\n\n"
        f"Кратко:\n{package.get('summary') or '—'}\n\n"
        f"Нужно уточнить: {missing_text}\n\n"
        "Откройте нужный раздел кнопками ниже — чат не будет перегружен длинным документом."
    )


def render_card_full(package: dict[str, Any]) -> str:
    sections = [
        ("КРАТКИЙ ИТОГ", package.get("summary")),
        ("ЦЕЛЕВАЯ АУДИТОРИЯ", package.get("audience")),
        ("ПОЗИЦИОНИРОВАНИЕ", package.get("positioning")),
        ("ВАРИАНТЫ НАЗВАНИЙ", package.get("titles")),
        ("РЕКОМЕНДУЕМОЕ НАЗВАНИЕ", package.get("recommended_title")),
        ("ОПИСАНИЕ", package.get("description")),
        ("ХАРАКТЕРИСТИКИ", package.get("characteristics")),
        ("SEO-КЛАСТЕРЫ", package.get("seo")),
        ("FAQ", package.get("faq")),
        ("ИНФОГРАФИКА", package.get("slides")),
        ("ГЛАВНОЕ ИЗОБРАЖЕНИЕ", package.get("main_image")),
        ("A/B-ТЕСТЫ", package.get("ab_tests")),
        ("ПРОМТЫ ФОТОСЕССИИ", package.get("photo_prompts")),
        ("RICH CONTENT", package.get("rich_content")),
        ("РЕКЛАМА", package.get("ads")),
        ("ВИДЕО", package.get("video")),
        ("ПЛАН ЗАПУСКА", package.get("launch_plan")),
        ("КОНТРОЛЬ КАЧЕСТВА", package.get("quality")),
        ("НУЖНО УТОЧНИТЬ", package.get("missing_data")),
    ]
    return "\n\n".join(
        f"{title}\n{_format_items(value)}" for title, value in sections
        if value not in (None, "", [], {})
    )


CARD_SECTION_MAP: dict[str, tuple[str, tuple[str, ...]]] = {
    "title": ("НАЗВАНИЯ И ПОЗИЦИОНИРОВАНИЕ", ("positioning", "titles", "recommended_title", "recommended_title_reason")),
    "desc": ("ОПИСАНИЕ И ХАРАКТЕРИСТИКИ", ("description", "characteristics", "missing_data")),
    "seo": ("SEO И FAQ", ("seo", "faq")),
    "visual": ("ВИЗУАЛ И ИНФОГРАФИКА", ("slides", "main_image", "ab_tests", "photo_prompts")),
    "promo": ("RICH CONTENT, РЕКЛАМА И ВИДЕО", ("rich_content", "ads", "video")),
    "launch": ("ПЛАН ЗАПУСКА", ("launch_plan",)),
    "quality": ("КОНТРОЛЬ КАЧЕСТВА", ("quality", "missing_data")),
}


def render_card_section(package: dict[str, Any], code: str) -> str:
    title, keys = CARD_SECTION_MAP.get(code, ("РАЗДЕЛ", tuple()))
    pieces = []
    for key in keys:
        value = package.get(key)
        if value not in (None, "", [], {}):
            pieces.append(f"{key.replace('_', ' ').upper()}\n{_format_items(value)}")
    return title + "\n\n" + ("\n\n".join(pieces) if pieces else "Данных нет.")


def deterministic_content_audit(source: str) -> dict[str, Any]:
    """Объективная проверка структуры текста. Не подменяет метрики продаж."""
    text = (source or "").strip()
    low = text.lower()
    score = 100
    checks: list[str] = []
    if len(text) < 300:
        score -= 22
        checks.append("Мало информации: текст короче 300 символов")
    elif len(text) < 700:
        score -= 8
        checks.append("Описание можно раскрыть подробнее")
    if len(text) > 6500:
        score -= 10
        checks.append("Текст перегружен и требует сокращения")
    first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
    if len(first_line) > 120:
        score -= 8
        checks.append("Первый заголовок слишком длинный")
    if not any(mark in text for mark in ("•", "- ", "1.", "2.")):
        score -= 8
        checks.append("Нет удобной структуры списками")
    characteristic_markers = ("материал", "состав", "размер", "комплект", "характерист", "вес", "объем", "объём")
    if sum(marker in low for marker in characteristic_markers) < 2:
        score -= 15
        checks.append("Недостаточно явных характеристик")
    benefit_markers = ("подходит", "помогает", "удоб", "защищ", "эконом", "для ", "благодаря")
    if sum(marker in low for marker in benefit_markers) < 2:
        score -= 12
        checks.append("Выгоды товара раскрыты слабо")
    risky = [phrase for phrase in ("лучший", "№1", "гарантированно", "100%", "лечит", "абсолютно безопас") if phrase in low]
    if risky:
        score -= min(20, 5 * len(risky))
        checks.append("Есть рискованные утверждения: " + ", ".join(risky))
    words = re.findall(r"[a-zа-яё0-9]+", low)
    if words:
        counts: dict[str, int] = {}
        for word in words:
            if len(word) >= 5:
                counts[word] = counts.get(word, 0) + 1
        top_word, top_count = max(counts.items(), key=lambda item: item[1], default=("", 0))
        density = top_count / max(len(words), 1) * 100
        if density > 5:
            score -= 10
            checks.append(f"Возможен переспам словом «{top_word}» — {density:.1f}%")
    score = int(clamp(score, 0, 100))
    return {
        "score": score,
        "checks": checks or ["Критических структурных ошибок не найдено"],
        "length": len(text),
        "words": len(words),
        "scope": "Оценка отражает только качество и полноту переданного контента, а не продажи или конверсию",
    }


async def audit_content(source: str, context: str = "") -> str:
    local = deterministic_content_audit(source)
    prompt = f"""Материал для аудита:
{source}

Дополнительный контекст/метрики:
{context or 'не переданы'}

Объективная локальная проверка контента:
{json.dumps(local, ensure_ascii=False)}

Подготовь:
- чётко раздели факты, расчётные сигналы и AI-гипотезы;
- не меняй локальную оценку структуры {local['score']}/100 и не называй её оценкой продаж;
- оцени отдельно название, SEO, описание, характеристики, оффер, визуал, доверие, цену и отзывы,
  но ставь «нет данных», если соответствующая информация отсутствует;
- поставь диагноз по этапам воронки только при наличии метрик;
- выдели 5 критических проблем;
- дай быстрые правки на 30 минут и план на 7 дней;
- предложи улучшенное название, первый экран и описание;
- сформируй 3 измеримых A/B-теста;
- таблицу рекомендаций: действие / основание / приоритет / сложность / ожидаемый эффект / уверенность.
"""
    ai_result = await AI.text(AUDIT_SYSTEM, prompt, max_tokens=3200)
    local_header = (
        "ОБЪЕКТИВНАЯ ПРОВЕРКА КОНТЕНТА\n"
        f"Структурная полнота: {local['score']}/100\n"
        f"Объём: {local['length']} символов · {local['words']} слов\n"
        + "\n".join(f"• {item}" for item in local["checks"])
        + f"\n\nПримечание: {local['scope']}"
    )
    return local_header + "\n\n" + ai_result


# ============================================================================
# MARKETPLACE DATA ADAPTERS
# ============================================================================

class MarketplaceError(RuntimeError):
    pass


class MarketplaceClient:
    def __init__(self):
        self.timeout = httpx.Timeout(CFG.request_timeout)
        self.headers = {"User-Agent": "Mozilla/5.0 MarketPRO/2.0", "Accept-Language": "ru-RU,ru;q=0.9"}

    async def fetch_public_card(self, url: str) -> dict[str, Any]:
        url = normalize_url(url)
        marketplace = marketplace_from_url(url)
        if marketplace not in {"wb", "ozon", "avito"}:
            raise MarketplaceError("Разрешены только ссылки Wildberries, Ozon и Авито")
        article = extract_article(url, marketplace)
        async with httpx.AsyncClient(timeout=self.timeout, follow_redirects=True, headers=self.headers) as client:
            r = await client.get(url)
            if marketplace_from_url(str(r.url)) not in {marketplace, "wb" if marketplace == "wb" else marketplace}:
                raise MarketplaceError("Ссылка перенаправила на неподдерживаемый домен")
            if r.status_code >= 400:
                raise MarketplaceError(f"Страница вернула HTTP {r.status_code}")
            content_type = r.headers.get("content-type", "")
            if "text/html" not in content_type and "application/xhtml" not in content_type:
                raise MarketplaceError("Ссылка не ведёт на HTML-карточку товара")
            text = r.text[:2_000_000]
        data: dict[str, Any] = {"marketplace": marketplace, "url": str(r.url), "article": article or ""}
        # JSON-LD is the most stable public source when available.
        blocks = re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', text, re.S | re.I)
        for block in blocks:
            try:
                parsed = json.loads(html.unescape(block.strip()))
                candidates = parsed if isinstance(parsed, list) else [parsed]
                for item in candidates:
                    if isinstance(item, dict) and item.get("@type") in {"Product", "IndividualProduct"}:
                        data.update({
                            "name": item.get("name", ""),
                            "description": item.get("description", ""),
                            "brand": item.get("brand", {}).get("name", "") if isinstance(item.get("brand"), dict) else item.get("brand", ""),
                            "image": item.get("image", ""),
                            "sku": item.get("sku", ""),
                            "rating": (item.get("aggregateRating") or {}).get("ratingValue", ""),
                            "reviews_count": (item.get("aggregateRating") or {}).get("reviewCount", ""),
                        })
                        offers = item.get("offers") or {}
                        if isinstance(offers, list):
                            offers = offers[0] if offers else {}
                        data["price"] = offers.get("price", "") if isinstance(offers, dict) else ""
            except Exception:
                continue
        if not data.get("name"):
            def meta(prop: str) -> str:
                patterns = [
                    rf'<meta[^>]+property=["\']{re.escape(prop)}["\'][^>]+content=["\'](.*?)["\']',
                    rf'<meta[^>]+content=["\'](.*?)["\'][^>]+property=["\']{re.escape(prop)}["\']',
                    rf'<meta[^>]+name=["\']{re.escape(prop)}["\'][^>]+content=["\'](.*?)["\']',
                ]
                for p in patterns:
                    m = re.search(p, text, re.I | re.S)
                    if m:
                        return html.unescape(m.group(1)).strip()
                return ""
            data.update({
                "name": meta("og:title") or meta("twitter:title"),
                "description": meta("og:description") or meta("description"),
                "image": meta("og:image"),
                "price": meta("product:price:amount"),
            })
        if not data.get("name") and not data.get("description"):
            raise MarketplaceError("Площадка не отдала публичные данные карточки. Используйте скриншоты или Seller API.")
        return data

    async def wb_validate(self, token: str) -> dict[str, Any]:
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.get("https://common-api.wildberries.ru/ping", headers={"Authorization": token})
            if r.status_code not in {200, 204}:
                raise MarketplaceError(f"WB API: HTTP {r.status_code}: {r.text[:300]}")
            return {"ok": True}

    async def wb_cards(self, token: str, limit: int = 100) -> list[dict[str, Any]]:
        payload = {"settings": {"cursor": {"limit": min(limit, 100)}, "filter": {"withPhoto": -1}}}
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.post(
                "https://content-api.wildberries.ru/content/v2/get/cards/list",
                headers={"Authorization": token, "Content-Type": "application/json"}, json=payload,
            )
            if r.status_code != 200:
                raise MarketplaceError(f"WB cards: HTTP {r.status_code}: {r.text[:400]}")
            body = r.json()
            return body.get("cards", []) if isinstance(body, dict) else []

    async def wb_orders(self, token: str, days: int = 14) -> list[dict[str, Any]]:
        date_from = (now() - timedelta(days=max(1, min(days, 90)))).date().isoformat()
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.get(
                "https://statistics-api.wildberries.ru/api/v1/supplier/orders",
                headers={"Authorization": token}, params={"dateFrom": date_from, "flag": 0},
            )
            if r.status_code != 200:
                raise MarketplaceError(f"WB orders: HTTP {r.status_code}: {r.text[:400]}")
            body = r.json()
            return body if isinstance(body, list) else []

    async def wb_sales(self, token: str, days: int = 14) -> list[dict[str, Any]]:
        date_from = (now() - timedelta(days=max(1, min(days, 90)))).date().isoformat()
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.get(
                "https://statistics-api.wildberries.ru/api/v1/supplier/sales",
                headers={"Authorization": token}, params={"dateFrom": date_from, "flag": 0},
            )
            if r.status_code != 200:
                raise MarketplaceError(f"WB sales: HTTP {r.status_code}: {r.text[:400]}")
            body = r.json()
            return body if isinstance(body, list) else []

    async def wb_stocks(self, token: str, nm_ids: list[int] | None = None) -> list[dict[str, Any]]:
        payload = {
            "nmIds": list(dict.fromkeys(nm_ids or []))[:1000],
            "chrtIds": [],
            "limit": 250000,
            "offset": 0,
        }
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.post(
                "https://seller-analytics-api.wildberries.ru/api/analytics/v1/stocks-report/wb-warehouses",
                headers={"Authorization": token, "Content-Type": "application/json"}, json=payload,
            )
            if r.status_code != 200:
                raise MarketplaceError(f"WB stocks: HTTP {r.status_code}: {r.text[:400]}")
            body = r.json()
            data = body.get("data") if isinstance(body, dict) else {}
            return (data or {}).get("items", []) if isinstance(data, dict) else []

    async def ozon_validate(self, client_id: str, token: str) -> dict[str, Any]:
        headers = {"Client-Id": client_id, "Api-Key": token, "Content-Type": "application/json"}
        payload = {"filter": {"visibility": "ALL"}, "last_id": "", "limit": 1}
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.post("https://api-seller.ozon.ru/v3/product/list", headers=headers, json=payload)
            if r.status_code != 200:
                raise MarketplaceError(f"Ozon API: HTTP {r.status_code}: {r.text[:300]}")
            return {"ok": True}

    async def ozon_products(self, client_id: str, token: str, limit: int = 100) -> list[dict[str, Any]]:
        headers = {"Client-Id": client_id, "Api-Key": token, "Content-Type": "application/json"}
        payload = {"filter": {"visibility": "ALL"}, "last_id": "", "limit": min(limit, 100)}
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            r = await client.post("https://api-seller.ozon.ru/v3/product/list", headers=headers, json=payload)
            if r.status_code != 200:
                raise MarketplaceError(f"Ozon products: HTTP {r.status_code}: {r.text[:400]}")
            return ((r.json().get("result") or {}).get("items") or [])


MP = MarketplaceClient()


async def upsert_synced_product(user_id: int, marketplace: str, item: dict[str, Any]) -> int | None:
    if marketplace == "wb":
        article = str(item.get("nmID") or item.get("nmId") or item.get("id") or "")
        name = str(item.get("title") or item.get("vendorCode") or f"WB {article}").strip()
        brand = str(item.get("brand") or "")
        category = str(item.get("subjectName") or item.get("subjectID") or "")
        source = json.dumps(item, ensure_ascii=False, default=str)[:12000]
    else:
        article = str(item.get("offer_id") or item.get("product_id") or item.get("sku") or "")
        name = str(item.get("name") or item.get("offer_id") or f"Ozon {article}").strip()
        brand = str(item.get("brand") or "")
        category = str(item.get("category_name") or "")
        source = json.dumps(item, ensure_ascii=False, default=str)[:12000]
    if not article:
        return None
    existing = await DB.fetchone(
        "SELECT id,name FROM products WHERE user_id=? AND marketplace=? AND article=? ORDER BY id LIMIT 1",
        (user_id, marketplace, article),
    )
    if existing:
        current_name = existing.get("name") or ""
        chosen_name = name if name and not name.startswith(("WB ", "Ozon ")) else current_name
        await DB.execute(
            "UPDATE products SET name=?,brand=COALESCE(NULLIF(?,''),brand),category=COALESCE(NULLIF(?,''),category),source_text=?,updated_at=? WHERE id=?",
            (chosen_name or current_name, brand, category, source, iso_now(), existing["id"]),
        )
        return existing["id"]
    return await DB.execute(
        "INSERT INTO products(user_id,name,marketplace,article,brand,category,source_text,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (user_id, name or f"{marketplace.upper()} {article}", marketplace, article, brand, category, source, iso_now(), iso_now()),
    )


def wb_record_date(item: dict[str, Any]) -> str:
    for key in ("date", "lastChangeDate", "rr_dt"):
        value = str(item.get(key) or "").strip()
        if value:
            return metric_date(value)
    return now().date().isoformat()


def wb_record_article(item: dict[str, Any]) -> str:
    return str(item.get("nmId") or item.get("nmID") or item.get("nm_id") or "").strip()


def wb_record_amount(item: dict[str, Any]) -> float:
    for key in ("finishedPrice", "priceWithDisc", "forPay", "totalPrice"):
        value = metric_number(item.get(key))
        if value:
            return value
    return 0.0


async def sync_wb_metric_snapshots(
    user_id: int, orders: list[dict[str, Any]], sales: list[dict[str, Any]], stocks: list[dict[str, Any]],
) -> int:
    """Aggregates verified WB API data into daily product snapshots without duplicating re-syncs."""
    aggregates: dict[tuple[str, str], dict[str, Any]] = {}

    def bucket(article: str, day: str) -> dict[str, Any]:
        key = (article, day)
        if key not in aggregates:
            aggregates[key] = {
                "orders": 0.0, "revenue": 0.0, "buyouts": 0.0, "returns": 0.0,
                "stock": 0.0, "price_sum": 0.0, "price_count": 0, "raw": {"orders": 0, "sales": 0, "stocks": 0},
            }
        return aggregates[key]

    for item in orders:
        article = wb_record_article(item)
        if not article or bool(item.get("isCancel")):
            continue
        data = bucket(article, wb_record_date(item))
        amount = wb_record_amount(item)
        data["orders"] += 1
        data["revenue"] += max(0.0, amount)
        if amount > 0:
            data["price_sum"] += amount
            data["price_count"] += 1
        data["raw"]["orders"] += 1

    for item in sales:
        article = wb_record_article(item)
        if not article:
            continue
        data = bucket(article, wb_record_date(item))
        sale_id = str(item.get("saleID") or item.get("saleId") or "").upper()
        is_return = sale_id.startswith("R") or bool(item.get("isStorno"))
        if is_return:
            data["returns"] += 1
        else:
            data["buyouts"] += 1
        data["raw"]["sales"] += 1

    today = now().date().isoformat()
    for item in stocks:
        article = wb_record_article(item)
        if not article:
            continue
        data = bucket(article, today)
        data["stock"] += max(0.0, metric_number(item.get("quantity")))
        data["raw"]["stocks"] += 1
        data["raw"]["in_way_to_client"] = data["raw"].get("in_way_to_client", 0) + metric_number(item.get("inWayToClient"))
        data["raw"]["in_way_from_client"] = data["raw"].get("in_way_from_client", 0) + metric_number(item.get("inWayFromClient"))

    saved = 0
    for (article, day), data in aggregates.items():
        product = await DB.fetchone(
            "SELECT id,price FROM products WHERE user_id=? AND marketplace='wb' AND article=? AND status='active' ORDER BY id LIMIT 1",
            (user_id, article),
        )
        product_id = product["id"] if product else None
        avg_price = data["price_sum"] / data["price_count"] if data["price_count"] else float((product or {}).get("price") or 0)
        await DB.execute(
            "DELETE FROM metric_snapshots WHERE user_id=? AND marketplace='wb' AND article=? AND metric_date=? AND source='wb_api'",
            (user_id, article, day),
        )
        await DB.execute(
            """INSERT INTO metric_snapshots(
                user_id,product_id,marketplace,article,metric_date,orders,revenue,buyouts,returns,stock,price,source,raw_json,created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user_id, product_id, "wb", article, day, data["orders"], data["revenue"], data["buyouts"],
             data["returns"], data["stock"], avg_price, "wb_api", json.dumps(data["raw"], ensure_ascii=False), iso_now()),
        )
        saved += 1
    return saved


async def sync_connection_products(connection: dict[str, Any]) -> dict[str, Any]:
    run_id = await DB.execute(
        "INSERT INTO sync_runs(user_id,connection_id,status,started_at) VALUES(?,?,?,?)",
        (connection["user_id"], connection["id"], "running", iso_now()),
    )
    try:
        token = CIPHER.decrypt(connection["token_encrypted"])
        if not token:
            raise MarketplaceError("Не удалось расшифровать токен")
        metrics_saved = 0
        metric_note = ""
        if connection["marketplace"] == "wb":
            items = await MP.wb_cards(token, 100)
        else:
            items = await MP.ozon_products(connection.get("client_id") or "", token, 100)
        synced = 0
        for item in items:
            pid = await upsert_synced_product(connection["user_id"], connection["marketplace"], item)
            if pid:
                synced += 1

        if connection["marketplace"] == "wb":
            try:
                nm_ids = [int(str(item.get("nmID") or item.get("nmId") or "0")) for item in items if str(item.get("nmID") or item.get("nmId") or "").isdigit()]
                orders, sales, stocks = await asyncio.gather(
                    MP.wb_orders(token, 14), MP.wb_sales(token, 14), MP.wb_stocks(token, nm_ids),
                )
                metrics_saved = await sync_wb_metric_snapshots(connection["user_id"], orders, sales, stocks)
                metric_note = f"; метрик сохранено {metrics_saved}"
            except Exception as metric_exc:
                # The content token may not include Statistics/Analytics scopes. Product sync stays successful.
                metric_note = f"; метрики недоступны: {str(metric_exc)[:180]}"
                log.warning("WB metrics sync skipped for connection %s: %s", connection["id"], metric_exc)

        await DB.execute(
            "UPDATE marketplace_connections SET status='active',last_sync_at=?,last_error='' WHERE id=?",
            (iso_now(), connection["id"]),
        )
        details = f"Получено {len(items)}, синхронизировано {synced}{metric_note}"
        await DB.execute(
            "UPDATE sync_runs SET status='success',items_count=?,details=?,finished_at=? WHERE id=?",
            (synced, details, iso_now(), run_id),
        )
        return {"ok": True, "received": len(items), "synced": synced, "metrics": metrics_saved, "details": details}
    except Exception as exc:
        await DB.execute(
            "UPDATE marketplace_connections SET status='error',last_error=? WHERE id=?",
            (str(exc)[:500], connection["id"]),
        )
        await DB.execute(
            "UPDATE sync_runs SET status='error',details=?,finished_at=? WHERE id=?",
            (str(exc)[:1000], iso_now(), run_id),
        )
        return {"ok": False, "error": str(exc)}


def card_data_text(data: dict[str, Any]) -> str:
    labels = {
        "marketplace": "Площадка", "article": "Артикул", "name": "Название", "description": "Описание",
        "brand": "Бренд", "price": "Цена", "rating": "Рейтинг", "reviews_count": "Отзывов", "url": "Ссылка",
    }
    lines = []
    for key, label in labels.items():
        value = data.get(key)
        if value not in (None, "", [], {}):
            lines.append(f"{label}: {value}")
    return "\n".join(lines)


# ============================================================================
# FINANCE ENGINE
# ============================================================================

@dataclass
class UnitEconomicsInput:
    price: float
    cost: float
    commission_pct: float
    logistics: float
    last_mile: float
    storage: float
    acquiring_pct: float
    tax_pct: float
    ads_pct: float
    buyout_pct: float
    return_cost: float
    packaging: float
    other: float


@dataclass
class UnitEconomicsResult:
    revenue_per_order: float
    expected_revenue: float
    commission: float
    acquiring: float
    tax: float
    ads: float
    expected_returns_cost: float
    total_cost: float
    profit: float
    margin_pct: float
    roi_pct: float
    break_even_price: float
    max_discount_pct: float
    max_ads_pct: float


def calculate_unit(inp: UnitEconomicsInput) -> UnitEconomicsResult:
    if inp.price <= 0 or inp.cost < 0:
        raise ValueError("Цена должна быть больше нуля, себестоимость — неотрицательной")
    buyout = clamp(inp.buyout_pct / 100, 0.01, 1)
    expected_revenue = inp.price * buyout
    commission = inp.price * inp.commission_pct / 100 * buyout
    acquiring = inp.price * inp.acquiring_pct / 100 * buyout
    tax = inp.price * inp.tax_pct / 100 * buyout
    ads = inp.price * inp.ads_pct / 100 * buyout
    expected_returns_cost = (1 - buyout) * inp.return_cost
    fixed = inp.cost + inp.logistics + inp.last_mile + inp.storage + inp.packaging + inp.other + expected_returns_cost
    total = fixed + commission + acquiring + tax + ads
    profit = expected_revenue - total
    margin = profit / expected_revenue * 100 if expected_revenue else -100
    roi = profit / max(inp.cost + inp.packaging, 1) * 100
    variable_rate = buyout * (1 - (inp.commission_pct + inp.acquiring_pct + inp.tax_pct + inp.ads_pct) / 100)
    break_even = fixed / variable_rate if variable_rate > 0 else math.inf
    max_discount = clamp((inp.price - break_even) / inp.price * 100, 0, 100) if math.isfinite(break_even) else 0
    base_without_ads = expected_revenue - (total - ads)
    max_ads_pct = clamp(base_without_ads / (inp.price * buyout) * 100, 0, 100) if inp.price * buyout else 0
    return UnitEconomicsResult(inp.price, expected_revenue, commission, acquiring, tax, ads,
                               expected_returns_cost, total, profit, margin, roi,
                               break_even, max_discount, max_ads_pct)


def scenario(inp: UnitEconomicsInput, price_factor: float, logistics_factor: float,
             buyout_delta: float, ads_delta: float) -> UnitEconomicsResult:
    adjusted = UnitEconomicsInput(
        price=inp.price * price_factor,
        cost=inp.cost,
        commission_pct=inp.commission_pct,
        logistics=inp.logistics * logistics_factor,
        last_mile=inp.last_mile * logistics_factor,
        storage=inp.storage * logistics_factor,
        acquiring_pct=inp.acquiring_pct,
        tax_pct=inp.tax_pct,
        ads_pct=max(0, inp.ads_pct + ads_delta),
        buyout_pct=clamp(inp.buyout_pct + buyout_delta, 1, 100),
        return_cost=inp.return_cost,
        packaging=inp.packaging,
        other=inp.other,
    )
    return calculate_unit(adjusted)


def render_unit(inp: UnitEconomicsInput) -> str:
    base = calculate_unit(inp)
    pessimistic = scenario(inp, 0.95, 1.2, -8, 3)
    optimistic = scenario(inp, 1.03, 0.95, 5, -2)
    verdict = "✅ здоровая экономика" if base.margin_pct >= 20 else "🟡 экономика требует контроля" if base.margin_pct >= 8 else "🔴 высокий риск убытка"
    return f"""🧮 ЮНИТ-ЭКОНОМИКА PRO

Цена: {money(inp.price)}
Себестоимость: {money(inp.cost)}
Выкуп: {pct(inp.buyout_pct)}

ОЖИДАЕМЫЙ РЕЗУЛЬТАТ НА ОДИН ЗАКАЗ
Выручка с учётом выкупа: {money(base.expected_revenue)}
Комиссия: −{money(base.commission)}
Логистика + последняя миля: −{money(inp.logistics + inp.last_mile)}
Хранение: −{money(inp.storage)}
Эквайринг: −{money(base.acquiring)}
Налог: −{money(base.tax)}
Реклама: −{money(base.ads)}
Возвраты: −{money(base.expected_returns_cost)}
Упаковка и прочее: −{money(inp.packaging + inp.other)}

Чистая прибыль: {money(base.profit)}
Маржинальность: {pct(base.margin_pct)}
ROI на товар и упаковку: {pct(base.roi_pct)}
Вердикт: {verdict}

БЕЗОПАСНЫЕ ГРАНИЦЫ
Точка безубыточности по цене: {money(base.break_even_price) if math.isfinite(base.break_even_price) else 'не рассчитывается'}
Максимальная скидка до нулевой прибыли: {pct(base.max_discount_pct)}
Максимальный ДРР до нулевой прибыли: {pct(base.max_ads_pct)}

СЦЕНАРИИ
🔴 Пессимистичный: {money(pessimistic.profit)} / маржа {pct(pessimistic.margin_pct)}
🟡 Базовый: {money(base.profit)} / маржа {pct(base.margin_pct)}
🟢 Оптимистичный: {money(optimistic.profit)} / маржа {pct(optimistic.margin_pct)}

⚠️ Комиссии и тарифы введены пользователем. Сверяйте их с актуальным кабинетом маркетплейса."""


def calculate_supply(sales_day: float, stock: int, in_transit: int, lead_days: int,
                     coverage_days: int, safety_days: int, growth_pct: float,
                     cost: float, buyout_pct: float) -> str:
    adjusted_sales = sales_day * (1 + growth_pct / 100)
    available = stock + in_transit
    days_left = available / adjusted_sales if adjusted_sales > 0 else math.inf
    target = math.ceil(adjusted_sales * (lead_days + coverage_days + safety_days) / max(buyout_pct / 100, 0.01))
    order_qty = max(0, target - available)
    order_date = now().date() + timedelta(days=max(0, math.floor(days_left - lead_days - safety_days))) if math.isfinite(days_left) else None
    risk = "🔴 риск out-of-stock" if days_left <= lead_days + safety_days else "🟡 готовьте поставку" if days_left <= lead_days + coverage_days else "🟢 запас достаточен"
    return f"""📦 ПЛАН ПОСТАВКИ PRO

Скорректированный спрос: {adjusted_sales:.1f} шт./день
Доступно с учётом товара в пути: {available} шт.
Запаса хватит: {'∞' if not math.isfinite(days_left) else f'{days_left:.1f} дней'}
Статус: {risk}

Целевой запас: {target} шт.
Рекомендуемый заказ: {order_qty} шт.
Инвестиции: {money(order_qty * cost)}
Рекомендуемая дата заказа: {order_date.strftime('%d.%m.%Y') if order_date else 'не требуется'}

Расчёт учитывает срок поставки, страховой запас, рост спроса и процент выкупа."""


# ============================================================================
# BOT UI
# ============================================================================

class Onboarding(StatesGroup):
    experience = State()
    marketplace = State()


class Card360(StatesGroup):
    platform = State()
    input_type = State()
    waiting_text = State()
    waiting_photo = State()
    save_name = State()


class ProductFlow(StatesGroup):
    add_platform = State()
    add_name = State()
    add_article = State()
    add_category = State()
    add_price = State()
    add_cost = State()
    add_dimensions = State()
    add_details = State()
    confirm = State()
    edit_field = State()


class BrandFlow(StatesGroup):
    name = State()
    tone = State()
    colors = State()
    fonts = State()
    style = State()
    audience = State()
    forbidden = State()
    logo = State()
    confirm = State()
    waiting = State()  # совместимость со старым сценарием


class AuditFlow(StatesGroup):
    mode = State()
    waiting_text = State()
    waiting_link = State()
    waiting_photo = State()


class CompetitorFlow(StatesGroup):
    waiting_link = State()


class NicheFlow(StatesGroup):
    waiting = State()


class ReviewFlow(StatesGroup):
    mode = State()
    waiting = State()


class AppealFlow(StatesGroup):
    type = State()
    waiting = State()


class UnitFlow(StatesGroup):
    choose_mode = State()
    waiting = State()  # быстрый ввод
    price = State()
    cost = State()
    commission = State()
    logistics = State()
    last_mile = State()
    storage = State()
    acquiring = State()
    tax = State()
    ads = State()
    buyout = State()
    return_cost = State()
    packaging = State()
    other = State()
    orders = State()
    confirm = State()


class DoctorFlow(StatesGroup):
    waiting = State()


class SupplyFlow(StatesGroup):
    choose_mode = State()
    waiting = State()
    sales = State()
    stock = State()
    in_transit = State()
    lead = State()
    coverage = State()
    safety = State()
    growth = State()
    cost = State()
    buyout = State()
    confirm = State()


class ImageFlow(StatesGroup):
    choose_mode = State()
    waiting_prompt = State()
    waiting_photo = State()
    waiting_scene = State()


class ConnectFlow(StatesGroup):
    marketplace = State()
    waiting_wb = State()
    waiting_ozon = State()


class PromoFlow(StatesGroup):
    waiting = State()


class PromoCodeFlow(StatesGroup):
    waiting = State()


class BroadcastFlow(StatesGroup):
    waiting = State()


class ImportFlow(StatesGroup):
    waiting = State()


class QuickIntentFlow(StatesGroup):
    choose_platform = State()
    choose_photo_action = State()
    brand_mismatch = State()


def kb(rows: list[list[tuple[str, str]]]) -> InlineKeyboardMarkup:
    buttons = []
    for row in rows:
        buttons.append([
            InlineKeyboardButton(text=text, url=value[4:]) if value.startswith("url:")
            else InlineKeyboardButton(text=text, callback_data=value)
            for text, value in row
        ])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def main_kb() -> InlineKeyboardMarkup:
    return kb([
        [("🚀 Карточка 360°", "card360"), ("📦 Мои товары", "products")],
        [("🎨 Визуальная студия", "visual"), ("📊 Аудит и рост", "analytics")],
        [("💰 Финансы", "finance"), ("💬 Отзывы", "reviews")],
        [("🛡 Апелляции", "appeals"), ("🤖 AI-директор", "director")],
        [("🔌 Кабинеты WB/Ozon", "connections"), ("👤 Профиль", "profile")],
        [("💳 Тарифы", "tariffs"), ("💬 Поддержка", f"url:{CFG.support_url}")],
    ])


def back_kb(target: str = "home") -> InlineKeyboardMarkup:
    return kb([[("⬅️ Назад", target), ("🏠 Меню", "home")]])


def card_platform_kb() -> InlineKeyboardMarkup:
    return kb([
        [("🟣 Wildberries", "cplat_wb"), ("🔵 Ozon", "cplat_ozon")],
        [("🟡 Авито", "cplat_avito")],
        [("🏠 Меню", "home")],
    ])


def card_result_kb(history_id: int) -> InlineKeyboardMarkup:
    return kb([
        [("🏷 Названия", f"cardsec_title_{history_id}"), ("📝 Описание", f"cardsec_desc_{history_id}")],
        [("🔑 SEO и FAQ", f"cardsec_seo_{history_id}"), ("🎨 Визуал", f"cardsec_visual_{history_id}")],
        [("📣 Продвижение", f"cardsec_promo_{history_id}"), ("🚀 План запуска", f"cardsec_launch_{history_id}")],
        [("✅ Контроль", f"cardsec_quality_{history_id}"), ("📄 Полный TXT", f"export_{history_id}")],
        [("💾 Сохранить товар", f"savehist_{history_id}"), ("🏠 Меню", "home")],
    ])


async def answer_long(message: Message, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    for index, part in enumerate(chunks(text)):
        await message.answer(part, reply_markup=reply_markup if index == len(chunks(text)) - 1 else None)


async def edit_or_answer(call: CallbackQuery, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    try:
        await call.message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest:
        await call.message.answer(text, reply_markup=reply_markup)
    await call.answer()


async def get_message_input(message: Message, bot: Bot) -> str:
    if message.text:
        return message.text.strip()
    if message.voice or message.audio:
        media = message.voice or message.audio
        file = await bot.get_file(media.file_id)
        bio = await bot.download_file(file.file_path)
        return await AI.transcribe(bio.read(), getattr(media, "file_name", None) or "voice.ogg")
    return ""


async def get_photo_bytes(message: Message, bot: Bot) -> tuple[bytes, str]:
    if not message.photo:
        raise ValueError("Фото не найдено")
    photo = message.photo[-1]
    if photo.file_size and photo.file_size > CFG.max_photo_mb * 1024 * 1024:
        raise ValueError(f"Максимальный размер фото — {CFG.max_photo_mb} МБ")
    file = await bot.get_file(photo.file_id)
    bio = await bot.download_file(file.file_path)
    return bio.read(), "image/jpeg"


async def refund_pending_state(user_id: int, state: FSMContext) -> None:
    """Возвращает зарезервированный кредит, если пользователь отменил незавершённый сценарий."""
    data = await state.get_data()
    if data.get("charge_consumed"):
        return
    if data.get("image_charged"):
        await refund(user_id, "image", image=True)
        return
    feature = data.get("charged_feature")
    if feature:
        await refund(user_id, feature)


async def access_or_paywall(call_or_message: CallbackQuery | Message, feature: str, image: bool = False) -> bool:
    uid = call_or_message.from_user.id
    ok, reason = await charge(uid, feature, image=image)
    if ok:
        return True
    text = f"⛔️ {reason}\n\nВыберите тариф или активируйте промокод."
    markup = kb([[("💳 Посмотреть тарифы", "tariffs"), ("🎁 Промокод", "promo")], [("🏠 Меню", "home")]])
    if isinstance(call_or_message, CallbackQuery):
        await call_or_message.message.answer(text, reply_markup=markup)
        await call_or_message.answer()
    else:
        await call_or_message.answer(text, reply_markup=markup)
    return False


router = Router()


@router.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await refund_pending_state(message.from_user.id, state)
    await state.clear()
    u = message.from_user
    await ensure_user(u.id, u.username or "", u.first_name or "")
    user = await DB.fetchone("SELECT onboarded FROM users WHERE user_id=?", (u.id,))

    # Atomic referral reward: unique invited_user_id prevents repeated bonuses.
    args = (message.text or "").split(maxsplit=1)
    if len(args) == 2 and args[1].startswith("ref_"):
        try:
            inviter = int(args[1][4:])
            if inviter != u.id and await DB.fetchone("SELECT user_id FROM users WHERE user_id=?", (inviter,)):
                try:
                    await DB.transaction([
                        ("INSERT INTO referral_events(invited_user_id,inviter_user_id,registered_rewarded,created_at) VALUES(?,?,1,?)",
                         (u.id, inviter, iso_now())),
                        ("UPDATE users SET ref_by=?,referral_rewarded=1 WHERE user_id=? AND referral_rewarded=0", (inviter, u.id)),
                        ("UPDATE balances SET credits=credits+? WHERE user_id=?", (CFG.referral_invited, u.id)),
                        ("UPDATE balances SET credits=credits+? WHERE user_id=?", (CFG.referral_inviter, inviter)),
                    ])
                    try:
                        await message.bot.send_message(inviter, f"🎉 Новый пользователь по вашей ссылке. Начислено +{CFG.referral_inviter} кредитов.")
                    except Exception:
                        pass
                except sqlite3.IntegrityError:
                    pass
        except ValueError:
            pass

    if not user or not user["onboarded"]:
        await state.set_state(Onboarding.experience)
        await message.answer(
            "👋 Добро пожаловать в МаркетПРО Premium.\n\n"
            "Я не просто генерирую тексты: сохраняю товары, диагностирую воронку, считаю прибыль "
            "и формирую план действий.\n\nКакой у вас опыт?",
            reply_markup=kb([[("🌱 Начинаю", "exp_newbie"), ("📈 Уже продаю", "exp_seller")], [("🏢 Команда/агентство", "exp_agency")]]),
        )
        return
    await show_home(message)


@router.message(Command("cancel"))
async def cancel(message: Message, state: FSMContext):
    await refund_pending_state(message.from_user.id, state)
    await state.clear()
    await message.answer("Действие отменено. Зарезервированные кредиты возвращены.", reply_markup=main_kb())


@router.callback_query(F.data.startswith("exp_"), Onboarding.experience)
async def onboarding_experience(call: CallbackQuery, state: FSMContext):
    experience = call.data[4:]
    await state.update_data(experience=experience)
    await state.set_state(Onboarding.marketplace)
    await edit_or_answer(call, "Основная площадка:", kb([[("🟣 WB", "onb_wb"), ("🔵 Ozon", "onb_ozon")], [("🟡 Авито", "onb_avito"), ("🌐 Несколько", "onb_multi")]]))


@router.callback_query(F.data.startswith("onb_"), Onboarding.marketplace)
async def onboarding_marketplace(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await DB.execute("UPDATE users SET onboarded=1,experience=?,primary_marketplace=? WHERE user_id=?",
                     (data.get("experience", "newbie"), call.data[4:], call.from_user.id))
    await state.clear()
    await call.message.answer("✅ Профиль настроен. Для первого результата рекомендую «Карточка 360°».", reply_markup=main_kb())
    await call.answer()


async def _home_text(user_id: int) -> str:
    plan, expiry = await user_plan(user_id)
    bal = await balance(user_id)
    recommendations = await build_recommendations_for_user(user_id)
    health = await health_report(user_id, recommendations)
    status = f"{PLANS[plan]['name']}" + (f" до {expiry.strftime('%d.%m.%Y')}" if expiry else "")
    urgent = [r for r in recommendations if r.get("severity") in {"critical", "high"}][:3]
    tasks = ""
    if urgent:
        tasks = "\n\nГЛАВНОЕ СЕЙЧАС\n" + "\n".join(
            f"• {'🔴' if r['severity']=='critical' else '🟠'} {r['title']}" for r in urgent
        )
    elif recommendations:
        tasks = f"\n\nАктивных задач: {len(recommendations)} — откройте AI-директора."
    else:
        tasks = "\n\n✅ Критических задач по доступным данным нет."
    completeness = health["data_completeness"]
    data_note = "полные данные" if completeness >= 90 else f"полнота данных {completeness}%"
    return (
        f"🏠 МАРКЕТПРО · AI-ДИРЕКТОР\n\n"
        f"Здоровье магазина: {health['score']}/100 · {data_note}\n"
        f"Тариф: {status}\n"
        f"Кредиты: {bal.get('credits', 0)} · Изображения: {bal.get('image_credits', 0)}\n"
        f"Товаров: {health['products']} · Активных задач: {len(recommendations)}"
        f"{tasks}\n\nВыберите действие:"
    )


async def show_home(message: Message):
    await message.answer(await _home_text(message.from_user.id), reply_markup=main_kb())


@router.callback_query(F.data == "home")
async def home_cb(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await edit_or_answer(call, await _home_text(call.from_user.id), main_kb())


# --- Card 360 ---------------------------------------------------------------
@router.callback_query(F.data == "card360")
async def card360_start(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await state.set_state(Card360.platform)
    await edit_or_answer(call, "🚀 КАРТОЧКА 360°\n\nСоздам контент, SEO, инфографику, Rich Content, рекламу и план запуска.\nВыберите площадку:", card_platform_kb())


@router.callback_query(F.data.startswith("cplat_"), Card360.platform)
async def card360_platform(call: CallbackQuery, state: FSMContext):
    await state.update_data(platform=call.data[6:])
    await state.set_state(Card360.input_type)
    await edit_or_answer(call, "Откуда взять данные о товаре?", kb([[("✍️ Текст/голос", "ctype_text"), ("📷 Фото", "ctype_photo")], [("🏠 Меню", "home")]]))


@router.callback_query(F.data == "ctype_text", Card360.input_type)
async def card360_text_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "card360_text"):
        return
    await state.update_data(charged_feature="card360_text")
    await state.set_state(Card360.waiting_text)
    await edit_or_answer(call, "Опишите товар: название, материал, размер, комплект, аудитория, преимущества, цена. Можно голосом.\n\nНеизвестные характеристики бот не станет выдумывать.", back_kb("card360"))


@router.message(Card360.waiting_text, F.text | F.voice | F.audio)
async def card360_text_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        if len(text) < 20:
            raise ValueError("Добавьте больше данных о товаре — минимум 20 символов")
        progress = await message.answer("⏳ Собираю стратегию, SEO, визуал и план запуска...")
        package = await generate_card360_package(message.from_user.id, data.get("platform", "wb"), text)
        full_result = render_card_full(package)
        hid = await save_history(
            message.from_user.id, "card360", text, full_result,
            metadata={"platform": data.get("platform"), "card_package": package},
        )
        await progress.delete()
        await message.answer(render_card_summary(package), reply_markup=card_result_kb(hid))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "card360_text"))
        log.exception("card360 text")
        await message.answer(f"❌ Не удалось создать карточку: {exc}", reply_markup=back_kb("card360"))
    await state.clear()


@router.callback_query(F.data == "ctype_photo", Card360.input_type)
async def card360_photo_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "card360_photo"):
        return
    await state.update_data(charged_feature="card360_photo")
    await state.set_state(Card360.waiting_photo)
    await edit_or_answer(call, "Отправьте чёткое фото товара. После распознавания я создам карточку 360° без выдумывания скрытых характеристик.", back_kb("card360"))


@router.message(Card360.waiting_photo, F.photo)
async def card360_photo_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        image, mime = await get_photo_bytes(message, message.bot)
        progress = await message.answer("⏳ Анализирую товар на фото...")
        recognized = await AI.vision(
            "Ты товаровед. Определи только видимые характеристики товара. Не угадывай бренд, материал и комплектность. Если это не товар, ответь НЕ_ТОВАР.",
            "Опиши товар максимально точно для последующего создания карточки маркетплейса. Отдельно перечисли неизвестные данные.", image, mime,
        )
        if "НЕ_ТОВАР" in recognized.upper():
            raise ValueError("На фото не удалось распознать товар")
        package = await generate_card360_package(message.from_user.id, data.get("platform", "wb"), recognized)
        full_result = render_card_full(package)
        hid = await save_history(
            message.from_user.id, "card360_photo", recognized, full_result,
            metadata={"platform": data.get("platform"), "card_package": package},
        )
        await progress.delete()
        await message.answer(render_card_summary(package), reply_markup=card_result_kb(hid))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "card360_photo"))
        log.exception("card360 photo")
        await message.answer(f"❌ {exc}", reply_markup=back_kb("card360"))
    await state.clear()


@router.callback_query(F.data.startswith("savehist_"))
async def save_history_as_product(call: CallbackQuery, state: FSMContext):
    hid = int(call.data.split("_", 1)[1])
    row = await DB.fetchone("SELECT * FROM history WHERE id=? AND user_id=?", (hid, call.from_user.id))
    if not row:
        await call.answer("Результат не найден", show_alert=True)
        return
    plan, _ = await user_plan(call.from_user.id)
    count = await DB.fetchone("SELECT COUNT(*) AS c FROM products WHERE user_id=? AND status='active'", (call.from_user.id,))
    if (count or {}).get("c", 0) >= PLANS[plan]["products"]:
        await call.answer("Достигнут лимит товаров тарифа", show_alert=True)
        return
    await state.update_data(history_id=hid)
    await state.set_state(Card360.save_name)
    await call.message.answer("Введите короткое название товара для раздела «Мои товары»: ", reply_markup=back_kb("products"))
    await call.answer()


@router.message(Card360.save_name, F.text)
async def save_product_name(message: Message, state: FSMContext):
    data = await state.get_data()
    row = await DB.fetchone("SELECT * FROM history WHERE id=? AND user_id=?", (data["history_id"], message.from_user.id))
    if not row:
        await message.answer("Результат не найден")
        await state.clear(); return
    meta = json.loads(row["metadata"] or "{}")
    package = meta.get("card_package") or {}
    seo = package.get("seo") or {}
    keywords = ", ".join(
        str(x) for group in seo.values() for x in (group if isinstance(group, list) else [])
    ) if isinstance(seo, dict) else ""
    pid = await DB.execute(
        """INSERT INTO products(
            user_id,name,marketplace,source_text,current_content,audience,benefits,characteristics,keywords,created_at,updated_at
           ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
        (message.from_user.id, message.text[:120], meta.get("platform", "wb"), row["input_text"], row["result"],
         _format_items(package.get("audience"))[:4000], _format_items(package.get("positioning"))[:4000],
         json.dumps(package.get("characteristics") or [], ensure_ascii=False)[:8000], keywords[:8000], iso_now(), iso_now()),
    )
    await DB.execute("UPDATE history SET product_id=? WHERE id=?", (pid, row["id"]))
    await state.clear()
    await message.answer("✅ Товар сохранён. Теперь все будущие результаты можно связывать с его паспортом.", reply_markup=back_kb("products"))


# --- Products and brand ------------------------------------------------------
@router.callback_query(F.data == "products")
async def products_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    rows = await DB.fetchall("SELECT id,name,marketplace,price,cost FROM products WHERE user_id=? AND status='active' ORDER BY updated_at DESC LIMIT 20", (call.from_user.id,))
    buttons = [[(f"{p['name']} · {p['marketplace'].upper()}", f"product_{p['id']}")] for p in rows]
    buttons += [[("➕ Добавить вручную", "product_add"), ("🎨 Brand Kit", "brandkit")], [("🏠 Меню", "home")]]
    text = "📦 МОИ ТОВАРЫ\n\n" + ("Выберите товар:" if rows else "Товаров пока нет. Создайте карточку 360° или добавьте товар вручную.")
    await edit_or_answer(call, text, kb(buttons))


@router.callback_query(F.data.startswith("product_") & ~F.data.in_({"product_add"}))
async def product_view(call: CallbackQuery):
    try:
        pid = int(call.data.split("_")[1])
    except ValueError:
        return
    p = await DB.fetchone("SELECT * FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not p:
        await call.answer("Товар не найден", show_alert=True); return
    text = f"""📦 {p['name']}
Площадка: {p['marketplace'].upper()}
Артикул: {p['article'] or '—'}
Категория: {p['category'] or '—'}
Цена: {money(p['price']) if p['price'] else '—'}
Себестоимость: {money(p['cost']) if p['cost'] else '—'}
Обновлён: {parse_dt(p['updated_at']).strftime('%d.%m.%Y') if parse_dt(p['updated_at']) else '—'}"""
    await edit_or_answer(call, text, kb([
        [("🚀 Карточка 360°", f"pcard_{pid}"), ("🔍 Аудит", f"paudit_{pid}")],
        [("🧮 Экономика", f"pfinance_{pid}"), ("📊 Метрики", f"pmetrics_{pid}")],
        [("📄 Контент", f"pcontent_{pid}"), ("✏️ Редактировать", f"pedit_{pid}")],
        [("🗄 Архивировать", f"parchive_{pid}"), ("⬅️ Товары", "products")],
        [("🏠 Меню", "home")],
    ]))


@router.callback_query(F.data.startswith("pcard_"))
async def product_card360(call: CallbackQuery):
    pid = int(call.data.split("_")[1])
    product = await DB.fetchone("SELECT * FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    if not await access_or_paywall(call, "card360_text"):
        return
    try:
        context = (
            f"Название: {product['name']}\nКатегория: {product['category']}\nАртикул: {product['article']}\n"
            f"Цена: {product['price']}\nСебестоимость: {product['cost']}\n"
            f"Характеристики и контекст: {product['source_text']}\n{product['characteristics']}"
        )
        progress = await call.message.answer("⏳ Создаю новый пакет карточки для сохранённого товара...")
        package = await generate_card360_package(call.from_user.id, product["marketplace"], context)
        full = render_card_full(package)
        hid = await save_history(call.from_user.id, "product_card360", context, full, product_id=pid,
                                 metadata={"platform": product["marketplace"], "card_package": package})
        await DB.execute("UPDATE products SET current_content=?,audience=?,benefits=?,characteristics=?,updated_at=? WHERE id=?",
                         (full, _format_items(package.get("audience"))[:4000], _format_items(package.get("positioning"))[:4000],
                          json.dumps(package.get("characteristics") or [], ensure_ascii=False)[:8000], iso_now(), pid))
        try:
            await progress.delete()
        except Exception:
            pass
        await call.answer()
        await call.message.answer(render_card_summary(package), reply_markup=card_result_kb(hid))
    except Exception as exc:
        await refund(call.from_user.id, "card360_text")
        await call.message.answer(f"❌ Не удалось создать карточку: {exc}", reply_markup=back_kb("products"))
        await call.answer()


@router.callback_query(F.data.startswith("paudit_"))
async def product_audit(call: CallbackQuery):
    pid = int(call.data.split("_")[1])
    product = await DB.fetchone("SELECT * FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    if not await access_or_paywall(call, "audit"):
        return
    try:
        metrics = await DB.fetchall("SELECT * FROM metric_snapshots WHERE user_id=? AND product_id=? ORDER BY metric_date DESC LIMIT 14", (call.from_user.id, pid))
        source = product["current_content"] or product["source_text"] or product["name"]
        result = await audit_content(source, json.dumps(metrics, ensure_ascii=False, default=str) if metrics else "Метрики не загружены")
        hid = await save_history(call.from_user.id, "product_audit", source, result, product_id=pid)
        await call.answer()
        await answer_long(call.message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("📦 Товар", f"product_{pid}")]]))
    except Exception as exc:
        await refund(call.from_user.id, "audit")
        await call.message.answer(f"❌ Ошибка аудита: {exc}")
        await call.answer()


@router.callback_query(F.data.startswith("pfinance_"))
async def product_finance(call: CallbackQuery, state: FSMContext):
    pid = int(call.data.split("_")[1])
    product = await DB.fetchone("SELECT * FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    if not await access_or_paywall(call, "finance"):
        return
    await state.clear()
    await state.update_data(charged_feature="finance", product_id=pid, price=float(product["price"] or 0), cost=float(product["cost"] or 0))
    if not product["price"] or not product["cost"]:
        await refund(call.from_user.id, "finance")
        await call.answer("Сначала укажите цену и себестоимость в паспорте товара", show_alert=True)
        return
    await state.set_state(UnitFlow.commission)
    await edit_or_answer(call, f"Цена {money(product['price'])} и себестоимость {money(product['cost'])} взяты из паспорта.\n\nЮНИТ-ЭКОНОМИКА · ШАГ 3/14\nВведите комиссию маркетплейса, %:", back_kb("products"))


@router.callback_query(F.data.startswith("pmetrics_"))
async def product_metrics(call: CallbackQuery):
    pid = int(call.data.split("_")[1])
    product = await DB.fetchone("SELECT name FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    rows = await DB.fetchall("SELECT * FROM metric_snapshots WHERE user_id=? AND product_id=? ORDER BY metric_date DESC,id DESC LIMIT 10", (call.from_user.id, pid))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    if not rows:
        await edit_or_answer(call, f"📊 {product['name']}\n\nМетрики пока не загружены. Импортируйте CSV/XLSX с артикулом товара.", kb([[("📥 Импортировать", "import_report"), ("⬅️ Товар", f"product_{pid}")]]))
        return
    lines = [f"📊 {product['name']}", ""]
    for row in rows:
        drr = row["ad_spend"] / row["revenue"] * 100 if row["revenue"] else 0
        ctr = row["clicks"] / row["views"] * 100 if row["views"] else 0
        lines.append(f"{row['metric_date']}: заказы {row['orders']:g} · выручка {money(row['revenue'])} · остаток {row['stock']:g} · CTR {ctr:.1f}% · ДРР {drr:.1f}%")
    await edit_or_answer(call, "\n".join(lines), back_kb(f"product_{pid}"))


@router.callback_query(F.data.startswith("pedit_") & ~F.data.startswith("peditfield_"))
async def product_edit_menu(call: CallbackQuery):
    try:
        pid = int(call.data.split("_")[1])
    except Exception:
        return
    product = await DB.fetchone("SELECT name FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    await edit_or_answer(
        call,
        f"✏️ РЕДАКТИРОВАНИЕ · {product['name']}\n\nВыберите поле:",
        kb([
            [("Название", f"peditfield_name_{pid}"), ("Артикул", f"peditfield_article_{pid}")],
            [("Категория", f"peditfield_category_{pid}"), ("Цена", f"peditfield_price_{pid}")],
            [("Себестоимость", f"peditfield_cost_{pid}"), ("Описание данных", f"peditfield_source_text_{pid}")],
            [("⬅️ Товар", f"product_{pid}")],
        ]),
    )


@router.callback_query(F.data.startswith("peditfield_"))
async def product_edit_field_start(call: CallbackQuery, state: FSMContext):
    match = re.match(r"^peditfield_(name|article|category|price|cost|source_text)_(\d+)$", call.data)
    if not match:
        return
    field, pid_text = match.groups()
    pid = int(pid_text)
    product = await DB.fetchone("SELECT id FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not product:
        await call.answer("Товар не найден", show_alert=True)
        return
    await state.update_data(edit_product_id=pid, edit_product_field=field)
    await state.set_state(ProductFlow.edit_field)
    labels = {
        "name": "новое название",
        "article": "новый артикул",
        "category": "новую категорию",
        "price": "новую цену в рублях",
        "cost": "новую себестоимость в рублях",
        "source_text": "обновлённые фактические данные о товаре",
    }
    await edit_or_answer(call, f"Введите {labels[field]}:", back_kb(f"product_{pid}"))


@router.message(ProductFlow.edit_field, F.text | F.voice | F.audio)
async def product_edit_field_save(message: Message, state: FSMContext):
    data = await state.get_data()
    pid = int(data.get("edit_product_id", 0))
    field = data.get("edit_product_field")
    if field not in {"name", "article", "category", "price", "cost", "source_text"}:
        await state.clear()
        await message.answer("Не удалось определить поле.")
        return
    text = await get_message_input(message, message.bot)
    try:
        value: Any = safe_float(text, 0, 100_000_000) if field in {"price", "cost"} else text.strip()
        if not value and field in {"name", "category"}:
            raise ValueError("Значение не может быть пустым")
        await DB.execute(f"UPDATE products SET {field}=?,updated_at=? WHERE id=? AND user_id=?", (value, iso_now(), pid, message.from_user.id))
        await state.clear()
        await message.answer("✅ Паспорт товара обновлён.", reply_markup=kb([[("📦 Открыть товар", f"product_{pid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await message.answer(f"❌ {exc}")


@router.callback_query(F.data.startswith("pcontent_"))
async def product_content(call: CallbackQuery):
    pid = int(call.data.split("_")[1])
    p = await DB.fetchone("SELECT current_content FROM products WHERE id=? AND user_id=?", (pid, call.from_user.id))
    if not p:
        await call.answer("Не найдено", show_alert=True); return
    await call.answer()
    await answer_long(call.message, p["current_content"] or "Контент ещё не создан.", back_kb("products"))


@router.callback_query(F.data.startswith("parchive_"))
async def product_archive(call: CallbackQuery):
    pid = int(call.data.split("_")[1])
    await DB.execute("UPDATE products SET status='archived',updated_at=? WHERE id=? AND user_id=?", (iso_now(), pid, call.from_user.id))
    await call.answer("Товар архивирован")
    await edit_or_answer(call, "✅ Товар перемещён в архив.", back_kb("products"))


@router.callback_query(F.data == "product_add")
async def product_add(call: CallbackQuery, state: FSMContext):
    plan, _ = await user_plan(call.from_user.id)
    count = await DB.fetchone("SELECT COUNT(*) c FROM products WHERE user_id=? AND status='active'", (call.from_user.id,))
    if int((count or {}).get("c", 0)) >= PLANS[plan]["products"]:
        await call.answer("Достигнут лимит товаров вашего тарифа", show_alert=True)
        return
    await state.clear()
    await state.set_state(ProductFlow.add_platform)
    await edit_or_answer(
        call,
        "➕ НОВЫЙ ТОВАР · ШАГ 1/8\n\nВыберите площадку:",
        kb([
            [("🟣 Wildberries", "paddplat_wb"), ("🔵 Ozon", "paddplat_ozon")],
            [("🟡 Авито", "paddplat_avito"), ("❌ Отмена", "products")],
        ]),
    )


@router.callback_query(F.data.startswith("paddplat_"), ProductFlow.add_platform)
async def product_add_platform(call: CallbackQuery, state: FSMContext):
    await state.update_data(marketplace=call.data.replace("paddplat_", ""))
    await state.set_state(ProductFlow.add_name)
    await edit_or_answer(call, "➕ НОВЫЙ ТОВАР · ШАГ 2/8\n\nВведите понятное название товара:", back_kb("products"))


@router.message(ProductFlow.add_name, F.text)
async def product_add_name(message: Message, state: FSMContext):
    name = (message.text or "").strip()
    if len(name) < 2:
        await message.answer("Название слишком короткое. Введите ещё раз.")
        return
    await state.update_data(name=name[:160])
    await state.set_state(ProductFlow.add_article)
    await message.answer(
        "➕ НОВЫЙ ТОВАР · ШАГ 3/8\n\nВведите артикул продавца или маркетплейса:",
        reply_markup=kb([[("Пропустить", "padd_skip_article"), ("❌ Отмена", "products")]]),
    )


@router.callback_query(F.data == "padd_skip_article", ProductFlow.add_article)
async def product_skip_article(call: CallbackQuery, state: FSMContext):
    await state.update_data(article="")
    await state.set_state(ProductFlow.add_category)
    await edit_or_answer(call, "➕ НОВЫЙ ТОВАР · ШАГ 4/8\n\nВведите категорию или предмет товара:", back_kb("products"))


@router.message(ProductFlow.add_article, F.text)
async def product_add_article(message: Message, state: FSMContext):
    await state.update_data(article=(message.text or "").strip()[:100])
    await state.set_state(ProductFlow.add_category)
    await message.answer("➕ НОВЫЙ ТОВАР · ШАГ 4/8\n\nВведите категорию или предмет товара:", reply_markup=back_kb("products"))


@router.message(ProductFlow.add_category, F.text)
async def product_add_category(message: Message, state: FSMContext):
    await state.update_data(category=(message.text or "").strip()[:160])
    await state.set_state(ProductFlow.add_price)
    await message.answer("➕ НОВЫЙ ТОВАР · ШАГ 5/8\n\nВведите текущую цену продажи в рублях:", reply_markup=back_kb("products"))


@router.message(ProductFlow.add_price, F.text)
async def product_add_price(message: Message, state: FSMContext):
    try:
        value = safe_float(message.text, 0, 100_000_000)
    except Exception:
        await message.answer("Введите число, например: 1990")
        return
    await state.update_data(price=value)
    await state.set_state(ProductFlow.add_cost)
    await message.answer("➕ НОВЫЙ ТОВАР · ШАГ 6/8\n\nВведите полную себестоимость единицы в рублях:", reply_markup=back_kb("products"))


@router.message(ProductFlow.add_cost, F.text)
async def product_add_cost(message: Message, state: FSMContext):
    try:
        value = safe_float(message.text, 0, 100_000_000)
    except Exception:
        await message.answer("Введите число, например: 650")
        return
    await state.update_data(cost=value)
    await state.set_state(ProductFlow.add_dimensions)
    await message.answer(
        "➕ НОВЫЙ ТОВАР · ШАГ 7/8\n\nВведите габариты и вес в формате:\nдлина × ширина × высота; вес_кг\n\nПример: 25 × 15 × 8; 0.7",
        reply_markup=kb([[("Пропустить", "padd_skip_dimensions"), ("❌ Отмена", "products")]]),
    )


async def _product_to_details(target: Message | CallbackQuery, state: FSMContext) -> None:
    await state.set_state(ProductFlow.add_details)
    text = (
        "➕ НОВЫЙ ТОВАР · ШАГ 8/8\n\n"
        "Опишите одним сообщением только фактические данные:\n"
        "• материал и состав;\n• размеры и комплект;\n• преимущества;\n"
        "• целевую аудиторию;\n• важные ограничения и сертификаты.\n\n"
        "Неизвестное можно не указывать."
    )
    if isinstance(target, CallbackQuery):
        await edit_or_answer(target, text, back_kb("products"))
    else:
        await target.answer(text, reply_markup=back_kb("products"))


@router.callback_query(F.data == "padd_skip_dimensions", ProductFlow.add_dimensions)
async def product_skip_dimensions(call: CallbackQuery, state: FSMContext):
    await state.update_data(length_cm=0, width_cm=0, height_cm=0, weight_kg=0)
    await _product_to_details(call, state)


@router.message(ProductFlow.add_dimensions, F.text)
async def product_add_dimensions(message: Message, state: FSMContext):
    try:
        parts = (message.text or "").replace("х", "x").replace("×", "x").split(";", 1)
        dims = [safe_float(x.strip(), 0, 10000) for x in parts[0].split("x")]
        if len(dims) != 3:
            raise ValueError
        weight = safe_float(parts[1].strip(), 0, 10000) if len(parts) > 1 else 0
    except Exception:
        await message.answer("Формат не распознан. Пример: 25 × 15 × 8; 0.7")
        return
    await state.update_data(length_cm=dims[0], width_cm=dims[1], height_cm=dims[2], weight_kg=weight)
    await _product_to_details(message, state)


@router.message(ProductFlow.add_details, F.text | F.voice | F.audio)
async def product_add_details(message: Message, state: FSMContext):
    details = await get_message_input(message, message.bot)
    if len(details) < 10:
        await message.answer("Добавьте немного больше фактических данных о товаре.")
        return
    await state.update_data(details=details[:12000])
    data = await state.get_data()
    await state.set_state(ProductFlow.confirm)
    preview = (
        "ПРОВЕРЬТЕ ПАСПОРТ ТОВАРА\n\n"
        f"Площадка: {data.get('marketplace', 'wb').upper()}\n"
        f"Название: {data.get('name')}\n"
        f"Артикул: {data.get('article') or '—'}\n"
        f"Категория: {data.get('category') or '—'}\n"
        f"Цена: {money(data.get('price', 0))}\n"
        f"Себестоимость: {money(data.get('cost', 0))}\n"
        f"Габариты: {data.get('length_cm', 0):g}×{data.get('width_cm', 0):g}×{data.get('height_cm', 0):g} см\n"
        f"Вес: {data.get('weight_kg', 0):g} кг\n\n"
        f"Данные: {details[:700]}"
    )
    await message.answer(preview, reply_markup=kb([[("✅ Сохранить", "padd_confirm"), ("❌ Отмена", "products")]]))


@router.callback_query(F.data == "padd_confirm", ProductFlow.confirm)
async def product_add_confirm(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    pid = await DB.execute(
        """INSERT INTO products(
            user_id,name,marketplace,article,category,source_text,cost,price,length_cm,width_cm,height_cm,weight_kg,created_at,updated_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (call.from_user.id, data.get("name", "Товар"), data.get("marketplace", "wb"), data.get("article", ""),
         data.get("category", ""), data.get("details", ""), data.get("cost", 0), data.get("price", 0),
         data.get("length_cm", 0), data.get("width_cm", 0), data.get("height_cm", 0), data.get("weight_kg", 0),
         iso_now(), iso_now()),
    )
    await state.clear()
    await edit_or_answer(
        call,
        "✅ Паспорт товара создан. Теперь расчёты, аудиты и рекомендации смогут использовать его цену, себестоимость и контекст.",
        kb([[("📦 Открыть товар", f"product_{pid}"), ("🚀 Создать карточку", "card360")], [("🏠 Меню", "home")]]),
    )


# --- Guided Brand Kit -------------------------------------------------------
@router.callback_query(F.data == "brandkit")
async def brandkit_start(call: CallbackQuery, state: FSMContext):
    await state.clear()
    current = await DB.fetchone("SELECT * FROM brand_kits WHERE user_id=?", (call.from_user.id,))
    preview = ""
    if current:
        preview = (
            f"\n\nСейчас: {current['brand_name'] or 'без названия'} · "
            f"{current['tone'] or 'тон не указан'} · {current['visual_style'] or 'стиль не указан'}"
        )
    await edit_or_answer(
        call,
        "🎨 BRAND KIT" + preview + "\n\nПошаговый мастер займёт 2–3 минуты. Он задаст единый стиль всем карточкам, ответам и визуалам.",
        kb([[("🧭 Заполнить пошагово", "brand_guided"), ("⚡ Быстро одним текстом", "brand_quick")], [("⬅️ Товары", "products")]]),
    )


@router.callback_query(F.data == "brand_guided")
async def brand_guided_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(BrandFlow.name)
    await edit_or_answer(call, "BRAND KIT · ШАГ 1/8\n\nВведите название бренда или магазина:", back_kb("products"))


@router.callback_query(F.data == "brand_quick")
async def brand_quick_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(BrandFlow.waiting)
    await edit_or_answer(
        call,
        "Отправьте одним сообщением: название; тон; цвета; шрифты; визуальный стиль; аудитория; запрещённые фразы.",
        back_kb("products"),
    )


@router.message(BrandFlow.name, F.text)
async def brand_name_step(message: Message, state: FSMContext):
    await state.update_data(brand_name=(message.text or "").strip()[:160])
    await state.set_state(BrandFlow.tone)
    await message.answer(
        "BRAND KIT · ШАГ 2/8\n\nВыберите тон коммуникации:",
        reply_markup=kb([
            [("Премиальный", "brandtone_premium"), ("Экспертный", "brandtone_expert")],
            [("Дружелюбный", "brandtone_friendly"), ("Энергичный", "brandtone_energy")],
            [("Минималистичный", "brandtone_minimal"), ("Заботливый", "brandtone_care")],
        ]),
    )


BRAND_TONES = {
    "premium": "премиальный, уверенный, сдержанный",
    "expert": "экспертный, конкретный, доказательный",
    "friendly": "дружелюбный, понятный, живой",
    "energy": "энергичный, динамичный, мотивирующий",
    "minimal": "минималистичный, короткий, без лишних слов",
    "care": "заботливый, спокойный, поддерживающий",
}


@router.callback_query(F.data.startswith("brandtone_"), BrandFlow.tone)
async def brand_tone_step(call: CallbackQuery, state: FSMContext):
    await state.update_data(tone=BRAND_TONES.get(call.data.replace("brandtone_", ""), "профессиональный"))
    await state.set_state(BrandFlow.colors)
    await edit_or_answer(
        call,
        "BRAND KIT · ШАГ 3/8\n\nУкажите основные цвета — названия или HEX. Пример: тёмно-синий #0B1220, золото #D4AF37.",
        kb([[("Пропустить", "brand_skip_colors"), ("❌ Отмена", "products")]]),
    )


@router.callback_query(F.data == "brand_skip_colors", BrandFlow.colors)
async def brand_skip_colors(call: CallbackQuery, state: FSMContext):
    await state.update_data(colors="")
    await state.set_state(BrandFlow.fonts)
    await edit_or_answer(call, "BRAND KIT · ШАГ 4/8\n\nУкажите шрифты или пожелание к типографике:", kb([[("Пропустить", "brand_skip_fonts")]]))


@router.message(BrandFlow.colors, F.text)
async def brand_colors_step(message: Message, state: FSMContext):
    await state.update_data(colors=(message.text or "").strip()[:500])
    await state.set_state(BrandFlow.fonts)
    await message.answer("BRAND KIT · ШАГ 4/8\n\nУкажите шрифты или пожелание к типографике:", reply_markup=kb([[("Пропустить", "brand_skip_fonts")]]))


@router.callback_query(F.data == "brand_skip_fonts", BrandFlow.fonts)
async def brand_skip_fonts(call: CallbackQuery, state: FSMContext):
    await state.update_data(fonts="")
    await state.set_state(BrandFlow.style)
    await edit_or_answer(call, "BRAND KIT · ШАГ 5/8\n\nВыберите визуальный характер:", kb([
        [("Luxury", "brandstyle_luxury"), ("Clean", "brandstyle_clean")],
        [("Natural", "brandstyle_natural"), ("Bold", "brandstyle_bold")],
        [("Editorial", "brandstyle_editorial"), ("Tech", "brandstyle_tech")],
    ]))


@router.message(BrandFlow.fonts, F.text)
async def brand_fonts_step(message: Message, state: FSMContext):
    await state.update_data(fonts=(message.text or "").strip()[:500])
    await state.set_state(BrandFlow.style)
    await message.answer("BRAND KIT · ШАГ 5/8\n\nВыберите визуальный характер:", reply_markup=kb([
        [("Luxury", "brandstyle_luxury"), ("Clean", "brandstyle_clean")],
        [("Natural", "brandstyle_natural"), ("Bold", "brandstyle_bold")],
        [("Editorial", "brandstyle_editorial"), ("Tech", "brandstyle_tech")],
    ]))


BRAND_STYLES = {
    "luxury": "luxury, много воздуха, дорогая типографика, сдержанные акценты",
    "clean": "чистый коммерческий минимализм, светлая сетка, высокая читаемость",
    "natural": "натуральный, тёплый, фактуры и мягкий свет",
    "bold": "яркий, контрастный, крупная типографика и сильный первый экран",
    "editorial": "редакционный, премиальная журнальная композиция",
    "tech": "современный технологичный стиль, точная сетка, функциональность",
}


@router.callback_query(F.data.startswith("brandstyle_"), BrandFlow.style)
async def brand_style_step(call: CallbackQuery, state: FSMContext):
    await state.update_data(visual_style=BRAND_STYLES.get(call.data.replace("brandstyle_", ""), "современный коммерческий"))
    await state.set_state(BrandFlow.audience)
    await edit_or_answer(call, "BRAND KIT · ШАГ 6/8\n\nКто ваш основной покупатель? Опишите возраст, ситуацию, потребности и уровень цены:", back_kb("products"))


@router.message(BrandFlow.audience, F.text | F.voice | F.audio)
async def brand_audience_step(message: Message, state: FSMContext):
    text = await get_message_input(message, message.bot)
    await state.update_data(target_audience=text[:2000])
    await state.set_state(BrandFlow.forbidden)
    await message.answer(
        "BRAND KIT · ШАГ 7/8\n\nКакие слова, обещания и визуальные приёмы запрещены?",
        reply_markup=kb([[("Ничего не запрещать", "brand_skip_forbidden")]]),
    )


async def _brand_to_logo(target: Message | CallbackQuery, state: FSMContext) -> None:
    await state.set_state(BrandFlow.logo)
    text = "BRAND KIT · ШАГ 8/8\n\nЗагрузите логотип как фото или пропустите шаг."
    markup = kb([[("Пропустить логотип", "brand_skip_logo"), ("❌ Отмена", "products")]])
    if isinstance(target, CallbackQuery):
        await edit_or_answer(target, text, markup)
    else:
        await target.answer(text, reply_markup=markup)


@router.callback_query(F.data == "brand_skip_forbidden", BrandFlow.forbidden)
async def brand_skip_forbidden(call: CallbackQuery, state: FSMContext):
    await state.update_data(forbidden_phrases="")
    await _brand_to_logo(call, state)


@router.message(BrandFlow.forbidden, F.text)
async def brand_forbidden_step(message: Message, state: FSMContext):
    await state.update_data(forbidden_phrases=(message.text or "").strip()[:2000])
    await _brand_to_logo(message, state)


async def _brand_confirm(target: Message | CallbackQuery, state: FSMContext, logo_file_id: str = "") -> None:
    await state.update_data(logo_file_id=logo_file_id)
    data = await state.get_data()
    await state.set_state(BrandFlow.confirm)
    preview = (
        "ПРОВЕРЬТЕ BRAND KIT\n\n"
        f"Бренд: {data.get('brand_name') or '—'}\n"
        f"Тон: {data.get('tone') or '—'}\n"
        f"Цвета: {data.get('colors') or '—'}\n"
        f"Шрифты: {data.get('fonts') or '—'}\n"
        f"Стиль: {data.get('visual_style') or '—'}\n"
        f"Аудитория: {data.get('target_audience') or '—'}\n"
        f"Запреты: {data.get('forbidden_phrases') or '—'}\n"
        f"Логотип: {'загружен' if logo_file_id else 'не загружен'}"
    )
    markup = kb([[("✅ Сохранить", "brand_confirm_save"), ("🔄 Заполнить заново", "brand_guided")], [("❌ Отмена", "products")]])
    if isinstance(target, CallbackQuery):
        await edit_or_answer(target, preview, markup)
    else:
        await target.answer(preview, reply_markup=markup)


@router.callback_query(F.data == "brand_skip_logo", BrandFlow.logo)
async def brand_skip_logo(call: CallbackQuery, state: FSMContext):
    await _brand_confirm(call, state, "")


@router.message(BrandFlow.logo, F.photo)
async def brand_logo_step(message: Message, state: FSMContext):
    await _brand_confirm(message, state, message.photo[-1].file_id)


@router.callback_query(F.data == "brand_confirm_save", BrandFlow.confirm)
async def brand_confirm_save(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await DB.execute(
        """INSERT INTO brand_kits(user_id,brand_name,tone,colors,fonts,visual_style,target_audience,forbidden_phrases,logo_file_id,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET
        brand_name=excluded.brand_name,tone=excluded.tone,colors=excluded.colors,fonts=excluded.fonts,
        visual_style=excluded.visual_style,target_audience=excluded.target_audience,
        forbidden_phrases=excluded.forbidden_phrases,logo_file_id=excluded.logo_file_id,updated_at=excluded.updated_at""",
        (call.from_user.id, data.get("brand_name", ""), data.get("tone", ""), data.get("colors", ""),
         data.get("fonts", ""), data.get("visual_style", ""), data.get("target_audience", ""),
         data.get("forbidden_phrases", ""), data.get("logo_file_id", ""), iso_now()),
    )
    await state.clear()
    await edit_or_answer(call, "✅ Brand Kit сохранён. Он автоматически применяется к новым карточкам, ответам и визуальным сценариям.", back_kb("products"))


async def _save_brandkit_text(message: Message, state: FSMContext, source_text: str) -> None:
    prompt = f"Разбери описание Brand Kit в JSON с ключами brand_name,tone,colors,fonts,visual_style,target_audience,forbidden_phrases:\n{source_text}"
    try:
        data = await AI.json("Ты бренд-стратег. Не добавляй факты, которых нет.", prompt, max_tokens=700)
    except Exception:
        data = {"brand_name": "", "tone": "профессиональный", "colors": "", "fonts": "", "visual_style": source_text, "target_audience": "", "forbidden_phrases": ""}
    await DB.execute(
        """INSERT INTO brand_kits(user_id,brand_name,tone,colors,fonts,visual_style,target_audience,forbidden_phrases,updated_at)
        VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(user_id) DO UPDATE SET brand_name=excluded.brand_name,tone=excluded.tone,
        colors=excluded.colors,fonts=excluded.fonts,visual_style=excluded.visual_style,target_audience=excluded.target_audience,
        forbidden_phrases=excluded.forbidden_phrases,updated_at=excluded.updated_at""",
        (message.from_user.id, data.get("brand_name", ""), data.get("tone", ""), data.get("colors", ""),
         data.get("fonts", ""), data.get("visual_style", ""), data.get("target_audience", ""),
         data.get("forbidden_phrases", ""), iso_now()),
    )
    await state.clear()
    await message.answer("✅ Brand Kit сохранён.", reply_markup=back_kb("products"))


@router.message(BrandFlow.waiting, F.text)
async def brandkit_quick_save(message: Message, state: FSMContext):
    await _save_brandkit_text(message, state, (message.text or "").strip())


# --- Analytics --------------------------------------------------------------
@router.callback_query(F.data == "analytics")
async def analytics_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await edit_or_answer(call, "📊 АУДИТ И РОСТ\n\nФакты и AI-гипотезы всегда разделяются.", kb([[("📝 Аудит текста", "audit_text"), ("🔗 Аудит ссылки", "audit_link")], [("📷 Аудит скриншота", "audit_photo"), ("🏆 Конкурент", "competitor")], [("🔭 Стратегия ниши", "niche"), ("📥 Импорт отчёта", "import_report")], [("🏠 Меню", "home")]]))


@router.callback_query(F.data == "audit_text")
async def audit_text_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "audit"): return
    await state.update_data(charged_feature="audit")
    await state.set_state(AuditFlow.waiting_text)
    await edit_or_answer(call, "Вставьте название, описание, характеристики и при наличии метрики воронки.", back_kb("analytics"))


@router.message(AuditFlow.waiting_text, F.text | F.voice | F.audio)
async def audit_text_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        result = await audit_content(text)
        hid = await save_history(message.from_user.id, "audit_text", text, result)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "audit"))
        await message.answer(f"❌ Ошибка аудита: {exc}", reply_markup=back_kb("analytics"))
    await state.clear()


@router.callback_query(F.data == "audit_link")
async def audit_link_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "audit_link"): return
    await state.update_data(charged_feature="audit_link")
    await state.set_state(AuditFlow.waiting_link)
    await edit_or_answer(call, "Отправьте ссылку WB, Ozon или Авито. Сначала бот загрузит реальные публичные данные; если площадка их скрывает, попросит скриншоты.", back_kb("analytics"))


@router.message(AuditFlow.waiting_link, F.text)
async def audit_link_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        card = await MP.fetch_public_card(message.text)
        source = card_data_text(card)
        await message.answer("✅ Получены реальные публичные данные:\n\n" + source[:1500])
        result = await audit_content(source, "Метрики кабинета не переданы; выводы по воронке являются гипотезами.")
        hid = await save_history(message.from_user.id, "audit_link", message.text, result, metadata=card)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "audit_link"))
        await message.answer(f"⚠️ Не удалось достоверно получить карточку: {exc}\n\nОтправьте скриншоты через «Аудит скриншота» — бот не будет выдумывать содержимое ссылки.", reply_markup=back_kb("analytics"))
    await state.clear()


@router.callback_query(F.data == "audit_photo")
async def audit_photo_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "audit"): return
    await state.update_data(charged_feature="audit")
    await state.set_state(AuditFlow.waiting_photo)
    await edit_or_answer(call, "Отправьте скриншот карточки, рекламы, аналитики или штрафа.", back_kb("analytics"))


@router.message(AuditFlow.waiting_photo, F.photo)
async def audit_photo_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        image, mime = await get_photo_bytes(message, message.bot)
        result = await AI.vision(AUDIT_SYSTEM, "Разбери этот скриншот. Сначала перечисли, что точно видно; затем проблемы, риски, приоритетные действия и данные, которых не хватает.", image, mime)
        hid = await save_history(message.from_user.id, "audit_screenshot", "screenshot", result)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "audit"))
        await message.answer(f"❌ {exc}", reply_markup=back_kb("analytics"))
    await state.clear()


@router.callback_query(F.data == "competitor")
async def competitor_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "competitor"): return
    await state.update_data(charged_feature="competitor")
    await state.set_state(CompetitorFlow.waiting_link)
    await edit_or_answer(call, "Отправьте ссылку конкурента. Анализ будет основан только на реально полученных данных.", back_kb("analytics"))


@router.message(CompetitorFlow.waiting_link, F.text)
async def competitor_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        card = await MP.fetch_public_card(message.text)
        source = card_data_text(card)
        prompt = f"""Данные конкурента:
{source}

Сделай анализ: позиционирование, оффер, SEO, доверие, цена, сильные и слабые стороны,
что нельзя копировать, как отстроиться, новая концепция продукта, таблица «конкурент / мы»,
5 A/B-тестов. Не выдумывай данные, которых нет."""
        result = await AI.text(AUDIT_SYSTEM, prompt, max_tokens=2600)
        hid = await save_history(message.from_user.id, "competitor", message.text, result, metadata=card)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "competitor"))
        await message.answer(f"⚠️ {exc}\nДля достоверного анализа используйте скриншоты.", reply_markup=back_kb("analytics"))
    await state.clear()


@router.callback_query(F.data == "niche")
async def niche_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "niche"): return
    await state.update_data(charged_feature="niche")
    await state.set_state(NicheFlow.waiting)
    await edit_or_answer(call, "Введите нишу, площадку, бюджет и ваш опыт. Это AI-стратегия; фактические объёмы продаж появятся только после импорта данных.", back_kb("analytics"))


@router.message(NicheFlow.waiting, F.text | F.voice | F.audio)
async def niche_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        result = await AI.text(CARD_SYSTEM, f"""Подготовь AI-стратегию ниши: {text}
Отдели гипотезы от фактов. Дай сегменты, Jobs-to-be-Done, ценовые уровни, продуктовые идеи,
комплектации, риски, сезонность как гипотезу, ключевые кластеры, отстройку, MVP,
план проверки спроса, бюджетный сценарий запуска и список данных, которые нужно собрать.""", max_tokens=3000)
        hid = await save_history(message.from_user.id, "niche_strategy", text, result)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "niche"))
        await message.answer(f"❌ {exc}")
    await state.clear()


@router.callback_query(F.data == "import_report")
async def import_report_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(ImportFlow.waiting)
    await edit_or_answer(call, "Отправьте CSV или XLSX с заголовками. Бот нормализует продажи, воронку, остатки и рекламу, свяжет строки с товарами и обновит задачи AI-директора. Максимум 5 МБ.", back_kb("analytics"))


@router.message(ImportFlow.waiting, F.document)
async def import_report_process(message: Message, state: FSMContext):
    doc = message.document
    filename = (doc.file_name or "report").lower()
    if not filename.endswith((".csv", ".xlsx")):
        await message.answer("Поддерживаются CSV и XLSX.")
        return
    if doc.file_size and doc.file_size > 5 * 1024 * 1024:
        await message.answer("Файл больше 5 МБ.")
        return
    try:
        file = await message.bot.get_file(doc.file_id)
        bio = await message.bot.download_file(file.file_path)
        raw = bio.read()
        rows: list[dict[str, Any]] = []
        if filename.endswith(".xlsx"):
            if not load_workbook:
                raise RuntimeError("Для XLSX установите openpyxl")
            book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = book.active
            iterator = sheet.iter_rows(values_only=True)
            headers_raw = next(iterator, None)
            if not headers_raw:
                raise ValueError("XLSX пуст")
            headers = [str(x or f"column_{i+1}").strip() for i, x in enumerate(headers_raw)]
            for values in iterator:
                if not any(v not in (None, "") for v in values):
                    continue
                rows.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
                if len(rows) >= 5000:
                    break
        else:
            decoded = None
            for enc in ("utf-8-sig", "cp1251", "utf-8"):
                try:
                    decoded = raw.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            if decoded is None:
                raise ValueError("Не удалось определить кодировку")
            sample = decoded[:8000]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
                dialect.delimiter = ";"
            rows = list(csv.DictReader(io.StringIO(decoded), dialect=dialect))[:5000]
        if not rows:
            raise ValueError("Отчёт пуст")

        normalize_stats = await import_metric_rows(message.from_user.id, rows, doc.file_name or "report")
        await DB.execute(
            "INSERT INTO imported_metrics(user_id,source,metrics_json,created_at) VALUES(?,?,?,?)",
            (message.from_user.id, doc.file_name or "report", json.dumps(rows[:1000], ensure_ascii=False, default=str), iso_now()),
        )
        recommendations = await build_recommendations_for_user(message.from_user.id)
        health = await health_report(message.from_user.id, recommendations)
        compact = json.dumps(rows[:100], ensure_ascii=False, default=str)[:CFG.max_input_chars]
        result = await AI.text(
            AUDIT_SYSTEM,
            f"""Проанализируй нормализованный отчёт продавца. Сначала перечисли, какие столбцы и периоды реально есть.
Не выдумывай отсутствующие метрики. Найди лидеров, аномалии, потери, риски остатков, рекламы и воронки.
Дай максимум 10 действий с приоритетом и укажи, какие выводы являются гипотезой.
Строки:
{compact}""",
            max_tokens=2800,
        )
        summary = (
            f"✅ ОТЧЁТ ЗАГРУЖЕН\n\n"
            f"Строк обработано: {normalize_stats['rows']}\n"
            f"Связано с товарами: {normalize_stats['linked']}\n"
            f"Создано новых паспортов: {normalize_stats['products_created']}\n"
            f"Индекс здоровья магазина: {health['score']}/100\n"
            f"Активных задач: {len(recommendations)}\n\n"
            f"{result}"
        )
        hid = await save_history(
            message.from_user.id, "import_analysis", doc.file_name or "report", summary,
            metadata={"normalized": normalize_stats, "health": health["score"]},
        )
        await answer_long(
            message,
            summary,
            kb([[("🤖 Открыть AI-директора", "director"), ("📄 Экспорт", f"export_{hid}")], [("🏠 Меню", "home")]]),
        )
    except Exception as exc:
        log.exception("report import")
        await message.answer(f"❌ Ошибка импорта: {exc}", reply_markup=back_kb("analytics"))
    await state.clear()


# --- Visual -----------------------------------------------------------------
@router.callback_query(F.data == "visual")
async def visual_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await edit_or_answer(
        call,
        "🎨 ВИЗУАЛЬНАЯ СТУДИЯ\n\nДля конкретного товара используйте режим с исходным фото — он лучше сохраняет форму и упаковку. Генерация использует image-кредиты.",
        kb([
            [("📷 Сцена из фото товара", "image_product"), ("🖼 Концепт с нуля", "image_generate")],
            [("🔍 Аудит обложки", "audit_photo"), ("🧩 Сценарий 8 слайдов", "card360")],
            [("🎨 Brand Kit", "brandkit"), ("🏠 Меню", "home")],
        ]),
    )


@router.callback_query(F.data == "image_product")
async def image_product_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "image", image=True):
        return
    await state.update_data(image_charged=True)
    await state.set_state(ImageFlow.waiting_photo)
    await edit_or_answer(
        call,
        "📷 Отправьте исходное фото товара на нейтральном фоне.\n\nЛучше всего работает крупный, резкий кадр без рук и лишних предметов.",
        back_kb("visual"),
    )


@router.message(ImageFlow.waiting_photo, F.photo)
async def image_product_photo(message: Message, state: FSMContext):
    photo = message.photo[-1]
    await state.update_data(product_photo_file_id=photo.file_id)
    await state.set_state(ImageFlow.waiting_scene)
    await message.answer(
        "Опишите нужную сцену:\n• фон и окружение;\n• аудиторию;\n• настроение;\n• формат 1:1, 3:4 или 4:5;\n• нужен ли текст.\n\nПример: премиальная ванная, мягкий утренний свет, товар на каменной полке, без текста.",
        reply_markup=back_kb("visual"),
    )


@router.message(ImageFlow.waiting_scene, F.text)
async def image_product_scene(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        file_id = data.get("product_photo_file_id")
        if not file_id:
            raise ValueError("Исходное фото потеряно — отправьте его ещё раз")
        file = await message.bot.get_file(file_id)
        bio = await message.bot.download_file(file.file_path)
        source = bio.read()
        brand = await brand_context(message.from_user.id)
        prompt = (
            "Отредактируй исходное фото как профессиональный коммерческий кадр для маркетплейса. "
            "Сохрани без изменений форму, пропорции, упаковку, логотип, маркировку и цвет товара. "
            "Не добавляй несуществующие свойства, награды и логотипы площадок. "
            f"Сцена пользователя: {message.text}. Brand Kit: {brand}."
        )
        progress = await message.answer("⏳ Создаю сцену и проверяю результат...")
        generated = await AI.image_edit(source, prompt)
        audit = await AI.vision(
            "Ты контролёр качества e-commerce визуалов. Пиши кратко и не идентифицируй бренды.",
            "Проверь изображение: читаемость, искажение товара и упаковки, артефакты, ложный текст, пригодность для маркетплейса. Дай вердикт ГОТОВО или НУЖНА ПРОВЕРКА и до 4 замечаний.",
            generated,
            "image/png",
        )
        try:
            await progress.delete()
        except Exception:
            pass
        await message.answer_photo(
            BufferedInputFile(generated, filename="marketpro_product_scene.png"),
            caption=f"✅ Сцена создана.\n\nКонтроль качества:\n{audit[:900]}",
            reply_markup=kb([[("🔄 Другой вариант", "image_product"), ("🔍 Аудит подробнее", "audit_photo")], [("🏠 Меню", "home")]]),
        )
    except Exception as exc:
        await refund(message.from_user.id, "image", image=True)
        await message.answer(f"❌ Не удалось создать сцену: {exc}", reply_markup=back_kb("visual"))
    await state.clear()


@router.callback_query(F.data == "image_generate")
async def image_generate_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "image", image=True): return
    await state.update_data(image_charged=True)
    await state.set_state(ImageFlow.waiting_prompt)
    await edit_or_answer(call, "Опишите товар и нужный кадр. Укажите формат: 1:1, 3:4 или 4:5; фон; аудиторию; текст на изображении.\n\nДля точного сохранения конкретного товара используйте специализированный image-to-image API — текстовая генерация может изменить детали упаковки.", back_kb("visual"))


@router.message(ImageFlow.waiting_prompt, F.text)
async def image_generate_process(message: Message, state: FSMContext):
    try:
        brand = await brand_context(message.from_user.id)
        prompt = f"Профессиональная карточка товара для маркетплейса. {message.text}. Brand Kit: {brand}. Чистая коммерческая композиция, высокая читаемость, без ложных наград и логотипов маркетплейса."
        progress = await message.answer("⏳ Создаю визуальный концепт...")
        image = await AI.image(prompt)
        await progress.delete()
        await message.answer_photo(BufferedInputFile(image, filename="marketpro_visual.png"), caption="✅ Визуальный концепт. Перед публикацией проверьте форму товара, упаковку, текст и маркировку.", reply_markup=back_kb("visual"))
    except Exception as exc:
        await refund(message.from_user.id, "image", image=True)
        await message.answer(f"❌ Генерация изображения недоступна: {exc}", reply_markup=back_kb("visual"))
    await state.clear()


# --- Finance ----------------------------------------------------------------
@router.callback_query(F.data == "finance")
async def finance_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await edit_or_answer(call, "💰 ФИНАНСЫ\n\nТарифы и комиссии вводятся пользователем — бот не выдаёт устаревшие проценты за актуальные.", kb([[("🧮 Юнит-экономика PRO", "unit"), ("🩺 Финансовый доктор", "doctor")], [("📦 План поставки", "supply"), ("🎯 Проверка акции", "promo_calc")], [("🏠 Меню", "home")]]))


UNIT_TEMPLATE = """Отправьте 14 чисел через точку с запятой:
цена; себестоимость; комиссия%; логистика; последняя миля; хранение; эквайринг%; налог%; ДРР%; выкуп%; стоимость возврата; упаковка; прочие расходы; продажи в месяц

Пример:
1990; 600; 20; 120; 40; 15; 1.5; 6; 12; 82; 100; 35; 20; 150"""


@router.callback_query(F.data == "unit")
async def unit_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "finance"):
        return
    await state.clear()
    await state.update_data(charged_feature="finance")
    await state.set_state(UnitFlow.choose_mode)
    await edit_or_answer(
        call,
        "🧮 ЮНИТ-ЭКОНОМИКА PRO\n\nВыберите способ заполнения. Пошаговый мастер удобнее и снижает риск ошибки.",
        kb([
            [("🧭 Пошаговый мастер", "unit_guided"), ("⚡ Быстрый ввод", "unit_quick")],
            [("⬅️ Финансы", "finance")],
        ]),
    )


@router.callback_query(F.data == "unit_quick", UnitFlow.choose_mode)
async def unit_quick(call: CallbackQuery, state: FSMContext):
    await state.set_state(UnitFlow.waiting)
    await edit_or_answer(call, UNIT_TEMPLATE, back_kb("finance"))


@router.callback_query(F.data == "unit_guided", UnitFlow.choose_mode)
async def unit_guided(call: CallbackQuery, state: FSMContext):
    await state.set_state(UnitFlow.price)
    await edit_or_answer(call, "ЮНИТ-ЭКОНОМИКА · ШАГ 1/14\n\nЦена продажи после скидок, ₽:", back_kb("finance"))


async def _unit_next(message: Message, state: FSMContext, key: str, next_state: State, prompt: str,
                     minimum: float = 0, maximum: float = 100_000_000) -> None:
    try:
        value = safe_float(message.text, minimum, maximum)
    except Exception:
        await message.answer("Введите число без пояснений. Можно использовать запятую или точку.")
        return
    await state.update_data(**{key: value})
    await state.set_state(next_state)
    await message.answer(prompt, reply_markup=back_kb("finance"))


@router.message(UnitFlow.price, F.text)
async def unit_price_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "price", UnitFlow.cost, "ЮНИТ-ЭКОНОМИКА · ШАГ 2/14\n\nПолная себестоимость единицы, ₽:")


@router.message(UnitFlow.cost, F.text)
async def unit_cost_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "cost", UnitFlow.commission, "ЮНИТ-ЭКОНОМИКА · ШАГ 3/14\n\nКомиссия маркетплейса, %:")


@router.message(UnitFlow.commission, F.text)
async def unit_commission_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "commission_pct", UnitFlow.logistics, "ЮНИТ-ЭКОНОМИКА · ШАГ 4/14\n\nЛогистика на один заказ, ₽:", 0, 100)


@router.message(UnitFlow.logistics, F.text)
async def unit_logistics_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "logistics", UnitFlow.last_mile, "ЮНИТ-ЭКОНОМИКА · ШАГ 5/14\n\nПоследняя миля на заказ, ₽:")


@router.message(UnitFlow.last_mile, F.text)
async def unit_last_mile_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "last_mile", UnitFlow.storage, "ЮНИТ-ЭКОНОМИКА · ШАГ 6/14\n\nХранение на один заказ/единицу, ₽:")


@router.message(UnitFlow.storage, F.text)
async def unit_storage_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "storage", UnitFlow.acquiring, "ЮНИТ-ЭКОНОМИКА · ШАГ 7/14\n\nЭквайринг, %:")


@router.message(UnitFlow.acquiring, F.text)
async def unit_acquiring_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "acquiring_pct", UnitFlow.tax, "ЮНИТ-ЭКОНОМИКА · ШАГ 8/14\n\nНалог с выручки, %:", 0, 100)


@router.message(UnitFlow.tax, F.text)
async def unit_tax_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "tax_pct", UnitFlow.ads, "ЮНИТ-ЭКОНОМИКА · ШАГ 9/14\n\nДРР — расходы на рекламу от выручки, %:", 0, 100)


@router.message(UnitFlow.ads, F.text)
async def unit_ads_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "ads_pct", UnitFlow.buyout, "ЮНИТ-ЭКОНОМИКА · ШАГ 10/14\n\nПроцент выкупа, %:", 0, 100)


@router.message(UnitFlow.buyout, F.text)
async def unit_buyout_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "buyout_pct", UnitFlow.return_cost, "ЮНИТ-ЭКОНОМИКА · ШАГ 11/14\n\nСтоимость обратной логистики одного возврата, ₽:", 1, 100)


@router.message(UnitFlow.return_cost, F.text)
async def unit_return_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "return_cost", UnitFlow.packaging, "ЮНИТ-ЭКОНОМИКА · ШАГ 12/14\n\nУпаковка и маркировка на единицу, ₽:")


@router.message(UnitFlow.packaging, F.text)
async def unit_packaging_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "packaging", UnitFlow.other, "ЮНИТ-ЭКОНОМИКА · ШАГ 13/14\n\nПрочие расходы на единицу, ₽:")


@router.message(UnitFlow.other, F.text)
async def unit_other_step(message: Message, state: FSMContext):
    await _unit_next(message, state, "other", UnitFlow.orders, "ЮНИТ-ЭКОНОМИКА · ШАГ 14/14\n\nПланируемое число заказов в месяц:")


@router.message(UnitFlow.orders, F.text)
async def unit_orders_step(message: Message, state: FSMContext):
    try:
        orders = safe_float(message.text, 0, 100_000_000)
    except Exception:
        await message.answer("Введите число заказов, например: 150")
        return
    await state.update_data(orders=orders)
    data = await state.get_data()
    await state.set_state(UnitFlow.confirm)
    preview = (
        "ПРОВЕРЬТЕ ДАННЫЕ\n\n"
        f"Цена: {money(data['price'])}\nСебестоимость: {money(data['cost'])}\n"
        f"Комиссия: {data['commission_pct']:g}%\nЛогистика: {money(data['logistics'])}\n"
        f"Последняя миля: {money(data['last_mile'])}\nХранение: {money(data['storage'])}\n"
        f"Эквайринг: {data['acquiring_pct']:g}%\nНалог: {data['tax_pct']:g}%\n"
        f"ДРР: {data['ads_pct']:g}%\nВыкуп: {data['buyout_pct']:g}%\n"
        f"Возврат: {money(data['return_cost'])}\nУпаковка: {money(data['packaging'])}\n"
        f"Прочее: {money(data['other'])}\nЗаказов в месяц: {orders:g}"
    )
    await message.answer(preview, reply_markup=kb([[("✅ Рассчитать", "unit_confirm"), ("🔄 Заново", "unit_guided")], [("❌ Отмена", "finance")]]))


async def _finish_unit(user_id: int, values: list[float], source_text: str) -> tuple[str, int]:
    inp = UnitEconomicsInput(*values[:13])
    result = render_unit(inp)
    monthly = calculate_unit(inp).profit * values[13]
    result += f"\n\nПРОГНОЗ ЗА МЕСЯЦ\nПри {values[13]:.0f} заказах: {money(monthly)}"
    hid = await save_history(user_id, "unit_economics", source_text, result, metadata={"values": values})
    return result, hid


@router.callback_query(F.data == "unit_confirm", UnitFlow.confirm)
async def unit_confirm(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    values = [
        data["price"], data["cost"], data["commission_pct"], data["logistics"], data["last_mile"],
        data["storage"], data["acquiring_pct"], data["tax_pct"], data["ads_pct"], data["buyout_pct"],
        data["return_cost"], data["packaging"], data["other"], data["orders"],
    ]
    try:
        result, hid = await _finish_unit(call.from_user.id, values, json.dumps(data, ensure_ascii=False))
        await state.clear()
        await call.answer()
        await call.message.answer(result, reply_markup=kb([[("📄 Экспорт", f"export_{hid}"), ("📦 План поставки", "supply")], [("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(call.from_user.id, data.get("charged_feature", "finance"))
        await state.clear()
        await call.message.answer(f"❌ Ошибка расчёта: {exc}", reply_markup=back_kb("finance"))
        await call.answer()


@router.message(UnitFlow.waiting, F.text)
async def unit_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        values = [safe_float(x.strip()) for x in message.text.split(";")]
        if len(values) != 14:
            raise ValueError("Нужно ровно 14 значений")
        result, hid = await _finish_unit(message.from_user.id, values, message.text)
        await state.clear()
        await message.answer(result, reply_markup=kb([[("📄 Экспорт", f"export_{hid}"), ("📦 План поставки", "supply")], [("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "finance"))
        await state.clear()
        await message.answer(f"❌ {exc}\n\n{UNIT_TEMPLATE}", reply_markup=back_kb("finance"))


@router.callback_query(F.data == "doctor")
async def doctor_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "doctor"): return
    await state.update_data(charged_feature="doctor")
    await state.set_state(DoctorFlow.waiting)
    await edit_or_answer(call, "Опишите финансовую ситуацию: выручка, продажи, себестоимость, комиссии, логистика, реклама, возвраты, налоги, акции, остатки и цель прибыли.", back_kb("finance"))


@router.message(DoctorFlow.waiting, F.text | F.voice | F.audio)
async def doctor_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        result = await AI.text(FINANCE_SYSTEM, f"""Проведи финансовую диагностику:
{text}

Найди, где теряются деньги; посчитай то, что можно посчитать; обозначь недостающие данные;
оцени максимальный ДРР, безопасную цену, влияние возвратов и акций; сформируй три сценария;
дай 5 действий по приоритету и ожидаемому эффекту в рублях. Не выдумывай тарифы.""", max_tokens=2600)
        hid = await save_history(message.from_user.id, "financial_doctor", text, result)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "doctor"))
        await message.answer(f"❌ {exc}")
    await state.clear()


@router.callback_query(F.data == "supply")
async def supply_start(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "supply"):
        return
    await state.clear()
    await state.update_data(charged_feature="supply")
    await state.set_state(SupplyFlow.choose_mode)
    await edit_or_answer(
        call,
        "📦 ПЛАН ПОСТАВКИ PRO\n\nВыберите способ заполнения:",
        kb([[("🧭 Пошагово", "supply_guided"), ("⚡ Быстро", "supply_quick")], [("⬅️ Финансы", "finance")]]),
    )


@router.callback_query(F.data == "supply_quick", SupplyFlow.choose_mode)
async def supply_quick(call: CallbackQuery, state: FSMContext):
    await state.set_state(SupplyFlow.waiting)
    await edit_or_answer(
        call,
        "Отправьте через ;\nпродаж/день; остаток; в пути; срок поставки дней; покрытие дней; страховой запас дней; рост%; себестоимость; выкуп%\n\nПример: 8; 70; 20; 14; 30; 7; 10; 450; 85",
        back_kb("finance"),
    )


@router.callback_query(F.data == "supply_guided", SupplyFlow.choose_mode)
async def supply_guided(call: CallbackQuery, state: FSMContext):
    await state.set_state(SupplyFlow.sales)
    await edit_or_answer(call, "ПЛАН ПОСТАВКИ · ШАГ 1/9\n\nСреднее число продаж в день:", back_kb("finance"))


async def _supply_next(message: Message, state: FSMContext, key: str, next_state: State, prompt: str,
                       minimum: float = 0, maximum: float = 100_000_000) -> None:
    try:
        value = safe_float(message.text, minimum, maximum)
    except Exception:
        await message.answer("Введите число без пояснений.")
        return
    await state.update_data(**{key: value})
    await state.set_state(next_state)
    await message.answer(prompt, reply_markup=back_kb("finance"))


@router.message(SupplyFlow.sales, F.text)
async def supply_sales_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "sales_day", SupplyFlow.stock, "ПЛАН ПОСТАВКИ · ШАГ 2/9\n\nТекущий доступный остаток, шт:")


@router.message(SupplyFlow.stock, F.text)
async def supply_stock_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "stock", SupplyFlow.in_transit, "ПЛАН ПОСТАВКИ · ШАГ 3/9\n\nСколько товара уже в пути, шт:")


@router.message(SupplyFlow.in_transit, F.text)
async def supply_transit_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "in_transit", SupplyFlow.lead, "ПЛАН ПОСТАВКИ · ШАГ 4/9\n\nСрок производства и доставки, дней:")


@router.message(SupplyFlow.lead, F.text)
async def supply_lead_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "lead_days", SupplyFlow.coverage, "ПЛАН ПОСТАВКИ · ШАГ 5/9\n\nНа сколько дней должен хватить новый запас:")


@router.message(SupplyFlow.coverage, F.text)
async def supply_coverage_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "coverage_days", SupplyFlow.safety, "ПЛАН ПОСТАВКИ · ШАГ 6/9\n\nСтраховой запас, дней:")


@router.message(SupplyFlow.safety, F.text)
async def supply_safety_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "safety_days", SupplyFlow.growth, "ПЛАН ПОСТАВКИ · ШАГ 7/9\n\nОжидаемый рост или падение спроса, % (можно 0):")


@router.message(SupplyFlow.growth, F.text)
async def supply_growth_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "growth_pct", SupplyFlow.cost, "ПЛАН ПОСТАВКИ · ШАГ 8/9\n\nСебестоимость единицы, ₽:", -90, 1000)


@router.message(SupplyFlow.cost, F.text)
async def supply_cost_step(message: Message, state: FSMContext):
    await _supply_next(message, state, "cost", SupplyFlow.buyout, "ПЛАН ПОСТАВКИ · ШАГ 9/9\n\nПроцент выкупа, %:")


@router.message(SupplyFlow.buyout, F.text)
async def supply_buyout_step(message: Message, state: FSMContext):
    try:
        buyout = safe_float(message.text, 1, 100)
    except Exception:
        await message.answer("Введите процент от 1 до 100.")
        return
    await state.update_data(buyout_pct=buyout)
    data = await state.get_data()
    await state.set_state(SupplyFlow.confirm)
    preview = (
        "ПРОВЕРЬТЕ ДАННЫЕ\n\n"
        f"Продажи/день: {data['sales_day']:g}\nОстаток: {data['stock']:g}\nВ пути: {data['in_transit']:g}\n"
        f"Срок поставки: {data['lead_days']:g} дней\nПокрытие: {data['coverage_days']:g} дней\n"
        f"Страховой запас: {data['safety_days']:g} дней\nРост: {data['growth_pct']:g}%\n"
        f"Себестоимость: {money(data['cost'])}\nВыкуп: {buyout:g}%"
    )
    await message.answer(preview, reply_markup=kb([[("✅ Рассчитать", "supply_confirm"), ("🔄 Заново", "supply_guided")], [("❌ Отмена", "finance")]]))


@router.callback_query(F.data == "supply_confirm", SupplyFlow.confirm)
async def supply_confirm(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    try:
        result = calculate_supply(
            data["sales_day"], int(data["stock"]), int(data["in_transit"]), int(data["lead_days"]),
            int(data["coverage_days"]), int(data["safety_days"]), data["growth_pct"], data["cost"], data["buyout_pct"],
        )
        hid = await save_history(call.from_user.id, "supply_plan", json.dumps(data, ensure_ascii=False), result)
        await state.clear()
        await call.answer()
        await call.message.answer(result, reply_markup=kb([[("📄 Экспорт", f"export_{hid}"), ("🧮 Экономика", "unit")], [("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(call.from_user.id, data.get("charged_feature", "supply"))
        await state.clear()
        await call.message.answer(f"❌ {exc}", reply_markup=back_kb("finance"))
        await call.answer()


@router.message(SupplyFlow.waiting, F.text)
async def supply_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        values = [safe_float(x) for x in message.text.split(";")]
        if len(values) != 9:
            raise ValueError("Нужно 9 значений")
        result = calculate_supply(values[0], int(values[1]), int(values[2]), int(values[3]), int(values[4]), int(values[5]), values[6], values[7], values[8])
        hid = await save_history(message.from_user.id, "supply_plan", message.text, result)
        await state.clear()
        await message.answer(result, reply_markup=kb([[("📄 Экспорт", f"export_{hid}"), ("🧮 Экономика", "unit")], [("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "supply"))
        await state.clear()
        await message.answer(f"❌ {exc}", reply_markup=back_kb("finance"))


@router.callback_query(F.data == "promo_calc")
async def promo_calc_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(PromoFlow.waiting)
    await edit_or_answer(call, "Отправьте: текущая цена; скидка%; чистая прибыль без акции; ожидаемый рост продаж%\nПример: 1990; 20; 420; 35", back_kb("finance"))


@router.message(PromoFlow.waiting, F.text)
async def promo_calc_process(message: Message, state: FSMContext):
    try:
        price, discount, profit, growth = [safe_float(x) for x in message.text.split(";")]
        new_price = price * (1 - discount / 100)
        lost = price - new_price
        new_profit = profit - lost
        volume_factor = 1 + growth / 100
        indexed = new_profit * volume_factor
        verdict = "✅ акция может увеличить общую прибыль" if indexed > profit else "❌ прогнозируемый рост не компенсирует скидку"
        result = f"🎯 ПРОВЕРКА АКЦИИ\n\nЦена в акции: {money(new_price)}\nПрибыль с единицы: {money(new_profit)}\nИндекс прибыли при росте продаж: {money(indexed)} против {money(profit)}\n\n{verdict}\n\n⚠️ Упрощённая проверка. Для точного решения используйте Юнит-экономику PRO с новой ценой."
        await message.answer(result, reply_markup=back_kb("finance"))
    except Exception as exc:
        await message.answer(f"❌ {exc}")
    await state.clear()


# --- Reviews and appeals -----------------------------------------------------
@router.callback_query(F.data == "reviews")
async def reviews_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    await edit_or_answer(call, "💬 ОТЗЫВЫ И ПОКУПАТЕЛИ", kb([[("⭐ Ответ на отзыв", "review_one"), ("📊 Анализ массива", "review_batch")], [("❓ FAQ из вопросов", "review_faq"), ("🏠 Меню", "home")]]))


@router.callback_query(F.data.in_({"review_one", "review_batch", "review_faq"}))
async def review_start(call: CallbackQuery, state: FSMContext):
    feature = "review_batch" if call.data != "review_one" else "review"
    if not await access_or_paywall(call, feature): return
    await state.update_data(mode=call.data, charged_feature=feature)
    await state.set_state(ReviewFlow.waiting)
    prompt = "Вставьте отзыв и при необходимости контекст заказа." if call.data == "review_one" else "Вставьте отзывы/вопросы, каждый с новой строки. Можно до 18 000 символов."
    await edit_or_answer(call, prompt, back_kb("reviews"))


@router.message(ReviewFlow.waiting, F.text | F.voice | F.audio)
async def review_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        brand = await brand_context(message.from_user.id)
        if data["mode"] == "review_one":
            prompt = f"Отзыв: {text}\nBrand Kit: {brand}\nДай 3 ответа: короткий, тёплый, официальный. На негативе не признавай юридическую вину без фактов; предложи конкретный следующий шаг."
        elif data["mode"] == "review_faq":
            prompt = f"Вопросы покупателей:\n{text}\nСоздай FAQ, предложения для карточки и недостающие слайды инфографики."
        else:
            prompt = f"Отзывы:\n{text}\nКлассифицируй тональность и темы; посчитай частоты приблизительно; выдели новые/критические жалобы; предложи изменения товара, упаковки, карточки, FAQ и ТЗ поставщику; подготовь шаблоны ответов."
        result = await AI.text("Ты руководитель клиентского сервиса маркетплейса. Не спорь с покупателем и не обещай невозможное.", prompt, max_tokens=2600)
        hid = await save_history(message.from_user.id, data["mode"], text, result)
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "review"))
        await message.answer(f"❌ {exc}")
    await state.clear()


@router.callback_query(F.data == "appeals")
async def appeals_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear(); await state.set_state(AppealFlow.type)
    await edit_or_answer(call, "🛡 АПЕЛЛЯЦИИ PRO\nВыберите ситуацию:", kb([[("💰 Штраф", "atype_fine"), ("🚫 Блокировка", "atype_block")], [("📦 Поставка", "atype_supply"), ("↩️ Возврат/удержание", "atype_return")], [("⭐ Отзыв", "atype_review"), ("📝 Модерация", "atype_moderation")], [("🏠 Меню", "home")]]))


@router.callback_query(F.data.startswith("atype_"), AppealFlow.type)
async def appeal_type(call: CallbackQuery, state: FSMContext):
    if not await access_or_paywall(call, "appeal"): return
    await state.update_data(appeal_type=call.data[6:], charged_feature="appeal")
    await state.set_state(AppealFlow.waiting)
    await edit_or_answer(call, "Опишите площадку, дату, артикул/поставку, сумму, причину, хронологию, доказательства и желаемое решение. Не отправляйте паспортные данные.", back_kb("appeals"))


@router.message(AppealFlow.waiting, F.text | F.voice | F.audio)
async def appeal_process(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        text = await get_message_input(message, message.bot)
        result = await AI.text("Ты специалист по обращениям в поддержку маркетплейсов, не адвокат. Не выдумывай нормы и пункты правил. Отмечай необходимость сверить актуальную оферту.", f"""Тип: {data['appeal_type']}
Ситуация: {text}

Подготовь: проверку достаточности данных; краткую и полную апелляцию; хронологию;
аргументы только из фактов; список приложений; формулировку требуемого решения;
следующие шаги при отказе; предупреждения о неподтверждённых утверждениях.""", max_tokens=2300)
        hid = await save_history(message.from_user.id, "appeal", text, result, metadata={"type": data["appeal_type"]})
        await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(message.from_user.id, data.get("charged_feature", "appeal"))
        await message.answer(f"❌ {exc}")
    await state.clear()


# --- Connections and AI director --------------------------------------------
@router.callback_query(F.data == "connections")
async def connections_menu(call: CallbackQuery, state: FSMContext):
    await refund_pending_state(call.from_user.id, state)
    await state.clear()
    rows = await DB.fetchall(
        "SELECT id,marketplace,name,status,last_sync_at,last_error FROM marketplace_connections WHERE user_id=? ORDER BY id",
        (call.from_user.id,),
    )
    text = "🔌 ПОДКЛЮЧЁННЫЕ КАБИНЕТЫ\n\n"
    buttons: list[list[tuple[str, str]]] = []
    if rows:
        for row in rows:
            icon = "✅" if row["status"] == "active" else "⚠️" if row["status"] == "error" else "⏳"
            synced = parse_dt(row.get("last_sync_at"))
            text += f"{icon} {row['name']} · {row['marketplace'].upper()} · синхр. {synced.strftime('%d.%m %H:%M') if synced else 'не было'}\n"
            buttons.append([(f"{icon} {row['name']} · {row['marketplace'].upper()}", f"conn_{row['id']}")])
    else:
        text += "Подключений нет."
    if not CIPHER.secure:
        text += "\n\n⚠️ Без ENCRYPTION_KEY новые кабинеты не сохраняются — это защита токенов."
    buttons += [
        [("➕ WB", "connect_wb"), ("➕ Ozon", "connect_ozon")],
        [("🔄 Синхронизировать все", "connections_check"), ("📥 Импорт отчёта", "import_report")],
        [("🏠 Меню", "home")],
    ]
    await edit_or_answer(call, text, kb(buttons))


async def _shop_limit_check(user_id: int) -> tuple[bool, str]:
    plan, _ = await user_plan(user_id)
    allowed = int(PLANS[plan]["shops"])
    if allowed <= 0:
        return False, "Подключение кабинета доступно на ПРО и Бизнес"
    count = await DB.fetchone("SELECT COUNT(*) c FROM marketplace_connections WHERE user_id=?", (user_id,))
    if int((count or {}).get("c", 0)) >= allowed:
        return False, f"На тарифе {PLANS[plan]['name']} доступно кабинетов: {allowed}"
    if not CIPHER.secure:
        return False, "На сервере не настроен ENCRYPTION_KEY — токен нельзя сохранить безопасно"
    return True, ""


@router.callback_query(F.data.regexp(r"^conn_\d+$"))
async def connection_view(call: CallbackQuery):
    try:
        cid = int(call.data.split("_")[1])
    except Exception:
        return
    row = await DB.fetchone("SELECT * FROM marketplace_connections WHERE id=? AND user_id=?", (cid, call.from_user.id))
    if not row:
        await call.answer("Подключение не найдено", show_alert=True)
        return
    last = parse_dt(row.get("last_sync_at"))
    text = (
        f"🔌 {row['name']}\n\n"
        f"Площадка: {row['marketplace'].upper()}\n"
        f"Статус: {row['status']}\n"
        f"Последняя синхронизация: {last.strftime('%d.%m.%Y %H:%M') if last else '—'}\n"
        f"Последняя ошибка: {row.get('last_error') or '—'}"
    )
    await edit_or_answer(
        call,
        text,
        kb([[("🔄 Синхронизировать", f"conn_sync_{cid}"), ("🗑 Отключить", f"conn_delete_{cid}")], [("⬅️ Кабинеты", "connections")]]),
    )


@router.callback_query(F.data == "connect_wb")
async def connect_wb(call: CallbackQuery, state: FSMContext):
    ok, reason = await _shop_limit_check(call.from_user.id)
    if not ok:
        await call.answer(reason, show_alert=True)
        return
    await state.set_state(ConnectFlow.waiting_wb)
    await edit_or_answer(call, "Отправьте одной строкой: название кабинета; WB API token\n\nТокен будет проверен и сохранён только в зашифрованном виде.", back_kb("connections"))


@router.message(ConnectFlow.waiting_wb, F.text)
async def connect_wb_save(message: Message, state: FSMContext):
    try:
        name, token = [x.strip() for x in message.text.split(";", 1)]
        if not name or len(token) < 10:
            raise ValueError("Проверьте название и токен")
        await MP.wb_validate(token)
        await DB.execute(
            """INSERT INTO marketplace_connections(user_id,marketplace,name,token_encrypted,status,created_at)
            VALUES(?,?,?,?,?,?) ON CONFLICT(user_id,marketplace,name) DO UPDATE SET token_encrypted=excluded.token_encrypted,status='active',last_error=''""",
            (message.from_user.id, "wb", name[:120], CIPHER.encrypt(token), "active", iso_now()),
        )
        row = await DB.fetchone("SELECT * FROM marketplace_connections WHERE user_id=? AND marketplace='wb' AND name=?", (message.from_user.id, name[:120]))
        sync = await sync_connection_products(row) if row else {"ok": True, "synced": 0}
        await message.answer(
            f"✅ WB-кабинет «{name}» подключён.\nТоваров: {sync.get('synced', 0)} · Снимков метрик: {sync.get('metrics', 0)}.",
            reply_markup=back_kb("connections"),
        )
    except Exception as exc:
        await message.answer(f"❌ Не удалось подключить WB: {exc}", reply_markup=back_kb("connections"))
    await state.clear()


@router.callback_query(F.data == "connect_ozon")
async def connect_ozon(call: CallbackQuery, state: FSMContext):
    ok, reason = await _shop_limit_check(call.from_user.id)
    if not ok:
        await call.answer(reason, show_alert=True)
        return
    await state.set_state(ConnectFlow.waiting_ozon)
    await edit_or_answer(call, "Отправьте: название кабинета; Client-Id; Api-Key\n\nДанные будут проверены и сохранены в зашифрованном виде.", back_kb("connections"))


@router.message(ConnectFlow.waiting_ozon, F.text)
async def connect_ozon_save(message: Message, state: FSMContext):
    try:
        name, client_id, token = [x.strip() for x in message.text.split(";", 2)]
        if not name or not client_id or len(token) < 10:
            raise ValueError("Проверьте название, Client-Id и Api-Key")
        await MP.ozon_validate(client_id, token)
        await DB.execute(
            """INSERT INTO marketplace_connections(user_id,marketplace,name,client_id,token_encrypted,status,created_at)
            VALUES(?,?,?,?,?,?,?) ON CONFLICT(user_id,marketplace,name) DO UPDATE SET client_id=excluded.client_id,token_encrypted=excluded.token_encrypted,status='active',last_error=''""",
            (message.from_user.id, "ozon", name[:120], client_id[:120], CIPHER.encrypt(token), "active", iso_now()),
        )
        row = await DB.fetchone("SELECT * FROM marketplace_connections WHERE user_id=? AND marketplace='ozon' AND name=?", (message.from_user.id, name[:120]))
        sync = await sync_connection_products(row) if row else {"ok": True, "synced": 0}
        await message.answer(
            f"✅ Ozon-кабинет «{name}» подключён.\nСинхронизировано товаров: {sync.get('synced', 0)}.",
            reply_markup=back_kb("connections"),
        )
    except Exception as exc:
        await message.answer(f"❌ Не удалось подключить Ozon: {exc}", reply_markup=back_kb("connections"))
    await state.clear()


@router.callback_query(F.data.startswith("conn_sync_"))
async def connection_sync_one(call: CallbackQuery):
    cid = int(call.data.rsplit("_", 1)[1])
    row = await DB.fetchone("SELECT * FROM marketplace_connections WHERE id=? AND user_id=?", (cid, call.from_user.id))
    if not row:
        await call.answer("Подключение не найдено", show_alert=True)
        return
    await call.answer("Синхронизация запущена")
    result = await sync_connection_products(row)
    if result.get("ok"):
        await call.message.answer(f"✅ Получено: {result['received']}; обновлено товаров: {result['synced']}; снимков метрик: {result.get('metrics', 0)}.", reply_markup=back_kb("connections"))
    else:
        await call.message.answer(f"❌ Ошибка синхронизации: {result.get('error')}", reply_markup=back_kb("connections"))


@router.callback_query(F.data.startswith("conn_delete_"))
async def connection_delete(call: CallbackQuery):
    cid = int(call.data.rsplit("_", 1)[1])
    await DB.execute("DELETE FROM marketplace_connections WHERE id=? AND user_id=?", (cid, call.from_user.id))
    await call.answer("Кабинет отключён")
    await edit_or_answer(call, "✅ Подключение удалено. Сохранённые товары и история не удалены.", back_kb("connections"))


@router.callback_query(F.data == "connections_check")
async def connections_check(call: CallbackQuery):
    rows = await DB.fetchall("SELECT * FROM marketplace_connections WHERE user_id=?", (call.from_user.id,))
    if not rows:
        await call.answer("Нет подключений", show_alert=True)
        return
    await call.answer("Синхронизация запущена")
    results = []
    for row in rows:
        result = await sync_connection_products(row)
        if result.get("ok"):
            results.append(f"✅ {row['name']}: товаров {result['synced']}, метрик {result.get('metrics', 0)}")
        else:
            results.append(f"❌ {row['name']}: {result.get('error')}")
    await build_recommendations_for_user(call.from_user.id)
    await call.message.answer("\n".join(results), reply_markup=back_kb("connections"))


DIRECTOR_ACTION_TARGETS = {
    "open_import": "import_report",
    "open_connections": "connections",
    "open_supply": "supply",
    "open_audit": "analytics",
    "open_visual_audit": "audit_photo",
    "open_reviews": "reviews",
    "open_finance": "finance",
    "open_card360": "card360",
}


def _health_bar(score: int) -> str:
    filled = int(clamp(score, 0, 100) // 10)
    return "●" * filled + "○" * (10 - filled)


async def _director_dashboard(user_id: int) -> tuple[str, InlineKeyboardMarkup]:
    recs = await build_recommendations_for_user(user_id)
    health = await health_report(user_id, recs)
    sections = health["sections"]
    lines = [
        "🤖 AI-ДИРЕКТОР МАГАЗИНА",
        "",
        f"Индекс здоровья: {health['score']}/100  {_health_bar(health['score'])}",
        f"Полнота данных: {health['data_completeness']}%",
        "",
        "СОСТОЯНИЕ",
        f"• Остатки: {sections.get('stock', 100)}/100",
        f"• Прибыль: {sections.get('profit', 100)}/100",
        f"• Карточки: {sections.get('card', 100)}/100",
        f"• Покупатели: {sections.get('customers', 100)}/100",
        "",
    ]
    buttons: list[list[tuple[str, str]]] = []
    if recs:
        lines.append("ЗАДАЧИ ПО ПРИОРИТЕТУ")
        for index, rec in enumerate(recs[:6], 1):
            icon = {"critical": "🔴", "high": "🟠", "medium": "🟡", "low": "🔵"}.get(rec["severity"], "•")
            effect = f" · до {money(rec['estimated_effect'])}" if rec.get("estimated_effect", 0) > 0 else ""
            lines.append(f"{index}. {icon} {rec['title']}{effect}")
            buttons.append([(f"{icon} {index}. {rec['title'][:32]}", f"rec_{rec['id']}")])
    else:
        lines.append("✅ По доступным данным критических задач нет.")
    if health["data_completeness"] < 70:
        lines.extend(["", "⚠️ Для точных денежных рекомендаций подключите кабинет или загрузите отчёт."])
    buttons += [
        [("🧠 Глубокий AI-разбор", "director_deep"), ("📥 Загрузить отчёт", "import_report")],
        [("📜 Журнал решений", "decisions"), ("🔄 Обновить", "director")],
        [("🏠 Меню", "home")],
    ]
    return "\n".join(lines), kb(buttons)


@router.callback_query(F.data == "director")
async def director(call: CallbackQuery):
    text, markup = await _director_dashboard(call.from_user.id)
    await edit_or_answer(call, text, markup)


@router.callback_query(F.data.regexp(r"^rec_\d+$"))
async def recommendation_view(call: CallbackQuery):
    try:
        rid = int(call.data.split("_", 1)[1])
    except Exception:
        return
    rec = await DB.fetchone("SELECT * FROM recommendations WHERE id=? AND user_id=?", (rid, call.from_user.id))
    if not rec:
        await call.answer("Задача не найдена", show_alert=True)
        return
    evidence = {}
    try:
        evidence = json.loads(rec.get("evidence_json") or "{}")
    except Exception:
        pass
    fact_lines = []
    for key, value in list(evidence.items())[:8]:
        if isinstance(value, float):
            value = round(value, 2)
        fact_lines.append(f"• {key}: {value}")
    effect = money(rec["estimated_effect"]) if rec.get("estimated_effect", 0) > 0 else "не рассчитан"
    text = (
        f"{ {'critical':'🔴','high':'🟠','medium':'🟡','low':'🔵'}.get(rec['severity'],'•') } {rec['title']}\n\n"
        f"{rec['description']}\n\n"
        f"Категория: {rec.get('category') or 'general'}\n"
        f"Уверенность: {rec.get('confidence', 50)}%\n"
        f"Оценочный эффект: {effect}"
    )
    if fact_lines:
        text += "\n\nОСНОВАНИЕ\n" + "\n".join(fact_lines)
    target = DIRECTOR_ACTION_TARGETS.get(rec.get("action_code") or "", "director")
    await edit_or_answer(
        call,
        text,
        kb([
            [("▶️ Перейти к действию", target), ("📌 В работу", f"rec_plan_{rid}")],
            [("✅ Выполнено", f"rec_done_{rid}"), ("⏰ Отложить 3 дня", f"rec_snooze_{rid}")],
            [("⬅️ AI-директор", "director")],
        ]),
    )


async def _log_decision(user_id: int, recommendation_id: int, action: str, status: str, note: str = "") -> None:
    rec = await DB.fetchone("SELECT evidence_json FROM recommendations WHERE id=? AND user_id=?", (recommendation_id, user_id))
    await DB.execute(
        """INSERT INTO decision_log(user_id,recommendation_id,action,status,note,baseline_json,result_json,created_at,updated_at)
           VALUES(?,?,?,?,?,?,?, ?,?)""",
        (user_id, recommendation_id, action, status, note, (rec or {}).get("evidence_json", "{}"), "{}", iso_now(), iso_now()),
    )


@router.callback_query(F.data.startswith("rec_plan_"))
async def recommendation_plan(call: CallbackQuery):
    rid = int(call.data.rsplit("_", 1)[1])
    await DB.execute("UPDATE recommendations SET status='in_progress' WHERE id=? AND user_id=?", (rid, call.from_user.id))
    await _log_decision(call.from_user.id, rid, "accepted", "in_progress")
    await call.answer("Задача добавлена в работу")
    text, markup = await _director_dashboard(call.from_user.id)
    await call.message.edit_text(text, reply_markup=markup)


@router.callback_query(F.data.startswith("rec_done_"))
async def recommendation_done(call: CallbackQuery):
    rid = int(call.data.rsplit("_", 1)[1])
    await DB.execute("UPDATE recommendations SET status='resolved',resolved_at=? WHERE id=? AND user_id=?", (iso_now(), rid, call.from_user.id))
    await _log_decision(call.from_user.id, rid, "completed", "resolved")
    await call.answer("Решение отмечено выполненным")
    text, markup = await _director_dashboard(call.from_user.id)
    await call.message.edit_text(text, reply_markup=markup)


@router.callback_query(F.data.startswith("rec_snooze_"))
async def recommendation_snooze(call: CallbackQuery):
    rid = int(call.data.rsplit("_", 1)[1])
    until = now() + timedelta(days=3)
    await DB.execute("UPDATE recommendations SET status='snoozed',resolved_at=? WHERE id=? AND user_id=?", (until.isoformat(), rid, call.from_user.id))
    await _log_decision(call.from_user.id, rid, "snoozed", "snoozed", f"до {until.strftime('%d.%m.%Y')}")
    await call.answer("Отложено на 3 дня")
    text, markup = await _director_dashboard(call.from_user.id)
    await call.message.edit_text(text, reply_markup=markup)


@router.callback_query(F.data == "decisions")
async def decisions_list(call: CallbackQuery):
    rows = await DB.fetchall(
        """SELECT d.*,r.title FROM decision_log d LEFT JOIN recommendations r ON r.id=d.recommendation_id
           WHERE d.user_id=? ORDER BY d.id DESC LIMIT 20""",
        (call.from_user.id,),
    )
    if not rows:
        await edit_or_answer(call, "📜 ЖУРНАЛ РЕШЕНИЙ\n\nПока пуст. Примите первую рекомендацию AI-директора.", back_kb("director"))
        return
    lines = ["📜 ЖУРНАЛ РЕШЕНИЙ", ""]
    for row in rows:
        dt = parse_dt(row["created_at"])
        lines.append(f"• {dt.strftime('%d.%m') if dt else '—'} · {row['status']} · {row.get('title') or row['action']}")
    await edit_or_answer(call, "\n".join(lines), back_kb("director"))


async def _director_factual_payload(user_id: int) -> dict[str, Any]:
    products = await DB.fetchall(
        "SELECT id,name,price,cost,article,marketplace,category FROM products WHERE user_id=? AND status='active' LIMIT 100",
        (user_id,),
    )
    snapshots = await DB.fetchall(
        "SELECT * FROM metric_snapshots WHERE user_id=? ORDER BY metric_date DESC,id DESC LIMIT 250",
        (user_id,),
    )
    recs = await DB.fetchall(
        "SELECT severity,title,description,estimated_effect,confidence,evidence_json FROM recommendations WHERE user_id=? AND status IN ('new','in_progress') ORDER BY id DESC LIMIT 20",
        (user_id,),
    )
    health = await health_report(user_id, recs)
    return {"health": health, "products": products, "snapshots": snapshots, "recommendations": recs}


@router.callback_query(F.data == "director_deep")
async def director_deep(call: CallbackQuery):
    if not await access_or_paywall(call, "director"):
        return
    try:
        factual = await _director_factual_payload(call.from_user.id)
        result = await AI.text(
            AUDIT_SYSTEM,
            """Ты AI-директор магазина. Используй только переданные факты и детерминированные рекомендации.
Сформируй управленческий отчёт: 1) что известно точно; 2) три главных денежно значимых проблемы;
3) план на сегодня; 4) план на 7 дней; 5) что измерить после изменений; 6) какие данные отсутствуют.
Не пересчитывай индекс здоровья самостоятельно и не придумывай продажи, CTR, остатки или денежный эффект.
Данные:
""" + json.dumps(factual, ensure_ascii=False, default=str)[:CFG.max_input_chars],
            max_tokens=3000,
        )
        hid = await save_history(call.from_user.id, "ai_director_deep", "connected data", result, metadata={"health": factual["health"]["score"]})
        await call.answer()
        await answer_long(call.message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("📜 Журнал решений", "decisions")], [("⬅️ AI-директор", "director")]]))
    except Exception as exc:
        await refund(call.from_user.id, "director")
        await call.message.answer(f"❌ Глубокий разбор недоступен: {exc}", reply_markup=back_kb("director"))
        await call.answer()


# --- Profile, history, export, tariffs --------------------------------------
@router.callback_query(F.data == "profile")
async def profile(call: CallbackQuery):
    plan, expiry = await user_plan(call.from_user.id)
    bal = await balance(call.from_user.id)
    user = await DB.fetchone("SELECT * FROM users WHERE user_id=?", (call.from_user.id,))
    refs = await DB.fetchone("SELECT COUNT(*) AS c FROM referral_events WHERE inviter_user_id=?", (call.from_user.id,))
    health = await health_report(call.from_user.id)
    text = f"""👤 ПРОФИЛЬ

Тариф: {PLANS[plan]['name']}
Действует до: {expiry.strftime('%d.%m.%Y') if expiry else '—'}
Кредиты: {bal.get('credits',0)}
Изображения: {bal.get('image_credits',0)}
Всего операций: {bal.get('lifetime_used',0)}
Приглашено: {(refs or {}).get('c',0)}
Здоровье магазина: {health['score']}/100
Полнота данных: {health['data_completeness']}%
Ежедневный отчёт: {'включён' if user and user['notifications_enabled'] else 'выключен'} · {int((user or {}).get('daily_report_hour', 9)):02d}:00"""
    toggle = "🔕 Выключить отчёт" if user and user["notifications_enabled"] else "🔔 Включить отчёт"
    await edit_or_answer(
        call,
        text,
        kb([
            [("📂 История", "history"), ("📜 Журнал решений", "decisions")],
            [(toggle, "notifications_toggle"), ("🕘 Время отчёта", "report_time")],
            [("👥 Реферальная ссылка", "referral"), ("🎁 Промокод", "promo")],
            [("💳 Тарифы", "tariffs"), ("🏠 Меню", "home")],
        ]),
    )


@router.callback_query(F.data == "notifications_toggle")
async def notifications_toggle(call: CallbackQuery):
    user = await DB.fetchone("SELECT notifications_enabled FROM users WHERE user_id=?", (call.from_user.id,))
    enabled = 0 if user and user["notifications_enabled"] else 1
    await DB.execute("UPDATE users SET notifications_enabled=? WHERE user_id=?", (enabled, call.from_user.id))
    await profile(call)


@router.callback_query(F.data == "report_time")
async def report_time(call: CallbackQuery):
    await edit_or_answer(
        call,
        "Во сколько присылать ежедневный отчёт по часовому поясу бота?",
        kb([
            [("08:00", "rtime_8"), ("09:00", "rtime_9"), ("10:00", "rtime_10")],
            [("18:00", "rtime_18"), ("20:00", "rtime_20")],
            [("⬅️ Профиль", "profile")],
        ]),
    )


@router.callback_query(F.data.startswith("rtime_"))
async def report_time_save(call: CallbackQuery):
    try:
        hour = int(call.data.split("_")[1])
    except Exception:
        return
    await DB.execute("UPDATE users SET daily_report_hour=? WHERE user_id=?", (hour, call.from_user.id))
    await profile(call)


@router.callback_query(F.data == "history")
async def history_list(call: CallbackQuery):
    rows = await DB.fetchall("SELECT id,feature,created_at FROM history WHERE user_id=? ORDER BY created_at DESC LIMIT 15", (call.from_user.id,))
    buttons = [[(f"{r['feature']} · {parse_dt(r['created_at']).strftime('%d.%m %H:%M')}", f"history_{r['id']}")] for r in rows]
    buttons.append([("⬅️ Профиль", "profile"), ("🏠 Меню", "home")])
    await edit_or_answer(call, "📂 ИСТОРИЯ\n\n" + ("Выберите результат:" if rows else "История пока пуста."), kb(buttons))


@router.callback_query(F.data.startswith("history_"))
async def history_view(call: CallbackQuery):
    hid = int(call.data.split("_")[1])
    row = await DB.fetchone("SELECT * FROM history WHERE id=? AND user_id=?", (hid, call.from_user.id))
    if not row:
        await call.answer("Не найдено", show_alert=True); return
    await call.answer()
    await answer_long(call.message, row["result"], kb([[("📄 Экспорт", f"export_{hid}"), ("⬅️ История", "history")]]))


@router.callback_query(F.data.startswith("cardsec_"))
async def card_section_view(call: CallbackQuery):
    parts = call.data.split("_")
    if len(parts) < 3:
        return
    code = parts[1]
    try:
        hid = int(parts[2])
    except ValueError:
        return
    row = await DB.fetchone("SELECT metadata FROM history WHERE id=? AND user_id=?", (hid, call.from_user.id))
    if not row:
        await call.answer("Результат не найден", show_alert=True)
        return
    try:
        package = (json.loads(row["metadata"] or "{}") or {}).get("card_package") or {}
    except Exception:
        package = {}
    if not package:
        await call.answer("Структурированные разделы недоступны", show_alert=True)
        return
    await call.answer()
    await answer_long(call.message, render_card_section(package, code), card_result_kb(hid))


@router.callback_query(F.data.startswith("export_"))
async def export_history(call: CallbackQuery):
    hid = int(call.data.split("_")[1])
    row = await DB.fetchone("SELECT * FROM history WHERE id=? AND user_id=?", (hid, call.from_user.id))
    if not row:
        await call.answer("Не найдено", show_alert=True); return
    content = f"МАРКЕТПРО PREMIUM\nФункция: {row['feature']}\nДата: {row['created_at']}\n\n{row['result']}"
    await call.message.answer_document(BufferedInputFile(content.encode("utf-8"), filename=f"marketpro_{hid}.txt"), caption="Экспорт результата")
    await call.answer()


@router.callback_query(F.data == "referral")
async def referral(call: CallbackQuery):
    link = f"https://t.me/{CFG.bot_username}?start=ref_{call.from_user.id}"
    await edit_or_answer(call, f"👥 РЕФЕРАЛЬНАЯ ПРОГРАММА\n\nВаша ссылка:\n{link}\n\nВы получаете +{CFG.referral_inviter} кредитов, друг +{CFG.referral_invited}. Повторное начисление защищено уникальным событием регистрации.", back_kb("profile"))


@router.callback_query(F.data == "promo")
async def promo_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(PromoCodeFlow.waiting)
    await state.update_data(promo_mode=True)
    await edit_or_answer(call, "Введите промокод:", back_kb("profile"))


@router.message(PromoCodeFlow.waiting, F.text)
async def promo_use(message: Message, state: FSMContext):
    data = await state.get_data()
    if not data.get("promo_mode"):
        return
    code = message.text.strip().upper()
    row = await DB.fetchone("SELECT * FROM promo_codes WHERE code=? AND active=1", (code,))
    if not row:
        await message.answer("❌ Промокод не найден.", reply_markup=back_kb("profile")); await state.clear(); return
    expiry = parse_dt(row["expires_at"])
    if expiry and expiry < now() or row["max_uses"] and row["uses"] >= row["max_uses"]:
        await message.answer("❌ Промокод недействителен."); await state.clear(); return
    try:
        await DB.transaction([
            ("INSERT INTO promo_uses(user_id,code,used_at) VALUES(?,?,?)", (message.from_user.id, code, iso_now())),
            ("UPDATE promo_codes SET uses=uses+1 WHERE code=?", (code,)),
            ("UPDATE balances SET credits=credits+? WHERE user_id=?", (row["bonus_credits"], message.from_user.id)),
        ])
        await message.answer(f"✅ Начислено +{row['bonus_credits']} кредитов.", reply_markup=back_kb("profile"))
    except sqlite3.IntegrityError:
        await message.answer("Этот промокод уже использован.")
    await state.clear()


@router.callback_query(F.data == "tariffs")
async def tariffs(call: CallbackQuery):
    text = "💳 ТАРИФЫ\n\n"
    for key in ("start", "pro", "business"):
        p = PLANS[key]
        text += f"{p['name']} — {p['price']} ₽ / 30 дней\n{p['credits']} кредитов · {p['images']} изображений · {p['products']} товаров · {p['shops']} кабинетов\n\n"
    await edit_or_answer(call, text + "Кредиты используются вместо опасного «безлимита» и защищают экономику сервиса.", kb([[("490 ₽ Старт", "pay_start"), ("990 ₽ ПРО", "pay_pro")], [("1990 ₽ Бизнес", "pay_business")], [("🏠 Меню", "home")]]))


@router.callback_query(F.data.startswith("pay_"))
async def create_payment(call: CallbackQuery):
    plan = call.data[4:]
    if plan not in PLANS or plan == "free":
        return
    if not Payment or not CFG.yookassa_shop_id or not CFG.yookassa_secret:
        await call.answer("YooKassa не настроена", show_alert=True); return
    try:
        p = await asyncio.to_thread(Payment.create, {
            "amount": {"value": f"{PLANS[plan]['price']}.00", "currency": "RUB"},
            "confirmation": {"type": "redirect", "return_url": CFG.payment_return_url},
            "capture": True,
            "description": f"МаркетПРО {PLANS[plan]['name']} — 30 дней",
            "metadata": {"user_id": str(call.from_user.id), "plan": plan},
        }, str(uuid.uuid4()))
        await DB.execute("INSERT INTO payments(payment_id,user_id,plan,amount,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
                         (p.id, call.from_user.id, plan, PLANS[plan]["price"], "pending", iso_now(), iso_now()))
        await call.message.answer(f"Тариф {PLANS[plan]['name']} — {PLANS[plan]['price']} ₽", reply_markup=kb([[("💳 Оплатить", f"url:{p.confirmation.confirmation_url}")], [("🏠 Меню", "home")]]))
        await call.answer()
    except Exception as exc:
        log.exception("payment create")
        await call.answer("Ошибка создания платежа", show_alert=True)
        await call.message.answer(f"Напишите в поддержку: {CFG.support_url}\nТехническая информация: {str(exc)[:200]}")


# --- Admin ------------------------------------------------------------------
@router.message(Command("stats"))
async def admin_stats(message: Message):
    if message.from_user.id != CFG.owner_id: return
    users = await DB.fetchone("SELECT COUNT(*) c FROM users")
    paid = await DB.fetchone("SELECT COUNT(*) c FROM subscriptions WHERE plan!='free' AND expires_at>?", (iso_now(),))
    revenue = await DB.fetchone("SELECT COALESCE(SUM(amount),0) s FROM payments WHERE status='succeeded'")
    operations = await DB.fetchone("SELECT COUNT(*) c FROM history")
    await message.answer(f"📊 МаркетПРО\nПользователи: {users['c']}\nАктивные подписки: {paid['c']}\nУспешные платежи: {money(revenue['s'])}\nРезультаты: {operations['c']}")


@router.message(Command("give"))
async def admin_give(message: Message):
    if message.from_user.id != CFG.owner_id: return
    parts = message.text.split()
    try:
        uid = int(parts[1]); kind = parts[2]
        if kind in PLANS and kind != "free":
            await ensure_user(uid); await activate_plan(uid, kind); await message.answer("✅ Тариф выдан")
        elif kind == "credits":
            amount = int(parts[3]); await DB.execute("UPDATE balances SET credits=credits+? WHERE user_id=?", (amount, uid)); await message.answer("✅ Кредиты начислены")
        else: raise ValueError
    except Exception:
        await message.answer("Формат: /give USER_ID pro | /give USER_ID credits 100")


@router.message(Command("promo_create"))
async def admin_promo_create(message: Message):
    if message.from_user.id != CFG.owner_id: return
    try:
        _, code, credits, max_uses = message.text.split()
        await DB.execute("INSERT OR REPLACE INTO promo_codes(code,bonus_credits,max_uses,active) VALUES(?,?,?,1)", (code.upper(), int(credits), int(max_uses)))
        await message.answer("✅ Промокод создан")
    except Exception:
        await message.answer("Формат: /promo_create CODE CREDITS MAX_USES")


@router.message(Command("broadcast"))
async def broadcast_start(message: Message, state: FSMContext):
    if message.from_user.id != CFG.owner_id: return
    await state.set_state(BroadcastFlow.waiting)
    await message.answer("Отправьте текст рассылки. Для отмены /cancel")


@router.message(BroadcastFlow.waiting, F.text)
async def broadcast_process(message: Message, state: FSMContext):
    if message.from_user.id != CFG.owner_id: return
    users = await DB.fetchall("SELECT user_id FROM users WHERE is_blocked=0")
    ok = fail = 0
    for item in users:
        try:
            await message.bot.send_message(item["user_id"], message.text)
            ok += 1
        except (TelegramForbiddenError, TelegramBadRequest):
            fail += 1
        except Exception:
            fail += 1
        await asyncio.sleep(0.04)
    await state.clear(); await message.answer(f"Рассылка: {ok} успешно, {fail} ошибок")


# Быстрый вход из главного меню ------------------------------------------------
def looks_like_marketplace_url(value: str) -> bool:
    low = value.lower()
    return ("wildberries.ru" in low or "wb.ru" in low or "ozon.ru" in low or "avito.ru" in low) and ("http://" in low or "https://" in low)


@router.callback_query(F.data.startswith("quickplat_"), QuickIntentFlow.choose_platform)
async def quick_platform_chosen(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    text_value = (data.get("quick_text") or "").strip()
    platform = call.data.replace("quickplat_", "")
    if not text_value:
        await state.clear()
        await call.answer("Описание товара потеряно. Отправьте его ещё раз.", show_alert=True)
        return
    if not await access_or_paywall(call, "card360_text"):
        await state.clear()
        return
    await call.answer()
    progress = await call.message.answer("⏳ Создаю карточку 360°...")
    try:
        package = await generate_card360_package(call.from_user.id, platform, text_value)
        full_result = render_card_full(package)
        hid = await save_history(
            call.from_user.id, "card360_quick", text_value, full_result,
            metadata={"platform": platform, "card_package": package},
        )
        try:
            await progress.delete()
        except Exception:
            pass
        await call.message.answer(render_card_summary(package), reply_markup=card_result_kb(hid))
    except Exception as exc:
        await refund(call.from_user.id, "card360_text")
        log.exception("quick card360")
        await call.message.answer(f"❌ Не удалось создать карточку: {exc}", reply_markup=back_kb("card360"))
    await state.clear()


@router.callback_query(F.data == "quick_photo_card", QuickIntentFlow.choose_photo_action)
async def quick_photo_card(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    file_id = data.get("quick_photo_file_id")
    if not file_id:
        await state.clear()
        await call.answer("Фото не найдено. Отправьте его ещё раз.", show_alert=True)
        return
    await call.answer()
    try:
        file = await call.bot.get_file(file_id)
        bio = await call.bot.download_file(file.file_path)
        image = bio.read()
        recognized = await AI.vision(
            "Ты товаровед. Определи только видимые характеристики товара. Не угадывай бренд, материал и комплектность. Если это не товар, ответь НЕ_ТОВАР.",
            "Опиши товар максимально точно для последующего создания карточки маркетплейса. Отдельно перечисли неизвестные данные.",
            image,
            "image/jpeg",
        )
        if "НЕ_ТОВАР" in recognized.upper():
            raise ValueError("На фото не удалось распознать товар")
        await state.set_state(QuickIntentFlow.choose_platform)
        await state.update_data(quick_text=recognized)
        await call.message.answer(
            "Товар распознан. Выберите площадку:",
            reply_markup=kb([
                [("🟣 Wildberries", "quickplat_wb"), ("🔵 Ozon", "quickplat_ozon")],
                [("🟡 Авито", "quickplat_avito"), ("🏠 Меню", "home")],
            ]),
        )
    except Exception as exc:
        await state.clear()
        await call.message.answer(f"❌ {exc}", reply_markup=main_kb())


@router.callback_query(F.data == "quick_photo_audit", QuickIntentFlow.choose_photo_action)
async def quick_photo_audit(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    file_id = data.get("quick_photo_file_id")
    if not file_id:
        await state.clear()
        await call.answer("Фото не найдено. Отправьте его ещё раз.", show_alert=True)
        return
    if not await access_or_paywall(call, "audit"):
        await state.clear()
        return
    await call.answer()
    try:
        file = await call.bot.get_file(file_id)
        bio = await call.bot.download_file(file.file_path)
        result = await AI.vision(
            AUDIT_SYSTEM,
            "Разбери этот скриншот. Сначала перечисли, что точно видно; затем проблемы, риски, приоритетные действия и данные, которых не хватает.",
            bio.read(),
            "image/jpeg",
        )
        hid = await save_history(call.from_user.id, "audit_screenshot_quick", "screenshot", result)
        await answer_long(call.message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
    except Exception as exc:
        await refund(call.from_user.id, "audit")
        await call.message.answer(f"❌ {exc}", reply_markup=main_kb())
    await state.clear()


# Fallbacks
@router.message(F.text)
async def fallback_text(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:
        await message.answer(
            "Сейчас открыт другой шаг. Отправьте данные в запрошенном формате или нажмите /cancel, чтобы выйти без потери зарезервированных кредитов."
        )
        return
    text_value = (message.text or "").strip()
    if not text_value:
        return

    if looks_like_marketplace_url(text_value):
        if not await access_or_paywall(message, "audit_link"):
            return
        try:
            card = await MP.fetch_public_card(text_value)
            source = card_data_text(card)
            await message.answer("✅ Получены реальные публичные данные:\n\n" + source[:1500])
            result = await audit_content(source, "Метрики кабинета не переданы; выводы по воронке являются гипотезами.")
            hid = await save_history(message.from_user.id, "audit_link_quick", text_value, result, metadata=card)
            await answer_long(message, result, kb([[("📄 Экспорт", f"export_{hid}"), ("🏠 Меню", "home")]]))
        except Exception as exc:
            await refund(message.from_user.id, "audit_link")
            await message.answer(
                f"⚠️ Не удалось достоверно получить карточку: {exc}\n\nПришлите скриншот карточки — бот не станет выдумывать данные.",
                reply_markup=main_kb(),
            )
        return

    await state.set_state(QuickIntentFlow.choose_platform)
    await state.update_data(quick_text=text_value)
    await message.answer(
        "Похоже, вы хотите создать карточку товара. Выберите площадку:",
        reply_markup=kb([
            [("🟣 Wildberries", "quickplat_wb"), ("🔵 Ozon", "quickplat_ozon")],
            [("🟡 Авито", "quickplat_avito"), ("❌ Это не товар", "home")],
        ]),
    )


@router.message(F.photo)
async def fallback_photo(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:
        await message.answer(
            "На этом шаге ожидается другой формат. Нажмите /cancel, чтобы выйти, или вернитесь к инструкции текущего сценария."
        )
        return
    photo = message.photo[-1]
    await state.set_state(QuickIntentFlow.choose_photo_action)
    await state.update_data(quick_photo_file_id=photo.file_id)
    await message.answer(
        "Что сделать с этим изображением?",
        reply_markup=kb([
            [("🚀 Создать карточку", "quick_photo_card"), ("🔍 Провести аудит", "quick_photo_audit")],
            [("🏠 Меню", "home")],
        ]),
    )


@router.message()
async def fallback_other(message: Message, state: FSMContext):
    if await state.get_state() is not None:
        await message.answer(
            "Этот тип сообщения не подходит для текущего шага. Следуйте инструкции выше или нажмите /cancel."
        )
        return
    await message.answer(
        "Используйте меню: текст подходит для быстрой карточки товара, фото — для карточки или аудита, CSV/XLSX — через «Аудит и рост → Импорт отчёта».",
        reply_markup=main_kb(),
    )


# ============================================================================
# BACKGROUND TASKS
# ============================================================================

async def payment_watcher(bot: Bot):
    if not Payment or not CFG.yookassa_shop_id or not CFG.yookassa_secret:
        return
    while True:
        await asyncio.sleep(20)
        rows = await DB.fetchall("SELECT * FROM payments WHERE status='pending' ORDER BY created_at LIMIT 100")
        for row in rows:
            try:
                payment = await asyncio.to_thread(Payment.find_one, row["payment_id"])
                if payment.status == "succeeded":
                    claimed = await DB.execute_rowcount(
                        "UPDATE payments SET status='processing',updated_at=? WHERE payment_id=? AND status='pending'",
                        (iso_now(), row["payment_id"]),
                    )
                    if not claimed:
                        continue
                    try:
                        await activate_plan(row["user_id"], row["plan"])
                        await DB.execute(
                            "UPDATE payments SET status='succeeded',updated_at=? WHERE payment_id=?",
                            (iso_now(), row["payment_id"]),
                        )
                        ref = await DB.fetchone(
                            "SELECT * FROM referral_events WHERE invited_user_id=? AND paid_rewarded=0",
                            (row["user_id"],),
                        )
                        if ref:
                            ref_claimed = await DB.execute_rowcount(
                                "UPDATE referral_events SET paid_rewarded=1 WHERE id=? AND paid_rewarded=0",
                                (ref["id"],),
                            )
                            if ref_claimed:
                                await DB.execute(
                                    "UPDATE balances SET credits=credits+15 WHERE user_id=?",
                                    (ref["inviter_user_id"],),
                                )
                        await bot.send_message(
                            row["user_id"],
                            f"✅ Тариф {PLANS[row['plan']]['name']} активирован на 30 дней.",
                            reply_markup=main_kb(),
                        )
                    except Exception:
                        await DB.execute(
                            "UPDATE payments SET status='pending',updated_at=? WHERE payment_id=? AND status='processing'",
                            (iso_now(), row["payment_id"]),
                        )
                        raise
                elif payment.status == "canceled":
                    await DB.execute(
                        "UPDATE payments SET status='canceled',updated_at=? WHERE payment_id=? AND status='pending'",
                        (iso_now(), row["payment_id"]),
                    )
            except Exception as exc:
                log.error("Payment check %s: %s", row["payment_id"], exc)


async def connection_sync_worker():
    """Периодически проверяет подключения и обновляет паспорта товаров без частых запросов."""
    while True:
        await asyncio.sleep(30 * 60)
        rows = await DB.fetchall("SELECT * FROM marketplace_connections WHERE status IN ('active','error')")
        for row in rows:
            last = parse_dt(row.get("last_sync_at"))
            if last and last > now() - timedelta(hours=6):
                continue
            try:
                await sync_connection_products(row)
            except Exception as exc:
                log.error("Background sync %s: %s", row.get("id"), exc)
            await asyncio.sleep(1)


async def recommendation_refresh_worker():
    while True:
        await asyncio.sleep(2 * 60 * 60)
        users = await DB.fetchall("SELECT user_id FROM users WHERE is_blocked=0")
        for user in users:
            try:
                await build_recommendations_for_user(user["user_id"])
            except Exception as exc:
                log.error("Recommendation refresh %s: %s", user["user_id"], exc)


async def daily_report_worker(bot: Bot):
    sent_date: dict[int, str] = {}
    while True:
        await asyncio.sleep(60)
        current = now()
        users = await DB.fetchall(
            "SELECT user_id,daily_report_hour,notifications_enabled FROM users WHERE notifications_enabled=1 AND is_blocked=0"
        )
        for user in users:
            uid = user["user_id"]
            today = current.date().isoformat()
            if current.hour != user["daily_report_hour"] or sent_date.get(uid) == today:
                continue
            try:
                recommendations = await build_recommendations_for_user(uid)
                health = await health_report(uid, recommendations)
                top = recommendations[:5]
                text = (
                    f"🤖 ОТЧЁТ AI-ДИРЕКТОРА\n\n"
                    f"Здоровье магазина: {health['score']}/100\n"
                    f"Полнота данных: {health['data_completeness']}%\n"
                    f"Активных задач: {len(recommendations)}\n\n"
                )
                if top:
                    text += "ГЛАВНОЕ НА СЕГОДНЯ\n" + "\n".join(
                        f"{i}. { {'critical':'🔴','high':'🟠','medium':'🟡','low':'🔵'}.get(rec['severity'],'•') } {rec['title']}"
                        for i, rec in enumerate(top, 1)
                    )
                else:
                    text += "✅ По доступным данным критических задач нет."
                await bot.send_message(
                    uid,
                    text,
                    reply_markup=kb([[("🤖 Открыть задачи", "director"), ("📥 Загрузить данные", "import_report")]]),
                )
                sent_date[uid] = today
            except TelegramForbiddenError:
                await DB.execute("UPDATE users SET notifications_enabled=0 WHERE user_id=?", (uid,))
            except Exception as exc:
                log.error("Daily report %s: %s", uid, exc)


async def main() -> None:
    if not CFG.bot_token:
        raise RuntimeError("BOT_TOKEN не задан")
    await DB.init()
    migration = await DB.migrate_legacy(CFG.legacy_db_path)
    if any(migration.values()):
        log.info("Legacy migration: %s", migration)
    if CFG.redis_url and RedisStorage:
        storage = RedisStorage.from_url(CFG.redis_url)
        log.info("FSM: Redis")
    else:
        storage = MemoryStorage()
        log.warning("FSM: MemoryStorage. Для production задайте REDIS_URL.")
    # Глобальный HTML parse mode отключён: AI-тексты могут содержать символы < и >.
    bot = Bot(CFG.bot_token)
    dp = Dispatcher(storage=storage)
    dp.include_router(router)
    tasks = [
        asyncio.create_task(payment_watcher(bot), name="payment_watcher"),
        asyncio.create_task(daily_report_worker(bot), name="daily_report"),
        asyncio.create_task(connection_sync_worker(), name="connection_sync"),
        asyncio.create_task(recommendation_refresh_worker(), name="recommendations"),
    ]
    log.info("МаркетПРО Premium 10/10 запущен")
    try:
        await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
    finally:
        for task in tasks:
            task.cancel()
        await asyncio.gather(*tasks, return_exceptions=True)
        await bot.session.close()
        if hasattr(storage, "close"):
            try:
                await storage.close()
            except Exception:
                pass


if __name__ == "__main__":
    asyncio.run(main())
